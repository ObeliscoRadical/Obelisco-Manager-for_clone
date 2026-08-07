"""
Modulo de Custos / Despesas com extracao automatica de faturas via IA (Gemini 2.5 Pro).
Funcionalidades:
- Upload de factura (PDF/JPG/PNG)
- Extracao IA de NIF, fornecedor, valor, IVA, data, numero
- Classificacao por categoria e tipo (fixo/variavel/obra)
- Associacao a obra especifica (centro de custo)
- Dashboard mensal: total gasto, por categoria, por obra
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, Request
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone, timedelta
import uuid
import os
import json
import logging
import base64
from pathlib import Path
import re
from difflib import SequenceMatcher
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from pymongo.errors import DuplicateKeyError

from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType

logger = logging.getLogger(__name__)

EXPENSES_DIR = Path("/app/backend/uploads/expenses")
EXPENSES_DIR.mkdir(parents=True, exist_ok=True)
RECONCILIATION_REPORTS_DIR = Path("/app/backend/uploads/reconciliation_reports")
RECONCILIATION_REPORTS_DIR.mkdir(parents=True, exist_ok=True)

expenses_router = APIRouter(prefix="/api/expenses", tags=["expenses"])

CATEGORIES = [
    "Combustível",
    "Material",
    "Fornecedor",
    "Serviços",
    "Comunicações",
    "Rendas",
    "Seguros",
    "Contabilidade/Advogado",
    "Ferramentas",
    "Viatura",
    "Alimentação",
    "Imposto/Taxa",
    "Outros",
]


class ExpenseCreate(BaseModel):
    date: str              # YYYY-MM-DD
    supplier: str = ""
    nif: str = ""
    invoice_number: str = ""
    category: str = "Outros"
    type: str = "variavel"    # fixo, variavel, obra
    obra_id: Optional[str] = None
    obra_name: Optional[str] = None
    value_net: float = 0       # sem IVA
    vat_rate: float = 23       # taxa IVA %
    vat_amount: float = 0      # valor do IVA
    value_gross: float = 0     # com IVA
    payment_method: str = ""
    notes: str = ""
    invoice_file: Optional[str] = None    # filename
    source_kind: Optional[str] = None
    bank_txn_id: Optional[str] = None
    bank_description: Optional[str] = None
    bank_analysis_id: Optional[str] = None
    dedupe_exempt: bool = False
    dedupe_exception_reason: str = ""


class ExpenseUpdate(BaseModel):
    date: Optional[str] = None
    supplier: Optional[str] = None
    nif: Optional[str] = None
    invoice_number: Optional[str] = None
    category: Optional[str] = None
    type: Optional[str] = None
    obra_id: Optional[str] = None
    obra_name: Optional[str] = None
    value_net: Optional[float] = None
    vat_rate: Optional[float] = None
    vat_amount: Optional[float] = None
    value_gross: Optional[float] = None
    payment_method: Optional[str] = None
    notes: Optional[str] = None
    dedupe_exempt: Optional[bool] = None
    dedupe_exception_reason: Optional[str] = None


async def extract_invoice_data(file_path: Path, mime_type: str) -> dict:
    """Use Gemini to extract structured data from a PT invoice."""
    api_key = os.environ.get("EMERGENT_LLM_KEY", "")
    if not api_key:
        return {"error": "EMERGENT_LLM_KEY not set"}

    system = (
        "És um assistente especializado em extrair dados de faturas portuguesas. "
        "Responde APENAS com JSON válido, sem texto antes ou depois. "
        "Se um campo não for visível na fatura, devolve string vazia ou 0. "
        "Valores monetários em euros como número (não como string com €)."
    )

    prompt = (
        "Analisa esta fatura portuguesa e extrai os seguintes dados em JSON:\n"
        "{\n"
        '  "supplier": "nome completo do fornecedor/emitente",\n'
        '  "nif": "NIF do fornecedor (9 dígitos)",\n'
        '  "invoice_number": "número da fatura",\n'
        '  "date": "data da fatura YYYY-MM-DD",\n'
        '  "value_net": valor líquido sem IVA (number),\n'
        '  "vat_rate": taxa de IVA em % (6, 13 ou 23),\n'
        '  "vat_amount": valor do IVA (number),\n'
        '  "value_gross": valor total com IVA (number),\n'
        '  "category": "uma de: Combustível, Material Elétrico, Material de Obra, Fornecedor, Serviços, Comunicações, Rendas, Seguros, Contabilidade/Advogado, Ferramentas, Viatura, Alimentação, Imposto/Taxa, Outros",\n'
        '  "description": "breve descrição do que foi faturado"\n'
        "}\n\n"
        "Responde APENAS com o JSON, sem markdown ```."
    )

    try:
        chat = LlmChat(
            api_key=api_key,
            session_id=f"invoice-{uuid.uuid4().hex[:8]}",
            system_message=system,
        ).with_model("gemini", "gemini-3.1-pro-preview")

        attach = FileContentWithMimeType(file_path=str(file_path), mime_type=mime_type)
        response = await chat.send_message(UserMessage(text=prompt, file_contents=[attach]))

        # Clean possible markdown code fences
        text = response.strip()
        if text.startswith("```"):
            text = text.split("```", 2)[1]
            if text.startswith("json"):
                text = text[4:]
            text = text.strip()
        if text.endswith("```"):
            text = text[:-3].strip()

        data = json.loads(text)
        result = {
            "supplier": str(data.get("supplier", "")).strip(),
            "nif": str(data.get("nif", "")).strip().replace(" ", ""),
            "invoice_number": str(data.get("invoice_number", "")).strip(),
            "date": str(data.get("date", "")).strip()[:10],
            "value_net": float(data.get("value_net", 0) or 0),
            "vat_rate": float(data.get("vat_rate", 23) or 23),
            "vat_amount": float(data.get("vat_amount", 0) or 0),
            "value_gross": float(data.get("value_gross", 0) or 0),
            "category": str(data.get("category", "Outros")).strip() or "Outros",
            "description": str(data.get("description", "")).strip(),
        }

        # Smart categorization: apply bank analysis patterns
        result["type"] = _smart_type_from_supplier(result["supplier"], result["category"])
        # Smart category from keywords (override AI if AI gave generic/Outros)
        kw_cat = _smart_category_from_supplier(result["supplier"])
        if kw_cat:
            if result["category"] in ("Outros", ""):
                result["category"] = kw_cat
                result["category_source"] = "palavras-chave"
            else:
                result["category_source"] = "IA"
        else:
            result["category_source"] = "IA" if result["category"] != "Outros" else "genérico"
        return result
    except Exception as e:
        logger.error(f"Invoice extraction failed: {e}")
        return {"error": str(e)}


# ── Smart categorization (same intelligence as bank analysis) ──────
_SUPPLIER_TYPE_MAP = {
    # Obra (electrical/construction suppliers)
    "worten": "obra", "leroy merlin": "obra", "aki": "obra", "bricomarche": "obra",
    "megaelectro": "obra", "janz": "obra", "schneider": "obra", "hager": "obra",
    "legrand": "obra", "abb": "obra", "siemens": "obra", "efapel": "obra",
    "cembre": "obra", "general cable": "obra", "cabelte": "obra", "solidal": "obra",
    "philips": "obra", "osram": "obra", "ledvance": "obra", "gewiss": "obra",
    "maxmat": "obra", "bigmat": "obra", "sotecnisol": "obra", "saint-gobain": "obra",
    # Fixo (utilities, insurance, rent)
    "vodafone": "fixo", "meo": "fixo", "nos ": "fixo", "nowo": "fixo",
    "edp": "fixo", "galp energia": "fixo", "endesa": "fixo", "iberdrola": "fixo",
    "epal": "fixo", "fidelidade": "fixo", "allianz": "fixo", "tranquilidade": "fixo",
    "ageas": "fixo", "ok teleseguros": "fixo", "contabilidade": "fixo",
    "contabilista": "fixo", "seguro": "fixo",
    # Variavel (fuel, meals, parking)
    "bp ": "variavel", "cepsa": "variavel", "repsol": "variavel", "prio": "variavel",
    "via verde": "variavel", "restaurante": "variavel", "estacionamento": "variavel",
}

_CATEGORY_TYPE_MAP = {
    "Material Elétrico": "obra", "Material de Obra": "obra", "Material": "obra",
    "Ferramentas": "obra", "Fornecedor": "obra",
    "Combustível": "variavel", "Viatura": "variavel", "Alimentação": "variavel",
    "Comunicações": "fixo", "Rendas": "fixo", "Seguros": "fixo",
    "Contabilidade/Advogado": "fixo", "Imposto/Taxa": "fixo",
    "Serviços": "variavel", "Outros": "variavel",
}


def _smart_type_from_supplier(supplier: str, ai_category: str) -> str:
    """Determine expense type (fixo/variavel/obra) from supplier name and AI category."""
    s = (supplier or "").lower()
    # 1. Check supplier name against known patterns
    for pattern, etype in _SUPPLIER_TYPE_MAP.items():
        if pattern in s:
            return etype
    # 2. Fall back to AI category mapping
    return _CATEGORY_TYPE_MAP.get(ai_category, "variavel")


# ── Smart category from supplier keywords ──────────────────────
_SUPPLIER_CATEGORY_MAP = {
    # Combustível
    "bp ": "Combustível", "cepsa": "Combustível", "repsol": "Combustível",
    "prio": "Combustível", "gasolina": "Combustível", "gasoleo": "Combustível",
    "combustivel": "Combustível", "galp": "Combustível",
    # Material (electrical/construction suppliers)
    "leroy merlin": "Material", "aki": "Material", "bricomarche": "Material",
    "worten": "Material", "megaelectro": "Material", "maxmat": "Material",
    "bigmat": "Material", "janz": "Material", "schneider": "Material",
    "hager": "Material", "legrand": "Material", "abb": "Material",
    "siemens": "Material", "efapel": "Material", "gewiss": "Material",
    "cembre": "Material", "general cable": "Material", "cabelte": "Material",
    "solidal": "Material", "philips": "Material", "osram": "Material",
    "ledvance": "Material", "sotecnisol": "Fornecedor", "saint-gobain": "Fornecedor",
    "material electrico": "Material", "material eletrico": "Material",
    # Ferramentas
    "ferramentas": "Ferramentas", "bosch": "Ferramentas", "dewalt": "Ferramentas",
    "makita": "Ferramentas", "hilti": "Ferramentas", "stanley": "Ferramentas",
    # Comunicações
    "vodafone": "Comunicações", "meo": "Comunicações", "nos ": "Comunicações", "nowo": "Comunicações",
    # Rendas
    "renda": "Rendas", "aluguer": "Rendas", "arrendamento": "Rendas",
    # Seguros
    "fidelidade": "Seguros", "allianz": "Seguros", "tranquilidade": "Seguros",
    "ageas": "Seguros", "ok teleseguros": "Seguros", "seguro": "Seguros",
    # Contabilidade/Advogado
    "contabilidade": "Contabilidade/Advogado", "contabilista": "Contabilidade/Advogado",
    "advogado": "Contabilidade/Advogado", "advocacia": "Contabilidade/Advogado",
    # Imposto/Taxa
    "at.gov": "Imposto/Taxa", "autoridade tributaria": "Imposto/Taxa",
    "impostos": "Imposto/Taxa", "seg social": "Imposto/Taxa",
    "seguranca social": "Imposto/Taxa",
    # Alimentação
    "restaurante": "Alimentação", "refeicao": "Alimentação", "cafe": "Alimentação",
    "supermercado": "Alimentação", "pingo doce": "Alimentação",
    "continente": "Alimentação", "lidl": "Alimentação",
    # Viatura (portagens, estacionamento)
    "via verde": "Viatura", "portagem": "Viatura", "estacionamento": "Viatura",
    "parking": "Viatura", "scut": "Viatura",
    # Serviços (energia/água)
    "edp": "Serviços", "endesa": "Serviços", "iberdrola": "Serviços", "epal": "Serviços",
}


def _smart_category_from_supplier(supplier: str) -> Optional[str]:
    """Determine expense category from supplier name using keyword matching."""
    s = (supplier or "").lower()
    for pattern, cat in _SUPPLIER_CATEGORY_MAP.items():
        if pattern in s:
            return cat
    return None


def _month_prefix(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}"


def _norm_text(value: str) -> str:
    value = (value or "").strip().lower()
    value = re.sub(r"[^\w\s]", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, _norm_text(a), _norm_text(b)).ratio()


def _parse_iso_date(value: str) -> Optional[datetime]:
    try:
        return datetime.strptime((value or "")[:10], "%Y-%m-%d")
    except Exception:
        return None


def _format_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value or "")[:10]


def _round_money(value) -> float:
    return round(abs(float(value or 0)), 2)


def _expense_anchor(doc: dict) -> str:
    invoice_number = _norm_text(doc.get("invoice_number") or "")
    if invoice_number:
        return invoice_number
    supplier = _norm_text(doc.get("supplier") or doc.get("description") or "")
    return supplier[:80]


def infer_expense_source_kind(doc: dict) -> str:
    if doc.get("source_kind"):
        return doc.get("source_kind")
    if doc.get("fixed_cost_instance_id"):
        return "fixed_cost"
    if doc.get("bank_txn_id") or doc.get("bank_analysis_id"):
        return "bank"
    if doc.get("invoice_number") or doc.get("invoice_file"):
        return "fiscal"
    return "manual"


def build_hard_dedupe_key(doc: dict) -> Optional[str]:
    date = _format_date(doc.get("date"))
    amount = _round_money(doc.get("value_gross"))
    anchor = _expense_anchor(doc)
    if not date or amount <= 0 or not anchor:
        return None
    return f"{date}|{anchor}|{amount:.2f}"


def _get_bank_history_entry(doc: dict) -> Optional[dict]:
    bank_txn_id = doc.get("bank_txn_id")
    explicit_description = (doc.get("bank_description") or "").strip()
    description = explicit_description or ((doc.get("supplier") or "").strip() if bank_txn_id else "")
    if not bank_txn_id and not explicit_description:
        return None
    return {
        "bank_txn_id": bank_txn_id,
        "date": _format_date(doc.get("bank_txn_date") or doc.get("date")),
        "description": description[:180],
        "amount": _round_money(doc.get("bank_amount") or doc.get("value_gross")),
        "analysis_id": doc.get("bank_analysis_id"),
        "matched_at": datetime.now(timezone.utc).isoformat(),
    }


def _merge_bank_history(existing: dict, incoming: dict) -> list:
    history = list(existing.get("bank_movement_history") or [])
    if existing.get("bank_txn_id") and not history:
        entry = _get_bank_history_entry({
            "bank_txn_id": existing.get("bank_txn_id"),
            "bank_txn_date": existing.get("date"),
            "bank_description": existing.get("supplier"),
            "bank_amount": existing.get("value_gross"),
            "bank_analysis_id": existing.get("bank_analysis_id"),
        })
        if entry:
            history.append(entry)
    incoming_entry = _get_bank_history_entry(incoming)
    if incoming_entry:
        seen = {item.get("bank_txn_id") or f"{item.get('date')}|{item.get('description')}|{item.get('amount')}" for item in history}
        key = incoming_entry.get("bank_txn_id") or f"{incoming_entry.get('date')}|{incoming_entry.get('description')}|{incoming_entry.get('amount')}"
        if key not in seen:
            history.append(incoming_entry)
    return history


def _merge_fiscal_history(existing: dict, incoming: dict) -> list:
    history = list(existing.get("fiscal_document_history") or [])
    invoice_number = (incoming.get("invoice_number") or "").strip()
    invoice_file = incoming.get("invoice_file")
    if not invoice_number and not invoice_file:
        return history
    entry = {
        "supplier": (incoming.get("supplier") or "")[:180],
        "invoice_number": invoice_number,
        "invoice_file": invoice_file,
        "nif": incoming.get("nif") or "",
        "date": _format_date(incoming.get("date")),
        "value_gross": _round_money(incoming.get("value_gross")),
        "linked_at": datetime.now(timezone.utc).isoformat(),
    }
    seen = {f"{item.get('invoice_number','')}|{item.get('invoice_file','')}|{item.get('date','')}|{item.get('value_gross',0)}" for item in history}
    key = f"{entry['invoice_number']}|{entry['invoice_file']}|{entry['date']}|{entry['value_gross']}"
    if key not in seen:
        history.append(entry)
    return history


def _collect_bank_history(doc: dict) -> list:
    history = list(doc.get("bank_movement_history") or [])
    extra = _get_bank_history_entry(doc)
    if extra:
        seen = {item.get("bank_txn_id") or f"{item.get('date')}|{item.get('description')}|{item.get('amount')}" for item in history}
        key = extra.get("bank_txn_id") or f"{extra.get('date')}|{extra.get('description')}|{extra.get('amount')}"
        if key not in seen:
            history.append(extra)
    return history


def _collect_fiscal_history(doc: dict) -> list:
    history = list(doc.get("fiscal_document_history") or [])
    invoice_number = (doc.get("invoice_number") or "").strip()
    invoice_file = doc.get("invoice_file")
    if invoice_number or invoice_file:
        entry = {
            "supplier": (doc.get("supplier") or "")[:180],
            "invoice_number": invoice_number,
            "invoice_file": invoice_file,
            "nif": doc.get("nif") or "",
            "date": _format_date(doc.get("date")),
            "value_gross": _round_money(doc.get("value_gross")),
            "linked_at": datetime.now(timezone.utc).isoformat(),
        }
        seen = {f"{item.get('invoice_number','')}|{item.get('invoice_file','')}|{item.get('date','')}|{item.get('value_gross',0)}" for item in history}
        key = f"{entry['invoice_number']}|{entry['invoice_file']}|{entry['date']}|{entry['value_gross']}"
        if key not in seen:
            history.append(entry)
    return history


def _merge_unique_notes(*notes: str) -> str:
    unique = []
    for note in notes:
        text = (note or "").strip()
        if text and text not in unique:
            unique.append(text)
    return "\n".join(unique)


def _is_generic_category(category: str) -> bool:
    return (category or "").strip().lower() in {"", "outros", "outro"}


def _expense_strength(doc: dict):
    has_fiscal = bool(doc.get("invoice_number") or doc.get("invoice_file") or doc.get("fiscal_document_history"))
    has_bank = bool(doc.get("bank_txn_id") or doc.get("bank_movement_history"))
    score = 0
    if doc.get("reconciled") or infer_expense_source_kind(doc) == "reconciled":
        score += 500
    if has_fiscal:
        score += 220
    if has_bank:
        score += 80
    if (doc.get("supplier") or "").strip():
        score += 25
    if (doc.get("nif") or "").strip():
        score += 15
    if (doc.get("notes") or "").strip():
        score += 5
    created_at = doc.get("created_at") or f"{_format_date(doc.get('date'))}T00:00:00"
    try:
        epoch = datetime.fromisoformat(created_at.replace("Z", "+00:00")).timestamp()
    except Exception:
        epoch = 0
    return (score, 1 if has_fiscal else 0, 1 if has_bank else 0, -epoch)


def _pick_primary_expense(docs: list) -> dict:
    return sorted(docs, key=_expense_strength, reverse=True)[0]


def merge_existing_expenses(primary: dict, secondary: dict, *, mode: str) -> dict:
    now_iso = datetime.now(timezone.utc).isoformat()
    merged = {**primary}

    for field in ("supplier", "nif", "invoice_number", "invoice_file", "payment_method", "obra_id", "obra_name", "fixed_cost_instance_id"):
        if not merged.get(field) and secondary.get(field):
            merged[field] = secondary.get(field)

    if _is_generic_category(merged.get("category")) and secondary.get("category"):
        merged["category"] = secondary.get("category")
    if not merged.get("type") and secondary.get("type"):
        merged["type"] = secondary.get("type")
    if not merged.get("value_net") and secondary.get("value_net"):
        merged["value_net"] = secondary.get("value_net")
    if not merged.get("vat_amount") and secondary.get("vat_amount"):
        merged["vat_amount"] = secondary.get("vat_amount")
    if not merged.get("vat_rate") and secondary.get("vat_rate"):
        merged["vat_rate"] = secondary.get("vat_rate")

    merged["notes"] = _merge_unique_notes(primary.get("notes", ""), secondary.get("notes", ""))
    merged["bank_movement_history"] = _collect_bank_history(primary)
    for item in _collect_bank_history(secondary):
        if item.get("bank_txn_id") not in {row.get("bank_txn_id") for row in merged["bank_movement_history"]}:
            merged["bank_movement_history"].append(item)
    merged["fiscal_document_history"] = _collect_fiscal_history(primary)
    existing_fiscal_keys = {f"{item.get('invoice_number','')}|{item.get('invoice_file','')}|{item.get('date','')}|{item.get('value_gross',0)}" for item in merged["fiscal_document_history"]}
    for item in _collect_fiscal_history(secondary):
        key = f"{item.get('invoice_number','')}|{item.get('invoice_file','')}|{item.get('date','')}|{item.get('value_gross',0)}"
        if key not in existing_fiscal_keys:
            merged["fiscal_document_history"].append(item)
            existing_fiscal_keys.add(key)

    if merged.get("bank_movement_history"):
        primary_bank = merged["bank_movement_history"][0]
        merged["bank_txn_id"] = primary_bank.get("bank_txn_id")
        merged["bank_description"] = primary_bank.get("description")
        merged["bank_analysis_id"] = primary_bank.get("analysis_id")

    source_kinds = {*(primary.get("source_kinds") or [infer_expense_source_kind(primary)]), *(secondary.get("source_kinds") or [infer_expense_source_kind(secondary)])}
    has_fiscal = bool(merged.get("invoice_number") or merged.get("invoice_file") or merged.get("fiscal_document_history"))
    has_bank = bool(merged.get("bank_txn_id") or merged.get("bank_movement_history"))
    merged["source_kinds"] = sorted(source_kinds)
    merged["reconciled"] = has_fiscal and has_bank
    merged["source_kind"] = "reconciled" if merged["reconciled"] else infer_expense_source_kind(merged)
    if merged["reconciled"]:
        merged["reconciled_at"] = now_iso
        merged["reconciliation_status"] = "matched_fiscal_bank"
        merged["reconciliation_rule"] = "date±2d+exact_amount"
    elif mode == "hard_duplicate":
        merged["reconciliation_status"] = "hard_duplicate_merged"
        merged["reconciliation_rule"] = "same_date+same_anchor+same_amount"

    merged["display_description"] = " · ".join(part for part in [merged.get("supplier"), merged.get("invoice_number")] if part)
    merged["merged_from_ids"] = sorted({*(primary.get("merged_from_ids") or []), *(secondary.get("merged_from_ids") or []), secondary.get("id")})
    merged["updated_at"] = now_iso
    merged["hard_dedupe_key"] = build_hard_dedupe_key(merged)
    return merged


def _build_expense_scope_query(month: Optional[int], year: Optional[int], category: Optional[str], type_value: Optional[str], scope: str = "month") -> tuple[dict, dict]:
    now = datetime.now(timezone.utc)
    q = {}
    resolved_year = year or now.year
    resolved_month = month or now.month
    if scope != "all":
        q["date"] = {"$regex": f"^{_month_prefix(resolved_year, resolved_month)}"}
    elif resolved_year:
        q["date"] = {"$regex": f"^{resolved_year:04d}"}
    if category:
        q["category"] = category
    if type_value:
        q["type"] = type_value
    return q, {"scope": scope, "month": resolved_month, "year": resolved_year, "category": category or "", "type": type_value or ""}


def _format_euro(value) -> float:
    return round(float(value or 0), 2)


def _autosize_worksheet(ws):
    for col in ws.columns:
        values = [str(cell.value or "") for cell in col]
        max_len = min(max((len(v) for v in values), default=10) + 2, 48)
        ws.column_dimensions[col[0].column_letter].width = max_len


def _write_sheet_rows(ws, title: str, rows: list[dict]):
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=13)
    if not rows:
        ws.append(["Sem registos"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="FACC15")
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    _autosize_worksheet(ws)


def _build_reconciliation_report_workbook(report_doc: dict, file_path: Path):
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Resumo"
    summary_rows = [
        {"campo": "Relatório ID", "valor": report_doc.get("id")},
        {"campo": "Gerado em", "valor": report_doc.get("created_at")},
        {"campo": "Gerado por", "valor": report_doc.get("created_by_name") or report_doc.get("created_by_email") or report_doc.get("created_by")},
        {"campo": "Escopo", "valor": report_doc.get("scope", {}).get("scope")},
        {"campo": "Mês", "valor": report_doc.get("scope", {}).get("month")},
        {"campo": "Ano", "valor": report_doc.get("scope", {}).get("year")},
        {"campo": "Categoria filtro", "valor": report_doc.get("scope", {}).get("category")},
        {"campo": "Tipo filtro", "valor": report_doc.get("scope", {}).get("type")},
        {"campo": "Registos analisados", "valor": report_doc.get("summary", {}).get("records_scanned")},
        {"campo": "Pares reconciliados", "valor": report_doc.get("result", {}).get("reconciled")},
        {"campo": "Duplicados removidos", "valor": report_doc.get("result", {}).get("duplicates_removed")},
        {"campo": "Grupos de duplicados tratados", "valor": report_doc.get("result", {}).get("duplicate_groups_processed")},
    ]
    _write_sheet_rows(summary_ws, "Resumo da operação", summary_rows)

    reconciled_ws = wb.create_sheet("Reconciliadas")
    _write_sheet_rows(reconciled_ws, "Itens reconciliados", report_doc.get("reconciled_items") or [])

    duplicates_ws = wb.create_sheet("Duplicados Removidos")
    _write_sheet_rows(duplicates_ws, "Itens removidos", report_doc.get("removed_duplicates") or [])

    wb.save(file_path)


async def _create_reconciliation_report(db, actor_user: dict, scope_meta: dict, summary: dict, result_meta: dict, reconciled_items: list, removed_duplicates: list) -> dict:
    created_at = datetime.now(timezone.utc).isoformat()
    report_id = str(uuid.uuid4())
    filename = f"reconciliacao_despesas_{scope_meta.get('year')}_{scope_meta.get('month'):02d}_{report_id[:8]}.xlsx"
    file_path = RECONCILIATION_REPORTS_DIR / filename
    report_doc = {
        "id": report_id,
        "created_at": created_at,
        "created_by": actor_user.get("id"),
        "created_by_name": actor_user.get("name") or "",
        "created_by_email": actor_user.get("email") or "",
        "scope": scope_meta,
        "summary": summary,
        "result": result_meta,
        "reconciled_items": reconciled_items,
        "removed_duplicates": removed_duplicates,
        "file_name": filename,
        "file_path": str(file_path),
        "download_url": f"/api/expenses/reconcile-reports/{report_id}/download",
    }
    _build_reconciliation_report_workbook(report_doc, file_path)
    await db.expense_reconciliation_reports.insert_one({**report_doc})
    report_doc.pop("_id", None)
    return report_doc


def _build_bulk_reconciliation_plan(expenses: list) -> dict:
    hard_groups_map = defaultdict(list)
    for doc in expenses:
        if doc.get("dedupe_exempt"):
            continue
        key = build_hard_dedupe_key(doc)
        if key:
            hard_groups_map[key].append(doc)

    hard_duplicate_groups = []
    loser_ids = set()
    for key, docs in hard_groups_map.items():
        if len(docs) <= 1:
            continue
        keeper = _pick_primary_expense(docs)
        losers = [doc for doc in docs if doc.get("id") != keeper.get("id")]
        loser_ids.update(doc.get("id") for doc in losers)
        hard_duplicate_groups.append({
            "hard_dedupe_key": key,
            "keep_id": keeper.get("id"),
            "keep_supplier": keeper.get("supplier"),
            "keep_invoice_number": keeper.get("invoice_number"),
            "remove_ids": [doc.get("id") for doc in losers],
            "remove_count": len(losers),
        })

    active_docs = [doc for doc in expenses if doc.get("id") not in loser_ids]
    fiscal_docs = [
        doc for doc in active_docs
        if bool(doc.get("invoice_number") or doc.get("invoice_file") or doc.get("fiscal_document_history"))
        and not bool(doc.get("bank_txn_id") or doc.get("bank_movement_history"))
    ]
    bank_docs = [
        doc for doc in active_docs
        if bool(doc.get("bank_txn_id") or doc.get("bank_movement_history"))
        and not bool(doc.get("invoice_number") or doc.get("invoice_file") or doc.get("fiscal_document_history"))
    ]

    used_bank_ids = set()
    reconciliation_pairs = []
    for fiscal in sorted(fiscal_docs, key=lambda doc: (_format_date(doc.get("date")), doc.get("supplier", ""))):
        fiscal_date = _parse_iso_date(fiscal.get("date"))
        fiscal_amount = _round_money(fiscal.get("value_gross"))
        if not fiscal_date or fiscal_amount <= 0:
            continue

        ranked = []
        for bank in bank_docs:
            if bank.get("id") in used_bank_ids:
                continue
            bank_date = _parse_iso_date(bank.get("date"))
            if not bank_date:
                continue
            bank_amount = _round_money(bank.get("value_gross"))
            if bank_amount != fiscal_amount:
                continue
            diff_days = abs((bank_date - fiscal_date).days)
            if diff_days > 2:
                continue
            sim = _similarity(fiscal.get("supplier", ""), bank.get("bank_description") or bank.get("supplier", ""))
            ranked.append((diff_days, -sim, bank))

        if ranked:
            ranked.sort(key=lambda row: (row[0], row[1]))
            bank = ranked[0][2]
            used_bank_ids.add(bank.get("id"))
            reconciliation_pairs.append({
                "fiscal_id": fiscal.get("id"),
                "bank_id": bank.get("id"),
                "supplier": fiscal.get("supplier") or bank.get("supplier"),
                "invoice_number": fiscal.get("invoice_number") or "",
                "amount": fiscal_amount,
                "date_diff_days": abs((_parse_iso_date(bank.get("date")) - fiscal_date).days),
            })

    return {
        "hard_duplicate_groups": hard_duplicate_groups,
        "reconciliation_pairs": reconciliation_pairs,
        "summary": {
            "records_scanned": len(expenses),
            "hard_duplicate_groups": len(hard_duplicate_groups),
            "duplicates_to_remove": sum(group.get("remove_count", 0) for group in hard_duplicate_groups),
            "reconcilable_pairs": len(reconciliation_pairs),
        },
    }


async def run_bulk_expense_reconciliation(db, *, month: Optional[int], year: Optional[int], category: Optional[str], type_value: Optional[str], scope: str = "month", apply_changes: bool = False, actor_user: Optional[dict] = None) -> dict:
    query, scope_meta = _build_expense_scope_query(month, year, category, type_value, scope=scope)
    expenses = await db.expenses.find(query, {"_id": 0}).sort("date", 1).to_list(5000)
    plan = _build_bulk_reconciliation_plan(expenses)

    preview = {
        "scope": scope_meta,
        "summary": plan["summary"],
        "reconciliation_preview": plan["reconciliation_pairs"][:12],
        "hard_duplicate_preview": plan["hard_duplicate_groups"][:12],
        "dry_run": not apply_changes,
    }
    if not apply_changes:
        return preview

    duplicates_removed = 0
    duplicate_groups_processed = 0
    reconciled = 0
    removed_duplicates_items = []
    reconciled_items = []

    for group in plan["hard_duplicate_groups"]:
        keeper = await db.expenses.find_one({"id": group["keep_id"]}, {"_id": 0})
        if not keeper:
            continue
        changed = False
        for loser_id in group["remove_ids"]:
            loser = await db.expenses.find_one({"id": loser_id}, {"_id": 0})
            if not loser:
                continue
            removed_duplicates_items.append({
                "tipo_operacao": "hard_duplicate_removido",
                "despesa_principal_id": keeper.get("id"),
                "despesa_removida_id": loser.get("id"),
                "fornecedor_principal": keeper.get("supplier") or "",
                "fornecedor_removido": loser.get("supplier") or "",
                "fatura_principal": keeper.get("invoice_number") or "",
                "fatura_removida": loser.get("invoice_number") or "",
                "data": _format_date(loser.get("date")),
                "valor_total": _format_euro(loser.get("value_gross")),
                "motivo": "Hard duplicate removido automaticamente",
                "regra_aplicada": "same_date+same_anchor+same_amount",
                "executado_em": datetime.now(timezone.utc).isoformat(),
            })
            keeper = merge_existing_expenses(keeper, loser, mode="hard_duplicate")
            if loser.get("fixed_cost_instance_id"):
                await db.fixed_cost_instances.update_one({"id": loser.get("fixed_cost_instance_id")}, {"$set": {"expense_id": keeper["id"]}})
            await db.expenses.delete_one({"id": loser_id})
            duplicates_removed += 1
            changed = True
        if changed:
            await db.expenses.update_one({"id": keeper["id"]}, {"$set": keeper})
            duplicate_groups_processed += 1

    for pair in plan["reconciliation_pairs"]:
        fiscal = await db.expenses.find_one({"id": pair["fiscal_id"]}, {"_id": 0})
        bank = await db.expenses.find_one({"id": pair["bank_id"]}, {"_id": 0})
        if not fiscal or not bank:
            continue
        reconciled_items.append({
            "tipo_operacao": "reconciliacao_fiscal_banco",
            "despesa_canonica_id": fiscal.get("id"),
            "despesa_bancaria_removida_id": bank.get("id"),
            "fornecedor": fiscal.get("supplier") or bank.get("supplier") or "",
            "numero_fatura": fiscal.get("invoice_number") or "",
            "data_fiscal": _format_date(fiscal.get("date")),
            "data_banco": _format_date(bank.get("date")),
            "valor_total": _format_euro(fiscal.get("value_gross") or bank.get("value_gross")),
            "motivo": "Documento fiscal e débito bancário conciliados",
            "regra_aplicada": "date±2d+exact_amount",
            "executado_em": datetime.now(timezone.utc).isoformat(),
        })
        merged = merge_existing_expenses(fiscal, bank, mode="reconciliation")
        await db.expenses.update_one({"id": fiscal["id"]}, {"$set": merged})
        if bank.get("fixed_cost_instance_id"):
            await db.fixed_cost_instances.update_one({"id": bank.get("fixed_cost_instance_id")}, {"$set": {"expense_id": fiscal["id"]}})
        await db.expenses.delete_one({"id": bank["id"]})
        reconciled += 1

    result_meta = {
        "reconciled": reconciled,
        "duplicates_removed": duplicates_removed,
        "duplicate_groups_processed": duplicate_groups_processed,
    }
    report_doc = None
    if actor_user:
        report_doc = await _create_reconciliation_report(
            db,
            actor_user=actor_user,
            scope_meta=scope_meta,
            summary=preview["summary"],
            result_meta=result_meta,
            reconciled_items=reconciled_items,
            removed_duplicates=removed_duplicates_items,
        )

    return {
        **preview,
        "dry_run": False,
        "result": result_meta,
        "report": {
            "id": report_doc.get("id"),
            "created_at": report_doc.get("created_at"),
            "file_name": report_doc.get("file_name"),
            "download_url": report_doc.get("download_url"),
        } if report_doc else None,
        "message": f"Verificação concluída: {reconciled} despesas reconciliadas e {duplicates_removed} duplicados removidos com sucesso!",
    }


async def find_hard_duplicate_expense(db, expense_doc: dict, exclude_id: Optional[str] = None) -> Optional[dict]:
    if expense_doc.get("dedupe_exempt"):
        return None
    hard_key = build_hard_dedupe_key(expense_doc)
    if hard_key:
        query = {"hard_dedupe_key": hard_key, "dedupe_exempt": {"$ne": True}}
        if exclude_id:
            query["id"] = {"$ne": exclude_id}
        match = await db.expenses.find_one(query, {"_id": 0})
        if match:
            return match

    date = _format_date(expense_doc.get("date"))
    amount = _round_money(expense_doc.get("value_gross"))
    anchor = _expense_anchor(expense_doc)
    if not date or amount <= 0 or not anchor:
        return None
    query = {
        "date": date,
        "value_gross": {"$gte": amount - 0.001, "$lte": amount + 0.001},
        "dedupe_exempt": {"$ne": True},
    }
    if exclude_id:
        query["id"] = {"$ne": exclude_id}
    candidates = await db.expenses.find(query, {"_id": 0}).to_list(20)
    for candidate in candidates:
        if _expense_anchor(candidate) == anchor:
            return candidate
    return None


async def find_reconciliation_candidate(db, expense_doc: dict, source_kind: str, exclude_id: Optional[str] = None) -> Optional[dict]:
    date_dt = _parse_iso_date(expense_doc.get("date"))
    amount = _round_money(expense_doc.get("value_gross"))
    if not date_dt or amount <= 0:
        return None

    query = {
        "date": {
            "$gte": _format_date(date_dt - timedelta(days=2)),
            "$lte": _format_date(date_dt + timedelta(days=2)),
        },
        "value_gross": {"$gte": amount - 0.001, "$lte": amount + 0.001},
    }
    if exclude_id:
        query["id"] = {"$ne": exclude_id}

    candidates = await db.expenses.find(query, {"_id": 0}).to_list(50)
    preferred = []
    for candidate in candidates:
        candidate_source = infer_expense_source_kind(candidate)
        candidate_has_fiscal = bool(candidate.get("invoice_number") or candidate.get("invoice_file") or candidate.get("fiscal_document_history"))
        candidate_has_bank = bool(candidate.get("bank_txn_id") or candidate.get("bank_movement_history"))
        if source_kind == "fiscal":
            incoming_invoice = (expense_doc.get("invoice_number") or "").strip()
            same_invoice = bool(incoming_invoice and incoming_invoice == (candidate.get("invoice_number") or "").strip())
            if not candidate_has_bank:
                continue
            if candidate_has_fiscal and not same_invoice:
                continue
        elif source_kind == "bank":
            existing_bank_ids = {item.get("bank_txn_id") for item in (candidate.get("bank_movement_history") or []) if item.get("bank_txn_id")}
            if candidate.get("bank_txn_id"):
                existing_bank_ids.add(candidate.get("bank_txn_id"))
            incoming_bank_txn = expense_doc.get("bank_txn_id")
            if not candidate_has_fiscal:
                continue
            if existing_bank_ids and incoming_bank_txn not in existing_bank_ids:
                continue
        else:
            continue

        diff_days = abs((_parse_iso_date(candidate.get("date")) - date_dt).days) if _parse_iso_date(candidate.get("date")) else 99
        supplier_score = _similarity(expense_doc.get("supplier", ""), candidate.get("supplier", ""))
        preferred.append((diff_days, -supplier_score, candidate))

    preferred.sort(key=lambda row: (row[0], row[1]))
    return preferred[0][2] if preferred else None


async def preview_expense_ingestion(db, expense_doc: dict, source_kind: str, exclude_id: Optional[str] = None) -> dict:
    hard_duplicate = await find_hard_duplicate_expense(db, expense_doc, exclude_id=exclude_id)
    reconciliation_candidate = await find_reconciliation_candidate(db, expense_doc, source_kind=source_kind, exclude_id=exclude_id)
    return {
        "hard_duplicate": hard_duplicate,
        "reconciliation_candidate": reconciliation_candidate,
    }


async def upsert_reconciled_expense(db, expense_doc: dict, source_kind: str, user_id: str = "", *, force: bool = False) -> dict:
    now_iso = datetime.now(timezone.utc).isoformat()
    incoming = {**expense_doc}
    incoming["date"] = _format_date(incoming.get("date"))
    incoming["value_gross"] = _round_money(incoming.get("value_gross"))
    incoming["value_net"] = round(float(incoming.get("value_net", 0) or 0), 2)
    incoming["vat_amount"] = round(float(incoming.get("vat_amount", 0) or 0), 2)
    incoming["source_kind"] = source_kind
    incoming.setdefault("dedupe_exempt", False)
    incoming.setdefault("dedupe_exception_reason", "")
    incoming.setdefault("created_at", now_iso)
    if user_id:
        incoming.setdefault("created_by", user_id)

    if not incoming.get("id"):
        incoming["id"] = str(uuid.uuid4())

    if incoming.get("value_gross") and not incoming.get("value_net"):
        vat_rate = float(incoming.get("vat_rate", 23) or 23)
        incoming["value_net"] = round(incoming["value_gross"] / (1 + vat_rate / 100), 2)
        incoming["vat_amount"] = round(incoming["value_gross"] - incoming["value_net"], 2)

    preview = await preview_expense_ingestion(db, incoming, source_kind=source_kind)
    if preview["hard_duplicate"] and not force:
        return {
            "action": "hard_duplicate",
            "expense": preview["hard_duplicate"],
            "reason": "duplicate_hard",
        }

    reconciliation_candidate = preview["reconciliation_candidate"]
    if reconciliation_candidate:
        update_doc = {**reconciliation_candidate}
        if source_kind == "fiscal":
            update_doc.update({
                "supplier": (incoming.get("supplier") or reconciliation_candidate.get("supplier") or "")[:100],
                "nif": incoming.get("nif") or reconciliation_candidate.get("nif") or "",
                "invoice_number": (incoming.get("invoice_number") or reconciliation_candidate.get("invoice_number") or "")[:80],
                "invoice_file": incoming.get("invoice_file") or reconciliation_candidate.get("invoice_file"),
                "category": incoming.get("category") or reconciliation_candidate.get("category") or "Outros",
                "type": incoming.get("type") or reconciliation_candidate.get("type") or "variavel",
                "payment_method": incoming.get("payment_method") or reconciliation_candidate.get("payment_method") or "",
                "notes": incoming.get("notes") or reconciliation_candidate.get("notes") or "",
                "value_net": incoming.get("value_net") or reconciliation_candidate.get("value_net") or 0,
                "vat_rate": incoming.get("vat_rate") or reconciliation_candidate.get("vat_rate") or 23,
                "vat_amount": incoming.get("vat_amount") or reconciliation_candidate.get("vat_amount") or 0,
                "value_gross": incoming.get("value_gross") or reconciliation_candidate.get("value_gross") or 0,
            })
        else:
            if incoming.get("payment_method"):
                update_doc["payment_method"] = incoming.get("payment_method")
            if incoming.get("notes"):
                base_notes = reconciliation_candidate.get("notes") or ""
                if incoming.get("notes") not in base_notes:
                    update_doc["notes"] = (base_notes + "\n" + incoming.get("notes")).strip()

        bank_history = _merge_bank_history(reconciliation_candidate, incoming)
        fiscal_history = _merge_fiscal_history(reconciliation_candidate, incoming)
        source_kinds = sorted({*(reconciliation_candidate.get("source_kinds") or [infer_expense_source_kind(reconciliation_candidate)]), source_kind})
        update_doc.update({
            "source_kind": "reconciled",
            "source_kinds": source_kinds,
            "reconciled": True,
            "reconciled_at": now_iso,
            "reconciliation_status": "matched_fiscal_bank",
            "reconciliation_rule": "date±2d+exact_amount",
            "bank_movement_history": bank_history,
            "fiscal_document_history": fiscal_history,
            "display_description": " · ".join(part for part in [update_doc.get("supplier"), update_doc.get("invoice_number")] if part),
            "hard_dedupe_key": build_hard_dedupe_key(update_doc),
        })
        if bank_history:
            update_doc["bank_txn_id"] = bank_history[0].get("bank_txn_id")
            update_doc["bank_description"] = bank_history[0].get("description")
            update_doc["bank_analysis_id"] = bank_history[0].get("analysis_id")

        await db.expenses.update_one({"id": reconciliation_candidate["id"]}, {"$set": update_doc})
        refreshed = await db.expenses.find_one({"id": reconciliation_candidate["id"]}, {"_id": 0})
        return {
            "action": "reconciled_existing",
            "expense": refreshed,
            "reason": "matched_fiscal_bank",
        }

    incoming["source_kinds"] = [source_kind]
    incoming["reconciled"] = bool(incoming.get("reconciled"))
    incoming["display_description"] = " · ".join(part for part in [incoming.get("supplier"), incoming.get("invoice_number")] if part)
    incoming["hard_dedupe_key"] = build_hard_dedupe_key(incoming)
    if source_kind == "bank":
        bank_history = _merge_bank_history({}, incoming)
        if bank_history:
            incoming["bank_movement_history"] = bank_history
            incoming["bank_description"] = bank_history[0].get("description")
    if source_kind == "fiscal":
        fiscal_history = _merge_fiscal_history({}, incoming)
        if fiscal_history:
            incoming["fiscal_document_history"] = fiscal_history

    try:
        await db.expenses.insert_one(incoming)
    except DuplicateKeyError:
        existing = await find_hard_duplicate_expense(db, incoming)
        return {
            "action": "hard_duplicate",
            "expense": existing or incoming,
            "reason": "duplicate_index",
        }
    incoming.pop("_id", None)
    return {
        "action": "created",
        "expense": incoming,
        "reason": None,
    }


def create_expenses_router(db, get_current_user):

    async def _find_duplicate(invoice_number: str, nif: str = "", supplier: str = "", exclude_id: Optional[str] = None) -> Optional[dict]:
        """Match on (invoice_number) AND (NIF OR supplier). Case-insensitive, trimmed."""
        inv = (invoice_number or "").strip()
        if not inv:
            return None
        import re as _re
        inv_re = f"^{_re.escape(inv)}$"
        or_clauses = []
        if nif and nif.strip():
            or_clauses.append({"nif": nif.strip()})
        if supplier and supplier.strip():
            or_clauses.append({"supplier": {"$regex": f"^{_re.escape(supplier.strip())}$", "$options": "i"}})
        if not or_clauses:
            return None
        query = {
            "invoice_number": {"$regex": inv_re, "$options": "i"},
            "$or": or_clauses,
        }
        if exclude_id:
            query["id"] = {"$ne": exclude_id}
        return await db.expenses.find_one(query, {"_id": 0})

    @expenses_router.get("/categories")
    async def get_categories(user=Depends(get_current_user)):
        return CATEGORIES

    @expenses_router.post("/extract")
    async def extract_from_upload(file: UploadFile = File(...), user=Depends(get_current_user)):
        """Upload a file and get AI-extracted invoice data with smart categorization and duplicate detection."""
        filename = file.filename or "invoice"
        ext = Path(filename).suffix.lower() or ".bin"
        if ext not in [".pdf", ".jpg", ".jpeg", ".png", ".webp"]:
            raise HTTPException(status_code=400, detail="Formato não suportado. Use PDF, JPG, PNG ou WEBP.")

        mime_map = {
            ".pdf": "application/pdf",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".webp": "image/webp",
        }
        mime = mime_map[ext]

        # Save to disk (for persistence)
        saved_name = f"{uuid.uuid4().hex}{ext}"
        saved_path = EXPENSES_DIR / saved_name
        content = await file.read()
        saved_path.write_bytes(content)

        # Call AI
        extracted = await extract_invoice_data(saved_path, mime)

        if isinstance(extracted, dict) and extracted.get("error"):
            return {"file_name": saved_name, "original_name": filename, "file_size": len(content), "extracted": extracted, "duplicate": None, "suggestions": None}

        # ── Smart suggestions from expense history ──────────────
        suggestions = None
        if isinstance(extracted, dict) and extracted.get("supplier"):
            supplier_lower = extracted["supplier"].lower()
            # Check historical expenses for this supplier
            hist = await db.expenses.find(
                {"supplier": {"$regex": extracted["supplier"][:20], "$options": "i"}},
                {"_id": 0, "category": 1, "type": 1, "obra_id": 1, "obra_name": 1}
            ).sort("created_at", -1).to_list(5)
            if hist:
                from collections import Counter
                cats = Counter(h.get("category", "Outros") for h in hist)
                types = Counter(h.get("type", "variavel") for h in hist)
                most_cat = cats.most_common(1)[0][0]
                most_type = types.most_common(1)[0][0]
                last_obra = next((h for h in hist if h.get("obra_id")), None)
                suggestions = {
                    "category": most_cat,
                    "type": most_type,
                    "source": "histórico",
                    "confidence": round(cats.most_common(1)[0][1] / len(hist) * 100),
                    "obra_id": last_obra.get("obra_id") if last_obra else None,
                    "obra_name": last_obra.get("obra_name") if last_obra else None,
                }
                # History overrides AI/keywords if AI didn't have a strong opinion
                if most_cat != "Outros":
                    extracted["category"] = most_cat
                    extracted["category_source"] = "histórico"
                if most_type:
                    extracted["type"] = most_type

        # ── Duplicate detection (3 layers) ──────────────────────
        duplicate = None
        if isinstance(extracted, dict) and extracted.get("date") and extracted.get("value_gross"):
            preview_doc = {
                "date": extracted.get("date"),
                "supplier": extracted.get("supplier", ""),
                "nif": extracted.get("nif", ""),
                "invoice_number": extracted.get("invoice_number", ""),
                "value_net": extracted.get("value_net", 0),
                "vat_rate": extracted.get("vat_rate", 23),
                "vat_amount": extracted.get("vat_amount", 0),
                "value_gross": extracted.get("value_gross", 0),
                "invoice_file": saved_name,
            }
            preview = await preview_expense_ingestion(db, preview_doc, source_kind="fiscal")
            match = preview.get("hard_duplicate") or preview.get("reconciliation_candidate")
            if match:
                duplicate = {
                    "id": match.get("id"),
                    "supplier": match.get("supplier"),
                    "invoice_number": match.get("invoice_number"),
                    "date": match.get("date"),
                    "value_gross": match.get("value_gross"),
                    "reason": "Fatura já existente" if preview.get("hard_duplicate") else "Será reconciliado com movimento bancário existente",
                    "match_mode": "hard_duplicate" if preview.get("hard_duplicate") else "reconciliation_candidate",
                }

        return {
            "file_name": saved_name,
            "original_name": filename,
            "file_size": len(content),
            "extracted": extracted,
            "duplicate": duplicate,
            "suggestions": suggestions,
        }

    @expenses_router.post("")
    async def create_expense(input: ExpenseCreate, force: bool = False, user=Depends(get_current_user)):
        doc = {
            **input.model_dump(),
            "id": str(uuid.uuid4()),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user["id"],
        }
        if not doc.get("value_gross") and doc.get("value_net"):
            doc["value_gross"] = round(doc["value_net"] * (1 + doc.get("vat_rate", 23) / 100), 2)
            doc["vat_amount"] = round(doc["value_gross"] - doc["value_net"], 2)
        source_kind = infer_expense_source_kind(doc)
        result = await upsert_reconciled_expense(db, doc, source_kind=source_kind, user_id=user["id"], force=force)
        if result["action"] == "hard_duplicate":
            existing = result.get("expense") or {}
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "duplicate_invoice",
                    "message": f"Despesa já registada para {existing.get('supplier', '—')} em {existing.get('date', '—')}.",
                    "existing": existing,
                },
            )
        expense = result["expense"]
        expense["ingestion_action"] = result["action"]
        return expense

    @expenses_router.get("")
    async def list_expenses(
        month: Optional[int] = None,
        year: Optional[int] = None,
        category: Optional[str] = None,
        type: Optional[str] = None,
        obra_id: Optional[str] = None,
        user=Depends(get_current_user),
    ):
        q = {}
        if year and month:
            q["date"] = {"$regex": f"^{_month_prefix(year, month)}"}
        elif year:
            q["date"] = {"$regex": f"^{year:04d}"}
        if category:
            q["category"] = category
        if type:
            q["type"] = type
        if obra_id:
            q["obra_id"] = obra_id
        expenses = await db.expenses.find(q, {"_id": 0}).sort("date", -1).to_list(2000)
        return expenses

    @expenses_router.get("/summary")
    async def summary(year: Optional[int] = None, month: Optional[int] = None, user=Depends(get_current_user)):
        now = datetime.now(timezone.utc)
        y = year or now.year
        all_exp = await db.expenses.find({"date": {"$regex": f"^{y:04d}"}}, {"_id": 0}).to_list(5000)

        by_month = {m: 0 for m in range(1, 13)}
        iva_by_month = {m: 0 for m in range(1, 13)}
        count_by_month = {m: 0 for m in range(1, 13)}
        by_category = {}
        by_type = {"fixo": 0, "variavel": 0, "obra": 0}
        by_obra = {}
        total_year = 0
        total_iva = 0

        for e in all_exp:
            try:
                m = int(e.get("date", "")[5:7])
            except Exception:
                m = 0
            gross = e.get("value_gross", 0) or 0
            iva = e.get("vat_amount", 0) or 0
            total_year += gross
            total_iva += iva
            if m:
                by_month[m] = round(by_month.get(m, 0) + gross, 2)
                iva_by_month[m] = round(iva_by_month.get(m, 0) + iva, 2)
                count_by_month[m] = count_by_month.get(m, 0) + 1
            cat = e.get("category", "Outros")
            by_category[cat] = round(by_category.get(cat, 0) + gross, 2)
            tp = e.get("type", "variavel")
            by_type[tp] = round(by_type.get(tp, 0) + gross, 2)
            if e.get("obra_id"):
                ob = e.get("obra_name") or e["obra_id"]
                by_obra[ob] = round(by_obra.get(ob, 0) + gross, 2)

        # Selected month: defaults to current month if same year, else December
        selected_month = month if (month and 1 <= month <= 12) else (now.month if now.year == y else 12)

        return {
            "year": y,
            "selected_month": selected_month,
            "total_year": round(total_year, 2),
            "total_iva": round(total_iva, 2),
            "count_year": len(all_exp),
            "month_total": round(by_month.get(selected_month, 0), 2),
            "month_iva": round(iva_by_month.get(selected_month, 0), 2),
            "month_count": count_by_month.get(selected_month, 0),
            "current_month_total": round(by_month.get(selected_month, 0), 2),  # backward compat
            "count": len(all_exp),  # backward compat
            "by_month": by_month,
            "iva_by_month": iva_by_month,
            "count_by_month": count_by_month,
            "by_category": by_category,
            "by_type": by_type,
            "by_obra": by_obra,
        }

    @expenses_router.get("/reconcile-preview")
    async def reconcile_preview(
        month: Optional[int] = None,
        year: Optional[int] = None,
        category: Optional[str] = None,
        type: Optional[str] = None,
        scope: str = "month",
        user=Depends(get_current_user),
    ):
        if user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Apenas administradores")
        return await run_bulk_expense_reconciliation(
            db,
            month=month,
            year=year,
            category=category,
            type_value=type,
            scope=scope,
            apply_changes=False,
        )

    @expenses_router.post("/reconcile-apply")
    async def reconcile_apply(
        month: Optional[int] = None,
        year: Optional[int] = None,
        category: Optional[str] = None,
        type: Optional[str] = None,
        scope: str = "month",
        user=Depends(get_current_user),
    ):
        if user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Apenas administradores")
        return await run_bulk_expense_reconciliation(
            db,
            month=month,
            year=year,
            category=category,
            type_value=type,
            scope=scope,
            apply_changes=True,
            actor_user=user,
        )

    @expenses_router.get("/reconcile-reports")
    async def list_reconciliation_reports(limit: int = 30, user=Depends(get_current_user)):
        if user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Apenas administradores")
        docs = await db.expense_reconciliation_reports.find({}, {"_id": 0, "reconciled_items": 0, "removed_duplicates": 0}).sort("created_at", -1).to_list(max(1, min(limit, 100)))
        return docs

    @expenses_router.get("/reconcile-reports/{report_id}/download")
    async def download_reconciliation_report(report_id: str, user=Depends(get_current_user)):
        if user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Apenas administradores")
        report = await db.expense_reconciliation_reports.find_one({"id": report_id}, {"_id": 0})
        if not report:
            raise HTTPException(status_code=404, detail="Relatório não encontrado")
        file_path = Path(report.get("file_path") or "")
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Ficheiro não encontrado")
        return FileResponse(path=file_path, filename=report.get("file_name") or file_path.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @expenses_router.get("/{expense_id}")
    async def get_expense(expense_id: str, user=Depends(get_current_user)):
        e = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
        if not e:
            raise HTTPException(status_code=404, detail="Despesa não encontrada")
        return e

    @expenses_router.put("/{expense_id}")
    async def update_expense(expense_id: str, input: ExpenseUpdate, force: bool = False, user=Depends(get_current_user)):
        data = input.model_dump(exclude_none=True)
        if not data:
            raise HTTPException(status_code=400, detail="Nada para atualizar")
        current = await db.expenses.find_one({"id": expense_id}, {"_id": 0}) or {}
        candidate = {**current, **data}
        candidate["hard_dedupe_key"] = build_hard_dedupe_key(candidate)
        if not force:
            dup = await find_hard_duplicate_expense(db, candidate, exclude_id=expense_id)
            if dup:
                raise HTTPException(
                    status_code=409,
                    detail={
                        "code": "duplicate_invoice",
                        "message": f"Despesa já registada para {dup.get('supplier', '—')} em {dup.get('date', '—')}.",
                        "existing": dup,
                    },
                )
        r = await db.expenses.update_one({"id": expense_id}, {"$set": candidate})
        if r.matched_count == 0:
            raise HTTPException(status_code=404, detail="Despesa não encontrada")
        e = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
        return e

    @expenses_router.delete("/{expense_id}")
    async def delete_expense(expense_id: str, user=Depends(get_current_user)):
        e = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
        if not e:
            raise HTTPException(status_code=404, detail="Despesa não encontrada")
        # Delete file
        if e.get("invoice_file"):
            fp = EXPENSES_DIR / e["invoice_file"]
            if fp.exists():
                try:
                    fp.unlink()
                except Exception:
                    pass
        await db.expenses.delete_one({"id": expense_id})
        return {"ok": True}

    @expenses_router.get("/file/{filename}")
    async def get_invoice_file(filename: str, user=Depends(get_current_user)):
        """Serve an uploaded invoice file (PDF/image)."""
        from fastapi.responses import FileResponse
        fp = EXPENSES_DIR / filename
        if not fp.exists():
            raise HTTPException(status_code=404, detail="Ficheiro não encontrado")
        return FileResponse(fp)

    @expenses_router.post("/ai-categorize")
    async def ai_categorize_expenses(user=Depends(get_current_user)):
        """Re-categorize all expenses using AI (keyword matching + LLM fallback).
        Only touches expenses with category 'Outros' or missing type."""
        api_key = os.environ.get("EMERGENT_LLM_KEY", "")

        all_exp = await db.expenses.find({}, {"_id": 0}).to_list(5000)
        updated = 0
        ai_batch = []

        for e in all_exp:
            eid = e.get("id")
            supplier = e.get("supplier", "")
            old_cat = e.get("category", "Outros")
            old_type = e.get("type", "variavel")
            changes = {}

            # 1. Keyword-based categorization
            kw_cat = _smart_category_from_supplier(supplier)
            if kw_cat and old_cat in ("Outros", "", "outro", "Outro"):
                changes["category"] = kw_cat

            # 2. Type inference from supplier
            new_type = _smart_type_from_supplier(supplier, changes.get("category", old_cat))
            if new_type != old_type:
                changes["type"] = new_type

            if changes:
                await db.expenses.update_one({"id": eid}, {"$set": changes})
                updated += 1
            elif old_cat in ("Outros", "", "outro", "Outro"):
                # Queue for AI categorization
                ai_batch.append(e)

        # 3. AI batch for remaining uncategorized
        ai_updated = 0
        if ai_batch and api_key:
            try:
                batches = [ai_batch[i:i+40] for i in range(0, len(ai_batch), 40)]
                for batch in batches:
                    items = "\n".join([f"{i}: {e.get('supplier','')} | {e.get('value_gross',0)}€ | {e.get('notes','')}" for i, e in enumerate(batch)])
                    prompt = f"""Categoriza estas despesas de uma empresa portuguesa de eletricidade (Obelisco Radical).
Para cada linha, responde APENAS com o número e a categoria, uma por linha.

Categorias possíveis:
- Combustível, Material, Fornecedor, Serviços, Comunicações, Rendas, Seguros
- Contabilidade/Advogado, Ferramentas, Viatura, Alimentação, Imposto/Taxa, Outros

Despesas:
{items}

Responde EXACTAMENTE no formato: NUMERO:CATEGORIA (uma por linha)"""

                    chat = LlmChat(
                        api_key=api_key,
                        session_id=f"exp-cat-{uuid.uuid4().hex[:8]}",
                        system_message="És um assistente de categorização de despesas empresariais portuguesas.",
                    ).with_model("openai", "gpt-4o-mini")
                    resp = await chat.send_message(UserMessage(text=prompt))

                    for line in resp.strip().split("\n"):
                        line = line.strip()
                        if ":" not in line:
                            continue
                        parts = line.split(":", 1)
                        try:
                            idx = int(parts[0].strip())
                            cat = parts[1].strip()
                            if 0 <= idx < len(batch) and cat in CATEGORIES:
                                eid = batch[idx].get("id")
                                new_type = _smart_type_from_supplier(batch[idx].get("supplier", ""), cat)
                                await db.expenses.update_one({"id": eid}, {"$set": {"category": cat, "type": new_type}})
                                ai_updated += 1
                        except (ValueError, IndexError):
                            continue
            except Exception as ex:
                logger.warning(f"AI expense categorization failed: {ex}")

        return {
            "total": len(all_exp),
            "updated_keywords": updated,
            "updated_ai": ai_updated,
            "unchanged": len(all_exp) - updated - ai_updated,
            "message": f"{updated + ai_updated} despesas recategorizadas ({updated} por palavras-chave, {ai_updated} por IA)",
        }

    return expenses_router
