from pathlib import Path
from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
import os
import logging
import uuid
import bcrypt
import jwt
import asyncio
from datetime import datetime, timezone, timedelta
from pydantic import BaseModel
from typing import List, Optional

# MongoDB
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# App
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# JWT Config
JWT_ALGORITHM = "HS256"

def get_jwt_secret():
    return os.environ["JWT_SECRET"]

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id, "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=60),
        "type": "access"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "refresh"
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request):
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Não autenticado")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Token inválido")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Utilizador não encontrado")
        return {
            "id": str(user["_id"]),
            "email": user["email"],
            "name": user["name"],
            "role": user.get("role", "user"),
            "module_permissions": user.get("module_permissions") or default_modules_for_role(user.get("role", "consulta")),
        }
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")
    except Exception:
        raise HTTPException(status_code=401, detail="Erro de autenticação")


# --- Models ---

class LoginInput(BaseModel):
    email: str
    password: str

class BudgetItemModel(BaseModel):
    category: str = ""
    name: str = ""
    unit: str = ""
    quantity: float = 1
    unit_cost: float = 0
    margin: float = 0.6
    discount_type: str = "percentage"  # "percentage" or "value"
    discount_value: float = 0

class BudgetCreate(BaseModel):
    title: str
    client_name: str
    client_phone: str = ""
    items: List[BudgetItemModel] = []
    discount_type: str = "percentage"  # "percentage" or "value"
    discount_value: float = 0
    payment_methods: List[str] = []
    payment_split: str = ""
    payment_notes: str = ""

class BudgetUpdate(BaseModel):
    title: Optional[str] = None
    client_name: Optional[str] = None
    client_phone: Optional[str] = None
    items: Optional[List[BudgetItemModel]] = None
    status: Optional[str] = None
    discount_type: Optional[str] = None
    discount_value: Optional[float] = None
    payment_methods: Optional[List[str]] = None
    payment_split: Optional[str] = None
    payment_notes: Optional[str] = None

class StatusUpdate(BaseModel):
    status: str

class WorkCreate(BaseModel):
    title: str
    client_name: str
    client_phone: str = ""
    budget_id: str = ""
    proposal_id: str = ""
    status: str = "orcamento"
    predicted_cost: float = 0
    real_cost: float = 0
    notes: str = ""
    start_date: str = ""
    end_date: str = ""

class WorkUpdate(BaseModel):
    title: Optional[str] = None
    client_name: Optional[str] = None
    client_phone: Optional[str] = None
    status: Optional[str] = None
    predicted_cost: Optional[float] = None
    real_cost: Optional[float] = None
    notes: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None

class AppointmentCreate(BaseModel):
    title: str
    client_name: str = ""
    date: str
    time_start: str
    time_end: str
    notes: str = ""
    employee_ids: List[str] = []      # Técnicos atribuídos a este compromisso
    location: str = ""
    work_id: Optional[str] = None     # Obra associada (opcional)


# --- Auth Endpoints ---

@api_router.post("/auth/login")
async def login(input: LoginInput, response: Response):
    email = input.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(input.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email ou password incorretos")

    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)

    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=3600, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=True, samesite="none", max_age=604800, path="/")

    return {
        "id": user_id,
        "email": user["email"],
        "name": user["name"],
        "role": user.get("role", "user"),
        "module_permissions": user.get("module_permissions") or default_modules_for_role(user.get("role", "consulta")),
        "access_token": access_token,
        "refresh_token": refresh_token,
    }

@api_router.get("/auth/me")
async def get_me(user=Depends(get_current_user)):
    return user

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logout com sucesso"}

@api_router.post("/auth/refresh")
async def refresh(request: Request, response: Response):
    # Accept refresh token from cookie OR body (for iframe contexts where cookies are blocked)
    token = request.cookies.get("refresh_token")
    if not token:
        try:
            body = await request.json()
            token = body.get("refresh_token") if isinstance(body, dict) else None
        except Exception:
            token = None
    if not token:
        raise HTTPException(status_code=401, detail="Sem refresh token")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Token inválido")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="Utilizador não encontrado")
        user_id = str(user["_id"])
        access_token = create_access_token(user_id, user["email"])
        response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=3600, path="/")
        return {
            "id": user_id,
            "email": user["email"],
            "name": user["name"],
            "role": user.get("role", "user"),
            "access_token": access_token,
        }
    except Exception:
        raise HTTPException(status_code=401, detail="Refresh token inválido")


# --- Budget Endpoints ---

def calc_budget_totals(items, discount_type="percentage", discount_value=0):
    total_cost = sum(i.get("unit_cost", 0) * i.get("quantity", 0) for i in items)
    # Price per item = unit_cost * (1 + margin) * qty, then apply per-item discount
    total_price = 0
    for i in items:
        line = i.get("unit_cost", 0) * (1 + i.get("margin", 0)) * i.get("quantity", 0)
        dtype = i.get("discount_type", "percentage")
        dval = i.get("discount_value", 0) or 0
        if dval > 0:
            if dtype == "percentage":
                line = line * (1 - dval / 100)
            else:
                line = max(0, line - dval)
        total_price += line
    # Apply global discount
    if discount_value and discount_value > 0:
        if discount_type == "percentage":
            total_price = total_price * (1 - discount_value / 100)
        else:
            total_price = max(0, total_price - discount_value)
    return round(total_cost, 2), round(total_price, 2)

@api_router.get("/budgets")
async def get_budgets(user=Depends(get_current_user)):
    budgets = await db.budgets.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return budgets

@api_router.post("/budgets")
async def create_budget(input: BudgetCreate, user=Depends(get_current_user)):
    items = [item.model_dump() for item in input.items]
    total_cost, total_price = calc_budget_totals(items, input.discount_type, input.discount_value)
    doc = {
        "id": str(uuid.uuid4()),
        "title": input.title,
        "client_name": input.client_name,
        "client_phone": input.client_phone,
        "items": items,
        "discount_type": input.discount_type,
        "discount_value": input.discount_value,
        "payment_methods": input.payment_methods,
        "payment_split": input.payment_split,
        "payment_notes": input.payment_notes,
        "total_cost": total_cost,
        "total_price": total_price,
        "status": "rascunho",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    await db.budgets.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/budgets/{budget_id}")
async def get_budget(budget_id: str, user=Depends(get_current_user)):
    budget = await db.budgets.find_one({"id": budget_id}, {"_id": 0})
    if not budget:
        raise HTTPException(status_code=404, detail="Orçamento não encontrado")
    return budget

@api_router.put("/budgets/{budget_id}")
async def update_budget(budget_id: str, input: BudgetUpdate, user=Depends(get_current_user)):
    update_data = {}
    input_dict = input.model_dump(exclude_none=True)
    for k, v in input_dict.items():
        if k == "items":
            update_data["items"] = [i.model_dump() if hasattr(i, 'model_dump') else i for i in v]
        else:
            update_data[k] = v

    if "items" in update_data:
        # Merge with existing budget to get latest discount values
        existing = await db.budgets.find_one({"id": budget_id}, {"_id": 0})
        dtype = update_data.get("discount_type", existing.get("discount_type", "percentage") if existing else "percentage")
        dval = update_data.get("discount_value", existing.get("discount_value", 0) if existing else 0)
        total_cost, total_price = calc_budget_totals(update_data["items"], dtype, dval)
        update_data["total_cost"] = total_cost
        update_data["total_price"] = total_price
    elif "discount_type" in update_data or "discount_value" in update_data:
        existing = await db.budgets.find_one({"id": budget_id}, {"_id": 0})
        if existing:
            dtype = update_data.get("discount_type", existing.get("discount_type", "percentage"))
            dval = update_data.get("discount_value", existing.get("discount_value", 0))
            total_cost, total_price = calc_budget_totals(existing.get("items", []), dtype, dval)
            update_data["total_cost"] = total_cost
            update_data["total_price"] = total_price

    if not update_data:
        raise HTTPException(status_code=400, detail="Nada para atualizar")

    result = await db.budgets.update_one({"id": budget_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Orçamento não encontrado")
    updated = await db.budgets.find_one({"id": budget_id}, {"_id": 0})
    return updated

@api_router.delete("/budgets/{budget_id}")
async def delete_budget(budget_id: str, user=Depends(get_current_user)):
    result = await db.budgets.delete_one({"id": budget_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Orçamento não encontrado")
    return {"message": "Orcamento eliminado"}


# --- Proposal Endpoints ---

@api_router.post("/budgets/{budget_id}/generate-proposals")
async def generate_proposals(budget_id: str, user=Depends(get_current_user)):
    budget = await db.budgets.find_one({"id": budget_id}, {"_id": 0})
    if not budget:
        raise HTTPException(status_code=404, detail="Orçamento não encontrado")

    await db.proposals.delete_many({"budget_id": budget_id})

    base_price = budget["total_price"]
    # Tier name stays internal (for sorting). Client-facing title/description removed tier mention.
    tiers = [
        {"tier": "basico", "label": "Básico", "multiplier": 1.0},
        {"tier": "profissional", "label": "Profissional", "multiplier": 1.15},
        {"tier": "premium", "label": "Premium", "multiplier": 1.30},
    ]
    client_description = "Proposta de serviços elétricos e de telecomunicações. Garantia de 2 anos sobre mão de obra e materiais fornecidos. Valores em euros, IVA não incluído."

    proposals = []
    for t in tiers:
        prop = {
            "id": str(uuid.uuid4()),
            "budget_id": budget_id,
            "tier": t["tier"],
            "label": t["label"],
            "title": budget['title'],
            "client_name": budget["client_name"],
            "client_phone": budget.get("client_phone", ""),
            "items": budget["items"],
            "base_value": round(base_price, 2),
            "multiplier": t["multiplier"],
            "final_value": round(base_price * t["multiplier"], 2),
            "description": client_description,
            "status": "pendente",
            "payment_methods": budget.get("payment_methods", []),
            "payment_split": budget.get("payment_split", ""),
            "payment_notes": budget.get("payment_notes", ""),
            "discount_type": budget.get("discount_type", "percentage"),
            "discount_value": budget.get("discount_value", 0),
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.proposals.insert_one(prop)
        prop.pop("_id", None)
        proposals.append(prop)

    await db.budgets.update_one({"id": budget_id}, {"$set": {"status": "proposta_gerada"}})
    return proposals

@api_router.get("/proposals")
async def get_proposals(user=Depends(get_current_user)):
    proposals = await db.proposals.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return proposals

@api_router.get("/proposals/{proposal_id}")
async def get_proposal(proposal_id: str, user=Depends(get_current_user)):
    proposal = await db.proposals.find_one({"id": proposal_id}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposta não encontrada")
    return proposal

@api_router.put("/proposals/{proposal_id}/status")
async def update_proposal_status(proposal_id: str, input: StatusUpdate, user=Depends(get_current_user)):
    result = await db.proposals.update_one({"id": proposal_id}, {"$set": {"status": input.status}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Proposta não encontrada")
    updated = await db.proposals.find_one({"id": proposal_id}, {"_id": 0})
    return updated

@api_router.delete("/proposals/{proposal_id}")
async def delete_proposal(proposal_id: str, user=Depends(get_current_user)):
    result = await db.proposals.delete_one({"id": proposal_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Proposta não encontrada")
    return {"message": "Proposta eliminada"}


# --- Signature Endpoints ---

@api_router.post("/proposals/{proposal_id}/sign-link")
async def create_sign_link(proposal_id: str, user=Depends(get_current_user)):
    proposal = await db.proposals.find_one({"id": proposal_id}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposta não encontrada")
    # Generate or reuse existing token
    token = proposal.get("sign_token")
    if not token:
        token = uuid.uuid4().hex
        await db.proposals.update_one(
            {"id": proposal_id},
            {"$set": {
                "sign_token": token,
                "sign_status": proposal.get("sign_status") or "pending",
                "sign_link_created_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
    return {"token": token, "sign_status": proposal.get("sign_status") or "pending"}


# Public router (no auth) - mounted separately
public_router = APIRouter(prefix="/api/public", tags=["public"])


@public_router.get("/proposal/{token}")
async def public_get_proposal(token: str):
    proposal = await db.proposals.find_one({"sign_token": token}, {"_id": 0, "sign_token": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Link de proposta inválido ou expirado")
    # Load settings for display (payment terms, etc)
    settings_doc = await db.proposal_settings.find_one({}, {"_id": 0}) or {}
    return {"proposal": proposal, "settings": settings_doc}


class SignatureInput(BaseModel):
    signature_data: str  # base64 PNG data URL
    signed_by_name: str
    signed_by_email: str = ""


@public_router.post("/proposal/{token}/sign")
async def public_sign_proposal(token: str, input: SignatureInput, request: Request):
    proposal = await db.proposals.find_one({"sign_token": token}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Link de proposta inválido ou expirado")
    if proposal.get("sign_status") == "signed":
        raise HTTPException(status_code=400, detail="Proposta já foi assinada")
    if not input.signature_data or not input.signature_data.startswith("data:image"):
        raise HTTPException(status_code=400, detail="Assinatura inválida")
    if not input.signed_by_name or len(input.signed_by_name.strip()) < 3:
        raise HTTPException(status_code=400, detail="Nome de quem assina é obrigatório")

    ip = request.client.host if request.client else ""
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.proposals.update_one(
        {"sign_token": token},
        {"$set": {
            "sign_status": "signed",
            "signature_data": input.signature_data,
            "signed_by_name": input.signed_by_name.strip(),
            "signed_by_email": input.signed_by_email.strip(),
            "signed_by_ip": ip,
            "signed_at": now_iso,
            "status": "aceite",
        }}
    )
    return {"ok": True, "signed_at": now_iso}


# --- Works Endpoints ---

@api_router.get("/pipeline")
async def get_pipeline(user=Depends(get_current_user)):
    """Consolida ciclo de vida completo por obra: orçamento → proposta → aceite → execução → guias → fatura → pago → concluída."""
    from datetime import datetime as _dt

    budgets = await db.budgets.find({}, {"_id": 0}).to_list(2000)
    proposals = await db.proposals.find({}, {"_id": 0}).to_list(5000)
    works = await db.works.find({}, {"_id": 0}).to_list(2000)
    guides = await db.transport_guides.find({}, {"_id": 0}).to_list(5000)
    invoices = await db.invoices.find({}, {"_id": 0}).to_list(5000)

    # Indexes for fast lookup
    proposals_by_budget = {}
    for p in proposals:
        proposals_by_budget.setdefault(p.get("budget_id"), []).append(p)

    guides_by_work = {}
    for g in guides:
        wid = g.get("work_id")
        if wid:
            guides_by_work.setdefault(wid, []).append(g)

    invoices_by_client = {}
    for i in invoices:
        cn = (i.get("client_name") or "").strip().lower()
        if cn:
            invoices_by_client.setdefault(cn, []).append(i)

    now = _dt.now(timezone.utc)

    def _pick_best_proposal(bid: str):
        """Escolhe a proposta mais avançada de um orçamento: signed > accepted > sent > draft."""
        arr = proposals_by_budget.get(bid, [])
        if not arr:
            return None
        rank = {"signed": 4, "accepted": 3, "aceite": 3, "sent": 2, "enviada": 2, "draft": 1, "rascunho": 1}
        best = None
        best_score = -1
        for p in arr:
            score = 0
            if p.get("sign_status") == "signed" or p.get("signed_at"):
                score = 4
            else:
                score = rank.get((p.get("status") or "").lower(), 1)
            if score > best_score:
                best = p
                best_score = score
        return best

    def _find_invoices_for(work, budget):
        """Encontra faturas relacionadas com esta obra por cliente + valor aproximado, ou por proposal_id se existir."""
        cn = (work.get("client_name") or budget.get("client_name") or "").strip().lower()
        if not cn:
            return []
        candidates = invoices_by_client.get(cn, [])
        return candidates

    pipeline_items = []
    handled_budget_ids = set()

    # 1) Iterar por obras (têm sempre o ciclo mais completo)
    for w in works:
        bid = w.get("budget_id") or ""
        handled_budget_ids.add(bid)
        budget = next((b for b in budgets if b.get("id") == bid), None) or {}
        prop = _pick_best_proposal(bid) if bid else None
        w_guides = guides_by_work.get(w["id"], [])
        w_invoices = _find_invoices_for(w, budget)

        # Milestones
        milestones = _build_milestones(budget, prop, w, w_guides, w_invoices)
        phase = _determine_phase(milestones, w)

        # Valor total (venda) e valor recebido
        sale_value = float((budget or {}).get("total_final", 0) or w.get("predicted_cost", 0) or 0)
        invoiced = sum(float(i.get("value_total", 0) or 0) for i in w_invoices)
        received = 0.0
        for i in w_invoices:
            for p in (i.get("payments") or []):
                received += float(p.get("amount", 0) or 0)
        pending = max(0.0, invoiced - received)

        pipeline_items.append({
            "id": w["id"],
            "kind": "work",
            "title": w.get("title") or budget.get("title") or "Obra sem título",
            "client_name": w.get("client_name") or budget.get("client_name") or "",
            "phase": phase,
            "milestones": milestones,
            "completed_count": sum(1 for m in milestones if m["done"]),
            "total_count": len(milestones),
            "sale_value": round(sale_value, 2),
            "invoiced_value": round(invoiced, 2),
            "received_value": round(received, 2),
            "pending_value": round(pending, 2),
            "budget_id": bid,
            "proposal_id": (prop or {}).get("id"),
            "invoices_count": len(w_invoices),
            "guides_count": len(w_guides),
            "created_at": w.get("created_at"),
            "updated_at": w.get("emitted_at") or w.get("created_at"),
        })

    # 2) Orçamentos sem obra (ainda em fase inicial)
    for b in budgets:
        if b.get("id") in handled_budget_ids:
            continue
        prop = _pick_best_proposal(b["id"])
        milestones = _build_milestones(b, prop, None, [], [])
        phase = _determine_phase(milestones, None)
        pipeline_items.append({
            "id": f"budget-{b['id']}",
            "kind": "budget",
            "title": b.get("title") or "Orçamento sem título",
            "client_name": b.get("client_name") or "",
            "phase": phase,
            "milestones": milestones,
            "completed_count": sum(1 for m in milestones if m["done"]),
            "total_count": len(milestones),
            "sale_value": round(float(b.get("total_final", 0) or 0), 2),
            "invoiced_value": 0.0,
            "received_value": 0.0,
            "pending_value": 0.0,
            "budget_id": b["id"],
            "proposal_id": (prop or {}).get("id"),
            "invoices_count": 0,
            "guides_count": 0,
            "created_at": b.get("created_at"),
            "updated_at": (prop or {}).get("created_at") or b.get("created_at"),
        })

    # Agrupar por fase
    phase_order = ["orcamento", "proposta_enviada", "aceite", "em_execucao", "guias_emitidas", "faturada", "paga", "concluida"]
    by_phase = {p: [] for p in phase_order}
    for it in pipeline_items:
        by_phase.setdefault(it["phase"], []).append(it)

    # Ordenar cada coluna por updated_at desc
    for phase in by_phase:
        by_phase[phase].sort(key=lambda x: x.get("updated_at") or "", reverse=True)

    # KPIs
    total_pending = sum(it["pending_value"] for it in pipeline_items)
    total_sale = sum(it["sale_value"] for it in pipeline_items)
    total_received = sum(it["received_value"] for it in pipeline_items)

    # Obras atrasadas: em execução há > 60 dias e sem fatura
    overdue = 0
    for it in pipeline_items:
        if it["phase"] in ("em_execucao", "guias_emitidas"):
            try:
                d = _dt.fromisoformat(it["updated_at"].replace("Z", "+00:00")) if it.get("updated_at") else None
                if d and (now - d).days > 60:
                    overdue += 1
            except Exception:
                pass

    counts_by_phase = {p: len(by_phase.get(p, [])) for p in phase_order}

    return {
        "phases": phase_order,
        "items": pipeline_items,
        "by_phase": by_phase,
        "kpis": {
            "total_items": len(pipeline_items),
            "total_sale_value": round(total_sale, 2),
            "total_pending_value": round(total_pending, 2),
            "total_received_value": round(total_received, 2),
            "overdue_count": overdue,
            "counts_by_phase": counts_by_phase,
        },
    }


def _build_milestones(budget, proposal, work, guides, invoices):
    """Constrói lista de 8 marcos com {key, label, done, at, meta}."""
    ms = []
    # 1) Orçamento criado
    ms.append({
        "key": "budget_created", "label": "Orçamento criado",
        "done": bool(budget and budget.get("id")),
        "at": (budget or {}).get("created_at"),
        "meta": None,
    })
    # 2) Proposta enviada
    prop_created = proposal and proposal.get("created_at")
    ms.append({
        "key": "proposal_sent", "label": "Proposta enviada",
        "done": bool(prop_created),
        "at": prop_created,
        "meta": (proposal or {}).get("label"),
    })
    # 3) Assinada/aceite
    signed = (proposal or {}).get("signed_at") or ((proposal or {}).get("sign_status") == "signed")
    accepted = (proposal or {}).get("status") in ("accepted", "aceite", "signed")
    ms.append({
        "key": "proposal_accepted", "label": "Proposta aceite/assinada",
        "done": bool(signed or accepted),
        "at": (proposal or {}).get("signed_at") or (proposal or {}).get("accepted_at"),
        "meta": "assinada digital" if signed else ("aceite" if accepted else None),
    })
    # 4) Obra em execução
    in_exec = bool(work and work.get("status") in ("em_execucao", "em_execução"))
    ms.append({
        "key": "work_in_progress", "label": "Obra em execução",
        "done": in_exec or bool(work and work.get("start_date")),
        "at": (work or {}).get("start_date") or (work or {}).get("created_at") if work else None,
        "meta": None,
    })
    # 5) Guias emitidas
    emitted_guides = [g for g in guides if g.get("status") in ("emitida", "em_transito", "recebida", "recebida_com_diferencas")]
    ms.append({
        "key": "guides_emitted", "label": "Guias emitidas",
        "done": len(emitted_guides) > 0,
        "at": max((g.get("emitted_at") or "" for g in emitted_guides), default=None) or None,
        "meta": f"{len(emitted_guides)}/{len(guides)}" if guides else None,
    })
    # 6) Fatura emitida
    ms.append({
        "key": "invoice_emitted", "label": "Fatura emitida",
        "done": len(invoices) > 0,
        "at": min((i.get("issue_date") or "" for i in invoices), default=None) or None,
        "meta": f"{len(invoices)} fatura(s)" if invoices else None,
    })
    # 7) Fatura paga (todas)
    all_paid = False
    if invoices:
        all_paid = all(
            (float(i.get("value_total", 0) or 0) - sum(float(p.get("amount", 0) or 0) for p in (i.get("payments") or []))) <= 0.01
            for i in invoices
        )
    ms.append({
        "key": "invoice_paid", "label": "Fatura paga",
        "done": bool(all_paid and invoices),
        "at": max((max((p.get("date") or "" for p in (i.get("payments") or [])), default="") for i in invoices), default=None) or None if all_paid else None,
        "meta": None,
    })
    # 8) Obra concluída
    concluded = bool(work and work.get("status") == "finalizado")
    ms.append({
        "key": "work_completed", "label": "Obra concluída",
        "done": concluded,
        "at": (work or {}).get("end_date") if concluded else None,
        "meta": None,
    })
    return ms


def _determine_phase(milestones, work):
    """Retorna a fase actual baseada nos marcos concluídos."""
    done_keys = {m["key"] for m in milestones if m["done"]}
    if "work_completed" in done_keys or (work and work.get("status") == "finalizado"):
        return "concluida"
    if "invoice_paid" in done_keys:
        return "paga"
    if "invoice_emitted" in done_keys:
        return "faturada"
    if "guides_emitted" in done_keys:
        return "guias_emitidas"
    if "work_in_progress" in done_keys:
        return "em_execucao"
    if "proposal_accepted" in done_keys:
        return "aceite"
    if "proposal_sent" in done_keys:
        return "proposta_enviada"
    return "orcamento"


@api_router.get("/works")
async def get_works(user=Depends(get_current_user)):
    works = await db.works.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return works

@api_router.post("/works")
async def create_work(input: WorkCreate, user=Depends(get_current_user)):
    doc = {
        "id": str(uuid.uuid4()),
        **input.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    await db.works.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/works/{work_id}")
async def get_work(work_id: str, user=Depends(get_current_user)):
    work = await db.works.find_one({"id": work_id}, {"_id": 0})
    if not work:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    return work

@api_router.put("/works/{work_id}")
async def update_work(work_id: str, input: WorkUpdate, user=Depends(get_current_user)):
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    result = await db.works.update_one({"id": work_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    updated = await db.works.find_one({"id": work_id}, {"_id": 0})
    return updated

@api_router.delete("/works/{work_id}")
async def delete_work(work_id: str, user=Depends(get_current_user)):
    result = await db.works.delete_one({"id": work_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    return {"message": "Obra eliminada"}

@api_router.post("/works/from-proposal/{proposal_id}")
async def create_work_from_proposal(proposal_id: str, user=Depends(get_current_user)):
    proposal = await db.proposals.find_one({"id": proposal_id}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposta não encontrada")
    budget_id = proposal.get("budget_id", "")
    budget = await db.budgets.find_one({"id": budget_id}, {"_id": 0}) if budget_id else None

    # Popular items da obra a partir do orçamento — para que sale_total, predicted, margens
    # apareçam de imediato na Caixa da Obra sem exigir sync manual.
    initial_items = []
    if budget:
        for idx, b_it in enumerate(budget.get("items", [])):
            initial_items.append({
                "id": str(uuid.uuid4()),
                "budget_item_idx": idx,
                "category": b_it.get("category", ""),
                "name": b_it.get("name", ""),
                "unit": b_it.get("unit", "un"),
                "quantity": b_it.get("quantity", 1),
                "predicted_unit_cost": b_it.get("unit_cost", 0),
                "margin": b_it.get("margin", 0.6),
                "real_unit_cost": 0,
                "real_quantity": None,
                "real_notes": "",
                "history": [],
                "is_extra": False,
            })

    doc = {
        "id": str(uuid.uuid4()),
        "title": proposal["title"],
        "client_name": proposal["client_name"],
        "client_phone": proposal.get("client_phone", ""),
        "budget_id": budget_id,
        "proposal_id": proposal_id,
        "status": "orcamento",
        "predicted_cost": proposal["final_value"],
        "real_cost": 0,
        "notes": f"Criada a partir da proposta {proposal['label']}",
        "start_date": "",
        "end_date": "",
        "items": initial_items,
        "items_synced_at": datetime.now(timezone.utc).isoformat() if initial_items else None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"]
    }
    await db.works.insert_one(doc)
    doc.pop("_id", None)
    return doc


# --- Work Items: Real vs Predicted Cost Tracking ---

class WorkItemUpdate(BaseModel):
    real_unit_cost: Optional[float] = None
    real_quantity: Optional[float] = None
    real_notes: Optional[str] = None
    execution_status: Optional[str] = None    # "pending" | "in_progress" | "done"
    executed_quantity: Optional[float] = None  # quantidade já executada
    execution_notes: Optional[str] = None


class WorkItemExtra(BaseModel):
    name: str
    category: str = "Extra"
    unit: str = "un"
    quantity: float = 1
    predicted_unit_cost: float = 0
    margin: float = 0.6
    real_unit_cost: float = 0
    real_quantity: Optional[float] = None
    notes: str = ""


def _compute_work_item_totals(it: dict) -> dict:
    """Calcula custos previsto vs real + preço de venda + margem por linha."""
    qty = float(it.get("quantity") or 0)
    real_qty = it.get("real_quantity")
    real_qty = float(real_qty) if real_qty is not None else qty
    pred_uc = float(it.get("predicted_unit_cost") or 0)
    real_uc = float(it.get("real_unit_cost") or 0)
    margin = float(it.get("margin") or 0)
    sale_unit = round(pred_uc * (1 + margin), 2)         # preço cobrado ao cliente
    return {
        **it,
        "real_quantity": real_qty,
        "predicted_total": round(pred_uc * qty, 2),
        "real_total": round(real_uc * real_qty, 2) if real_uc > 0 else 0,
        "sale_unit_price": sale_unit,
        "sale_total": round(sale_unit * qty, 2),
        "delta": round((real_uc * real_qty) - (pred_uc * qty), 2) if real_uc > 0 else 0,
    }


@api_router.post("/works/{work_id}/sync-budget")
async def sync_work_from_budget(work_id: str, user=Depends(get_current_user)):
    """Carrega/actualiza os itens do orçamento de origem para a obra.
    Items que já tinham custo real preenchido na obra são preservados."""
    work = await db.works.find_one({"id": work_id}, {"_id": 0})
    if not work:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    if not work.get("budget_id"):
        raise HTTPException(status_code=400, detail="Esta obra não tem orçamento associado")
    budget = await db.budgets.find_one({"id": work["budget_id"]}, {"_id": 0})
    if not budget:
        raise HTTPException(status_code=404, detail="Orçamento de origem não encontrado")

    existing = {it.get("budget_item_idx"): it for it in (work.get("items") or [])}
    new_items = []
    for idx, b_it in enumerate(budget.get("items", [])):
        prev = existing.get(idx) or {}
        item = {
            "id": prev.get("id") or str(uuid.uuid4()),
            "budget_item_idx": idx,
            "category": b_it.get("category", ""),
            "name": b_it.get("name", ""),
            "unit": b_it.get("unit", "un"),
            "quantity": b_it.get("quantity", 1),
            "predicted_unit_cost": b_it.get("unit_cost", 0),
            "margin": b_it.get("margin", 0.6),
            "real_unit_cost": prev.get("real_unit_cost", 0),
            "real_quantity": prev.get("real_quantity"),
            "real_notes": prev.get("real_notes", ""),
            "history": prev.get("history", []),
            "is_extra": False,
        }
        new_items.append(item)
    # Preserve extras
    for it in (work.get("items") or []):
        if it.get("is_extra"):
            new_items.append(it)

    await db.works.update_one({"id": work_id}, {"$set": {"items": new_items, "items_synced_at": datetime.now(timezone.utc).isoformat()}})
    return await get_work_full(work_id, user)


@api_router.get("/works/{work_id}/caixa")
async def get_work_caixa(work_id: str, user=Depends(get_current_user)):
    """Retorna todos os KPIs financeiros e de execução da caixa da obra."""

    async def _work_attendance_today(wid: str):
        """Lista os técnicos que picaram nesta obra HOJE (in/out)."""
        today_iso = datetime.now(timezone.utc).date().isoformat()
        atts = await db.attendance.find({"date": today_iso}, {"_id": 0}).to_list(200)
        result = []
        for att in atts:
            work_punches = [p for p in (att.get("punches") or []) if p.get("work_id") == wid]
            if not work_punches:
                continue
            emp = await db.employees.find_one({"id": att.get("employee_id")}, {"_id": 0, "name": 1})
            last = work_punches[-1]
            result.append({
                "employee_id": att.get("employee_id"),
                "employee_name": (emp or {}).get("name", "Técnico"),
                "current_status": "in" if last.get("action") == "in" else "out",
                "punches": work_punches,
            })
        return result
    """
    Caixa da Obra — balanço financeiro completo:
    - Valor de venda (do orçamento aprovado)
    - Facturas emitidas + pagamentos recebidos + saldo em dívida
    - Despesas pagas / a pagar
    - Margem prevista vs real, cashflow líquido
    """
    work = await db.works.find_one({"id": work_id}, {"_id": 0})
    if not work:
        raise HTTPException(status_code=404, detail="Obra não encontrada")

    # Auto-sync: se a obra foi criada antes deste ajuste e não tem items,
    # mas tem budget_id, carrega os items do orçamento uma única vez.
    if not (work.get("items") or []) and work.get("budget_id"):
        budget = await db.budgets.find_one({"id": work["budget_id"]}, {"_id": 0})
        if budget and budget.get("items"):
            new_items = []
            for idx, b_it in enumerate(budget.get("items", [])):
                new_items.append({
                    "id": str(uuid.uuid4()),
                    "budget_item_idx": idx,
                    "category": b_it.get("category", ""),
                    "name": b_it.get("name", ""),
                    "unit": b_it.get("unit", "un"),
                    "quantity": b_it.get("quantity", 1),
                    "predicted_unit_cost": b_it.get("unit_cost", 0),
                    "margin": b_it.get("margin", 0.6),
                    "real_unit_cost": 0,
                    "real_quantity": None,
                    "real_notes": "",
                    "history": [],
                    "is_extra": False,
                })
            await db.works.update_one(
                {"id": work_id},
                {"$set": {"items": new_items, "items_synced_at": datetime.now(timezone.utc).isoformat()}},
            )
            work["items"] = new_items

    # Items para calcular venda + previsto + real
    items = [_compute_work_item_totals(it) for it in (work.get("items") or [])]
    sale_total = round(sum(it.get("sale_total", 0) for it in items), 2)
    predicted_total = round(sum(it.get("predicted_total", 0) for it in items), 2)
    real_items_total = round(sum(it.get("real_total", 0) for it in items), 2)

    # --- Execução por item (peso = sale_total de cada linha) ---
    execution_done = 0.0
    execution_in_progress = 0.0
    execution_pending = 0.0
    items_done_count = 0
    items_in_progress_count = 0
    items_pending_count = 0
    exec_weighted_pct = 0.0
    raw_items = work.get("items") or []
    for computed, raw in zip(items, raw_items):
        st = raw.get("execution_status") or "pending"
        qty_total = float(raw.get("quantity") or 0)
        exec_qty = float(raw.get("executed_quantity") or 0)
        line_sale = computed.get("sale_total", 0)
        if st == "done":
            items_done_count += 1
            execution_done += line_sale
            item_pct = 1.0
        elif st == "in_progress":
            items_in_progress_count += 1
            item_pct = (exec_qty / qty_total) if qty_total > 0 else 0
            execution_in_progress += line_sale * item_pct
        else:
            items_pending_count += 1
            item_pct = 0
        if sale_total > 0:
            exec_weighted_pct += (line_sale / sale_total) * item_pct
    execution_pct = round(exec_weighted_pct * 100, 1)
    executed_value = round(execution_done + execution_in_progress, 2)

    # Facturas emitidas para esta obra
    invoices = await db.invoices.find({"obra_id": work_id}, {"_id": 0}).to_list(500)
    total_invoiced = round(sum(float(i.get("value_total") or 0) for i in invoices), 2)
    total_received = 0.0
    for inv in invoices:
        for p in (inv.get("payments") or []):
            total_received += float(p.get("amount") or 0)
    total_received = round(total_received, 2)
    to_receive = round(total_invoiced - total_received, 2)  # em dívida
    to_invoice = round(max(0, sale_total - total_invoiced), 2)  # ainda por facturar

    # Despesas
    expenses = await db.expenses.find({"obra_id": work_id}, {"_id": 0}).to_list(2000)
    expenses_total = round(sum(float(e.get("value_gross") or 0) for e in expenses), 2)
    expenses_paid = round(sum(float(e.get("value_gross") or 0) for e in expenses if e.get("paid")), 2)
    expenses_to_pay = round(expenses_total - expenses_paid, 2)

    # Custo real total (items + expenses)
    real_total_cost = round(real_items_total + expenses_total, 2)

    # Margens
    predicted_profit = round(sale_total - predicted_total, 2)
    real_profit = round(sale_total - real_total_cost, 2)
    margin_predicted_pct = round((predicted_profit / sale_total * 100) if sale_total > 0 else 0, 1)
    margin_real_pct = round((real_profit / sale_total * 100) if sale_total > 0 else 0, 1)

    # Cashflow — o que está de facto na caixa desta obra
    cash_balance = round(total_received - expenses_paid, 2)  # dinheiro efectivo
    projected_cash_balance = round(sale_total - real_total_cost, 2)  # se tudo for cobrado/pago

    # ---- ALERTAS por obra ----
    today = datetime.now(timezone.utc).date()
    alerts = []

    # (a) Margem real muito abaixo da prevista
    if margin_predicted_pct > 0 and margin_real_pct < margin_predicted_pct * 0.7:
        gap = round(margin_predicted_pct - margin_real_pct, 1)
        alerts.append({
            "code": "margem_baixa",
            "severity": "high" if margin_real_pct < margin_predicted_pct * 0.5 else "medium",
            "title": "Margem real abaixo do previsto",
            "message": f"Margem real {margin_real_pct}% vs {margin_predicted_pct}% prevista ({gap}pp abaixo).",
            "meta": {"margin_real_pct": margin_real_pct, "margin_predicted_pct": margin_predicted_pct},
        })

    # (b) Custo real ultrapassa o previsto
    if predicted_total > 0 and real_total_cost > predicted_total:
        overrun_pct = round((real_total_cost / predicted_total - 1) * 100, 1)
        alerts.append({
            "code": "custo_excedido",
            "severity": "high" if overrun_pct > 20 else "medium",
            "title": "Custo real acima do previsto",
            "message": f"Custo real {round(real_total_cost, 2)}€ excede o previsto em {overrun_pct}% (+{round(real_total_cost - predicted_total, 2)}€).",
            "meta": {"real_total_cost": real_total_cost, "predicted_total": predicted_total, "overrun_pct": overrun_pct},
        })

    # (c) Facturas vencidas
    overdue_inv = []
    for inv in invoices:
        total_i = float(inv.get("value_total") or 0)
        paid_i = sum(float(p.get("amount") or 0) for p in (inv.get("payments") or []))
        balance_i = round(total_i - paid_i, 2)
        due = inv.get("due_date") or ""
        if balance_i > 0.01 and due:
            try:
                due_d = datetime.fromisoformat(due).date() if "T" in due else datetime.strptime(due[:10], "%Y-%m-%d").date()
                if due_d < today:
                    overdue_inv.append({
                        "id": inv.get("id"), "number": inv.get("number"),
                        "balance": balance_i, "days_overdue": (today - due_d).days,
                    })
            except Exception:
                pass
    if overdue_inv:
        total_owed = round(sum(x["balance"] for x in overdue_inv), 2)
        max_days = max(x["days_overdue"] for x in overdue_inv)
        alerts.append({
            "code": "faturas_vencidas",
            "severity": "high" if max_days > 30 else "medium",
            "title": f"{len(overdue_inv)} fatura(s) vencida(s)",
            "message": f"{total_owed}€ em dívida · atraso máx. {max_days} dia(s).",
            "meta": {"count": len(overdue_inv), "total_owed": total_owed, "max_days_overdue": max_days, "invoices": overdue_inv},
        })

    # (d) Despesas por pagar em atraso (>30 dias após data da despesa e não pagas)
    def _exp_paid(e):
        if e.get("paid") == True:
            return True
        st = (e.get("payment_status") or "").lower()
        return st in ("pago", "paid")

    overdue_exp = []
    for e in expenses:
        if _exp_paid(e):
            continue
        d = e.get("date") or ""
        if not d:
            continue
        try:
            ed = datetime.strptime(d[:10], "%Y-%m-%d").date()
            days_since = (today - ed).days
            if days_since > 30:
                overdue_exp.append({
                    "id": e.get("id"),
                    "supplier": e.get("supplier"),
                    "value_gross": float(e.get("value_gross") or 0),
                    "days_since": days_since,
                })
        except Exception:
            pass
    if overdue_exp:
        total_exp_owed = round(sum(x["value_gross"] for x in overdue_exp), 2)
        alerts.append({
            "code": "despesas_atraso",
            "severity": "medium",
            "title": f"{len(overdue_exp)} despesa(s) por pagar >30 dias",
            "message": f"{total_exp_owed}€ em despesas por regularizar.",
            "meta": {"count": len(overdue_exp), "total": total_exp_owed, "expenses": overdue_exp},
        })

    # (e) Obra em curso há >30 dias sem qualquer factura emitida
    status_l = (work.get("status") or "").lower()
    is_active = status_l not in ("finalizado", "concluida", "concluída", "cancelada", "cancelado", "")
    start = work.get("start_date") or ""
    if is_active and start and len(invoices) == 0:
        try:
            sd = datetime.strptime(start[:10], "%Y-%m-%d").date()
            days_running = (today - sd).days
            if days_running > 30:
                alerts.append({
                    "code": "sem_faturacao",
                    "severity": "medium",
                    "title": "Obra em curso sem faturação",
                    "message": f"Obra iniciada há {days_running} dias e ainda sem faturas emitidas.",
                    "meta": {"days_running": days_running, "start_date": start},
                })
        except Exception:
            pass

    # (f) Grande % em dívida do que foi facturado (>50%)
    if total_invoiced > 0:
        debt_ratio = to_receive / total_invoiced
        if debt_ratio > 0.5:
            alerts.append({
                "code": "recebimento_lento",
                "severity": "medium",
                "title": "Recebimento lento",
                "message": f"{round(debt_ratio * 100, 1)}% do valor facturado ainda por receber ({round(to_receive, 2)}€ de {round(total_invoiced, 2)}€).",
                "meta": {"debt_ratio_pct": round(debt_ratio * 100, 1), "to_receive": to_receive, "total_invoiced": total_invoiced},
            })

    return {
        "work": {
            "id": work.get("id"),
            "title": work.get("title"),
            "client_name": work.get("client_name"),
            "status": work.get("status"),
            "start_date": work.get("start_date"),
            "end_date": work.get("end_date"),
        },
        "resumo": {
            "sale_total": sale_total,
            "predicted_total": predicted_total,
            "real_total_cost": real_total_cost,
            "predicted_profit": predicted_profit,
            "real_profit": real_profit,
            "margin_predicted_pct": margin_predicted_pct,
            "margin_real_pct": margin_real_pct,
        },
        "receitas": {
            "total_invoiced": total_invoiced,
            "total_received": total_received,
            "to_receive": to_receive,
            "to_invoice": to_invoice,
            "invoices_count": len(invoices),
            "invoices": [{
                "id": i.get("id"), "number": i.get("number"), "issue_date": i.get("issue_date"),
                "due_date": i.get("due_date"), "value_total": i.get("value_total"),
                "paid_total": sum(float(p.get("amount") or 0) for p in (i.get("payments") or [])),
                "status": i.get("status"),
            } for i in invoices],
        },
        "despesas": {
            "expenses_total": expenses_total,
            "expenses_paid": expenses_paid,
            "expenses_to_pay": expenses_to_pay,
            "count": len(expenses),
            "expenses": [{
                "id": e.get("id"), "description": e.get("description"), "date": e.get("date"),
                "supplier": e.get("supplier"), "category": e.get("category"),
                "value_gross": e.get("value_gross"), "paid": e.get("paid", False),
                "type": e.get("type"),
            } for e in expenses],
        },
        "caixa": {
            "cash_balance": cash_balance,                     # dinheiro efectivo na caixa da obra
            "projected_cash_balance": projected_cash_balance, # o que sobra quando tudo cobrado/pago
            "receipts_progress_pct": round((total_received / sale_total * 100) if sale_total > 0 else 0, 1),
            "cost_progress_pct": round((real_total_cost / predicted_total * 100) if predicted_total > 0 else 0, 1),
        },
        "execution": {
            "pct": execution_pct,
            "executed_value": executed_value,
            "remaining_value": round(sale_total - executed_value, 2),
            "items_done": items_done_count,
            "items_in_progress": items_in_progress_count,
            "items_pending": items_pending_count,
            "items_total": len(raw_items),
        },
        "attendance": await _work_attendance_today(work_id),
        "alerts": alerts,
    }


class WorkLinkInput(BaseModel):
    obra_id: Optional[str] = None   # None => desassocia


def _require_finance_module(user, module_key: str):
    """Permite acesso se user for admin OU tiver a permissão do módulo."""
    if user.get("role") == "admin":
        return
    perms = user.get("module_permissions") or {}
    if not perms.get(module_key):
        raise HTTPException(status_code=403, detail=f"Sem permissão para gerir o módulo '{module_key}'.")


@api_router.put("/invoices/{invoice_id}/link-work")
async def link_invoice_to_work(invoice_id: str, input: WorkLinkInput, user=Depends(get_current_user)):
    """Associa (ou desassocia) uma fatura existente a uma obra. Requer permissão do módulo Faturas ou admin."""
    _require_finance_module(user, "faturas")
    invoice = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="Fatura não encontrada")

    new_obra = (input.obra_id or "").strip() or None
    if new_obra:
        work = await db.works.find_one({"id": new_obra}, {"_id": 0})
        if not work:
            raise HTTPException(status_code=404, detail="Obra não encontrada")

    await db.invoices.update_one({"id": invoice_id}, {"$set": {"obra_id": new_obra}})
    updated = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    return updated


@api_router.put("/expenses/{expense_id}/link-work")
async def link_expense_to_work(expense_id: str, input: WorkLinkInput, user=Depends(get_current_user)):
    """Associa (ou desassocia) uma despesa existente a uma obra. Requer permissão do módulo Despesas ou admin."""
    _require_finance_module(user, "despesas")
    expense = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="Despesa não encontrada")

    new_obra = (input.obra_id or "").strip() or None
    obra_name = None
    if new_obra:
        work = await db.works.find_one({"id": new_obra}, {"_id": 0})
        if not work:
            raise HTTPException(status_code=404, detail="Obra não encontrada")
        obra_name = work.get("title")
        # Se estiver a associar a obra, garantir que o tipo passa a "obra"
        expense_type = expense.get("type", "variavel")
        if expense_type != "obra":
            await db.expenses.update_one({"id": expense_id}, {"$set": {"type": "obra"}})

    await db.expenses.update_one(
        {"id": expense_id},
        {"$set": {"obra_id": new_obra, "obra_name": obra_name}},
    )
    updated = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    return updated


@api_router.get("/works/{work_id}/full")
async def get_work_full(work_id: str, user=Depends(get_current_user)):
    """Devolve a obra + items computados + despesas vinculadas + KPIs agregados."""
    work = await db.works.find_one({"id": work_id}, {"_id": 0})
    if not work:
        raise HTTPException(status_code=404, detail="Obra não encontrada")

    # Auto-sync se nunca foi sincronizado e tem budget_id
    if not work.get("items") and work.get("budget_id"):
        budget = await db.budgets.find_one({"id": work["budget_id"]}, {"_id": 0})
        if budget:
            items = []
            for idx, b_it in enumerate(budget.get("items", [])):
                items.append({
                    "id": str(uuid.uuid4()),
                    "budget_item_idx": idx,
                    "category": b_it.get("category", ""),
                    "name": b_it.get("name", ""),
                    "unit": b_it.get("unit", "un"),
                    "quantity": b_it.get("quantity", 1),
                    "predicted_unit_cost": b_it.get("unit_cost", 0),
                    "margin": b_it.get("margin", 0.6),
                    "real_unit_cost": 0,
                    "real_quantity": None,
                    "real_notes": "",
                    "history": [],
                    "is_extra": False,
                })
            work["items"] = items
            await db.works.update_one({"id": work_id}, {"$set": {"items": items, "items_synced_at": datetime.now(timezone.utc).isoformat()}})

    items = [_compute_work_item_totals(it) for it in (work.get("items") or [])]

    # Despesas vinculadas a esta obra
    expenses = await db.expenses.find({"obra_id": work_id}, {"_id": 0}).to_list(2000)
    expenses_total = round(sum(float(e.get("value_gross") or 0) for e in expenses), 2)

    # KPIs
    sale_total = round(sum(it.get("sale_total", 0) for it in items), 2)
    predicted_total = round(sum(it.get("predicted_total", 0) for it in items), 2)
    real_total_items = round(sum(it.get("real_total", 0) for it in items), 2)
    real_total = round(real_total_items + expenses_total, 2)
    predicted_profit = round(sale_total - predicted_total, 2)
    real_profit = round(sale_total - real_total, 2)
    margin_predicted_pct = round((predicted_profit / sale_total * 100) if sale_total > 0 else 0, 1)
    margin_real_pct = round((real_profit / sale_total * 100) if sale_total > 0 else 0, 1)
    overrun_pct = round(((real_total - predicted_total) / predicted_total * 100) if predicted_total > 0 else 0, 1)

    return {
        "work": work,
        "items": items,
        "expenses": expenses,
        "kpis": {
            "sale_total": sale_total,
            "predicted_total": predicted_total,
            "real_total_items": real_total_items,
            "expenses_total": expenses_total,
            "real_total": real_total,
            "predicted_profit": predicted_profit,
            "real_profit": real_profit,
            "margin_predicted_pct": margin_predicted_pct,
            "margin_real_pct": margin_real_pct,
            "overrun_pct": overrun_pct,
            "is_overrun": overrun_pct > 10,
        },
    }


@api_router.put("/works/{work_id}/items/{item_id}")
async def update_work_item(work_id: str, item_id: str, input: WorkItemUpdate, user=Depends(get_current_user)):
    _require_admin(user)   # só admin/escritório marca execução e custos reais
    work = await db.works.find_one({"id": work_id}, {"_id": 0})
    if not work:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    items = list(work.get("items") or [])
    idx = next((i for i, it in enumerate(items) if it.get("id") == item_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail="Item não encontrado")
    data = input.model_dump(exclude_none=True)
    if not data:
        raise HTTPException(status_code=400, detail="Nada para atualizar")

    # Normaliza execution_status
    if "execution_status" in data and data["execution_status"] not in ("pending", "in_progress", "done"):
        raise HTTPException(status_code=400, detail="execution_status inválido")

    # Se marcado como concluído mas sem executed_quantity, assume = quantity total
    if data.get("execution_status") == "done" and "executed_quantity" not in data:
        data["executed_quantity"] = float(items[idx].get("quantity") or 0)
    # Se em curso e sem quantidade, mantém a existente ou 0
    if data.get("execution_status") == "pending":
        data["executed_quantity"] = 0

    # Histórico de custo real
    prev_real = float(items[idx].get("real_unit_cost") or 0)
    new_real = data.get("real_unit_cost", prev_real)
    if new_real != prev_real:
        items[idx].setdefault("history", []).append({
            "at": datetime.now(timezone.utc).isoformat(),
            "by": user.get("name", ""),
            "from": prev_real,
            "to": float(new_real),
        })

    # Histórico de execução
    prev_status = items[idx].get("execution_status") or "pending"
    prev_qty = float(items[idx].get("executed_quantity") or 0)
    new_status = data.get("execution_status", prev_status)
    new_qty = float(data.get("executed_quantity", prev_qty))
    if new_status != prev_status or new_qty != prev_qty:
        items[idx].setdefault("execution_history", []).append({
            "at": datetime.now(timezone.utc).isoformat(),
            "by": user.get("name", ""),
            "status_from": prev_status,
            "status_to": new_status,
            "qty_from": prev_qty,
            "qty_to": new_qty,
            "note": data.get("execution_notes", ""),
        })

    items[idx] = {**items[idx], **data}
    await db.works.update_one({"id": work_id}, {"$set": {"items": items}})
    return await get_work_full(work_id, user)


@api_router.post("/works/{work_id}/items")
async def add_work_item_extra(work_id: str, input: WorkItemExtra, user=Depends(get_current_user)):
    """Adiciona item extra à obra (algo que não estava no orçamento original)."""
    work = await db.works.find_one({"id": work_id}, {"_id": 0})
    if not work:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    item = {
        "id": str(uuid.uuid4()),
        "budget_item_idx": None,
        "is_extra": True,
        "category": input.category,
        "name": input.name,
        "unit": input.unit,
        "quantity": input.quantity,
        "predicted_unit_cost": input.predicted_unit_cost,
        "margin": input.margin,
        "real_unit_cost": input.real_unit_cost,
        "real_quantity": input.real_quantity,
        "real_notes": input.notes,
        "history": [],
        "added_at": datetime.now(timezone.utc).isoformat(),
        "added_by": user.get("name", ""),
    }
    await db.works.update_one({"id": work_id}, {"$push": {"items": item}})
    return await get_work_full(work_id, user)


@api_router.delete("/works/{work_id}/items/{item_id}")
async def delete_work_item(work_id: str, item_id: str, user=Depends(get_current_user)):
    work = await db.works.find_one({"id": work_id}, {"_id": 0})
    if not work:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    new_items = [it for it in (work.get("items") or []) if it.get("id") != item_id]
    await db.works.update_one({"id": work_id}, {"$set": {"items": new_items}})
    return await get_work_full(work_id, user)



# --- Appointment Endpoints ---

@api_router.get("/appointments")
async def get_appointments(user=Depends(get_current_user)):
    appointments = await db.appointments.find({}, {"_id": 0}).sort("date", 1).to_list(1000)
    return appointments

def _require_admin(user):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Só o administrador pode gerir agendamentos")

def _overlap_query(appt, exclude_id=None):
    """Devolve query Mongo que detecta conflitos apenas quando os employee_ids se sobrepõem."""
    q = {
        "date": appt.date,
        "time_start": {"$lt": appt.time_end},
        "time_end": {"$gt": appt.time_start},
    }
    if exclude_id:
        q["id"] = {"$ne": exclude_id}
    if appt.employee_ids:
        q["employee_ids"] = {"$in": appt.employee_ids}
    return q

# --- Google Calendar helpers for Agenda ---
def _gcal_check_and_create(date_str, time_start, time_end, title, client_name="", location=""):
    """Check Google Calendar and create event. Returns (created_event_id, conflicts, suggestions)."""
    from service_orders import _get_calendar_service, _check_calendar_availability, GOOGLE_CALENDAR_ID
    preferred = f"{date_str}T{time_start}"
    # Calculate duration in hours
    try:
        from datetime import datetime as _dt
        t1 = _dt.strptime(time_start, "%H:%M")
        t2 = _dt.strptime(time_end, "%H:%M")
        dur = max((t2 - t1).total_seconds() / 3600, 1)
    except Exception:
        dur = 2

    has_conflict, conflicts, suggestions = _check_calendar_availability(preferred, int(dur))

    # Create event regardless (admin chose to create it)
    svc = _get_calendar_service()
    event_id = None
    if svc and GOOGLE_CALENDAR_ID:
        try:
            start_dt = datetime.strptime(f"{date_str}T{time_start}", "%Y-%m-%dT%H:%M")
            end_dt = datetime.strptime(f"{date_str}T{time_end}", "%Y-%m-%dT%H:%M")
            if start_dt.tzinfo is None:
                start_dt = start_dt.replace(tzinfo=timezone.utc)
            if end_dt.tzinfo is None:
                end_dt = end_dt.replace(tzinfo=timezone.utc)
            event = {
                'summary': f"📅 {title}" + (f" - {client_name}" if client_name else ""),
                'description': f"Agendamento Obelisco Manager\nCliente: {client_name}\nNotas: criado via Agenda",
                'location': location,
                'start': {'dateTime': start_dt.isoformat(), 'timeZone': 'Europe/Lisbon'},
                'end': {'dateTime': end_dt.isoformat(), 'timeZone': 'Europe/Lisbon'},
                'reminders': {'useDefault': False, 'overrides': [{'method': 'popup', 'minutes': 60}]},
            }
            created = svc.events().insert(calendarId=GOOGLE_CALENDAR_ID, body=event).execute()
            event_id = created.get('id')
            logger.info(f"Agenda: Calendar event created: {event_id}")
        except Exception as e:
            logger.warning(f"Agenda: Calendar event creation failed: {e}")

    return event_id, has_conflict, conflicts, suggestions

@api_router.get("/appointments/check-calendar")
async def check_calendar_for_appointment(date: str, time_start: str, time_end: str, user=Depends(get_current_user)):
    """Check Google Calendar availability for a given slot."""
    from service_orders import _check_calendar_availability
    preferred = f"{date}T{time_start}"
    try:
        from datetime import datetime as _dt
        t1 = _dt.strptime(time_start, "%H:%M")
        t2 = _dt.strptime(time_end, "%H:%M")
        dur = max((t2 - t1).total_seconds() / 3600, 1)
    except Exception:
        dur = 2
    has_conflict, conflicts, suggestions = _check_calendar_availability(preferred, int(dur))
    return {"available": not has_conflict, "conflicts": conflicts, "suggested_times": suggestions}

@api_router.post("/appointments")
async def create_appointment(input: AppointmentCreate, user=Depends(get_current_user)):
    _require_admin(user)
    existing = await db.appointments.find_one(_overlap_query(input))
    if existing:
        raise HTTPException(status_code=400, detail="Conflito: o(s) técnico(s) atribuído(s) já têm marcação nesse horário.")
    doc = {
        "id": str(uuid.uuid4()),
        **input.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    # Create Google Calendar event (sync, fast)
    try:
        event_id, _, _, _ = _gcal_check_and_create(
            input.date, input.time_start, input.time_end,
            input.title, input.client_name, input.location
        )
        if event_id:
            doc["gcal_event_id"] = event_id
    except Exception as e:
        logger.warning(f"Calendar event skipped: {e}")

    await db.appointments.insert_one(doc)
    doc.pop("_id", None)
    # Notificar os técnicos atribuídos
    when_label = f"{input.date} {input.time_start}"
    for emp_id in (input.employee_ids or []):
        await create_notification(
            db, user_id=emp_id, user_kind="employee", type="agenda",
            title="Nova marcação atribuída",
            message=f"{input.title} — {when_label}",
            link="/tech/agenda",
            meta={"appointment_id": doc["id"]},
        )
        # Push notification
        try:
            from push_notifications import send_push_to_user, PushMessage
            import asyncio
            asyncio.ensure_future(send_push_to_user(db, emp_id, PushMessage(
                title="📅 Nova Marcação",
                body=f"{input.title} — {when_label}" + (f" · {input.client_name}" if input.client_name else ""),
                tag=f"agenda-{doc['id']}",
                url="/tech/agenda",
            )))
        except Exception:
            pass
    return doc

@api_router.put("/appointments/{appointment_id}")
async def update_appointment(appointment_id: str, input: AppointmentCreate, user=Depends(get_current_user)):
    _require_admin(user)
    existing = await db.appointments.find_one(_overlap_query(input, exclude_id=appointment_id))
    if existing:
        raise HTTPException(status_code=400, detail="Conflito: o(s) técnico(s) atribuído(s) já têm marcação nesse horário.")
    before = await db.appointments.find_one({"id": appointment_id}, {"_id": 0}) or {}
    result = await db.appointments.update_one({"id": appointment_id}, {"$set": input.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Agendamento não encontrado")
    updated = await db.appointments.find_one({"id": appointment_id}, {"_id": 0})
    # Notificar
    old_ids = set(before.get("employee_ids") or [])
    new_ids = set(input.employee_ids or [])
    when_label = f"{input.date} {input.time_start}"
    changed = (
        before.get("date") != input.date
        or before.get("time_start") != input.time_start
        or before.get("time_end") != input.time_end
        or before.get("title") != input.title
    )
    for emp_id in (new_ids - old_ids):
        await create_notification(db, user_id=emp_id, user_kind="employee", type="agenda",
            title="Nova marcação atribuída",
            message=f"{input.title} — {when_label}",
            link="/tech/agenda", meta={"appointment_id": appointment_id})
    for emp_id in (new_ids & old_ids):
        if changed:
            await create_notification(db, user_id=emp_id, user_kind="employee", type="agenda",
                title="Marcação alterada",
                message=f"{input.title} — {when_label}",
                link="/tech/agenda", meta={"appointment_id": appointment_id})
    for emp_id in (old_ids - new_ids):
        await create_notification(db, user_id=emp_id, user_kind="employee", type="agenda",
            title="Marcação removida",
            message=f"Já não estás atribuído a: {before.get('title', '')}",
            link="/tech/agenda", meta={"appointment_id": appointment_id})
    return updated

@api_router.delete("/appointments/{appointment_id}")
async def delete_appointment(appointment_id: str, user=Depends(get_current_user)):
    _require_admin(user)
    before = await db.appointments.find_one({"id": appointment_id}, {"_id": 0}) or {}
    result = await db.appointments.delete_one({"id": appointment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Agendamento não encontrado")
    when_label = f"{before.get('date', '')} {before.get('time_start', '')}".strip()
    for emp_id in (before.get("employee_ids") or []):
        await create_notification(db, user_id=emp_id, user_kind="employee", type="agenda",
            title="Marcação cancelada",
            message=f"{before.get('title', 'Marcação')} — {when_label}",
            link="/tech/agenda", meta={"appointment_id": appointment_id})
    return {"message": "Agendamento eliminado"}


# --- Schedule from Proposal (auto next slot) ---

class ScheduleProposalInput(BaseModel):
    duration_hours: float = 4              # tamanho do bloco a procurar
    window: str = "any"                    # "morning" (09-13), "afternoon" (14-18) ou "any"
    start_from: Optional[str] = None       # ISO date — opcional, default = amanhã


@api_router.post("/proposals/{proposal_id}/schedule")
async def schedule_from_proposal(proposal_id: str, input: ScheduleProposalInput, user=Depends(get_current_user)):
    """Encontra o próximo slot livre em horário comercial e cria o agendamento.
    Bloqueia duplicados: se já existe um appointment para esta proposta, devolve 409 com o existente."""
    from datetime import date as _date_cls, timedelta as _td, datetime as _dt, time as _tm

    proposal = await db.proposals.find_one({"id": proposal_id}, {"_id": 0})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposta não encontrada")

    # Já existe agendamento para esta proposta? → 409
    existing_for_prop = await db.appointments.find_one({"proposal_id": proposal_id}, {"_id": 0})
    if existing_for_prop:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "Já existe um agendamento para esta proposta.",
                "appointment": existing_for_prop,
            },
        )

    # Janelas comerciais
    WINDOWS = {
        "morning": [("09:00", "13:00")],
        "afternoon": [("14:00", "18:00")],
        "any": [("09:00", "13:00"), ("14:00", "18:00")],
    }
    candidate_windows = WINDOWS.get((input.window or "any").lower(), WINDOWS["any"])
    duration_min = int(round((input.duration_hours or 4) * 60))

    # Carrega todos os appointments próximos (90 dias)
    if input.start_from:
        try:
            start_date = _date_cls.fromisoformat(input.start_from)
        except Exception:
            start_date = _date_cls.today() + _td(days=1)
    else:
        start_date = _date_cls.today() + _td(days=1)
    end_date = start_date + _td(days=60)

    apts = await db.appointments.find(
        {"date": {"$gte": start_date.isoformat(), "$lte": end_date.isoformat()}},
        {"_id": 0},
    ).to_list(2000)
    by_day = {}
    for a in apts:
        by_day.setdefault(a["date"], []).append(a)

    def _hm_to_min(s: str) -> int:
        h, m = s.split(":")
        return int(h) * 60 + int(m)

    def _min_to_hm(m: int) -> str:
        return f"{m // 60:02d}:{m % 60:02d}"

    found = None
    d = start_date
    while d <= end_date and not found:
        # skip weekends
        if d.weekday() >= 5:
            d += _td(days=1)
            continue
        day_apts = by_day.get(d.isoformat(), [])
        # Para cada janela do dia, gerar slots possíveis em passos de 30min e ver se cabe
        for win_start, win_end in candidate_windows:
            ws = _hm_to_min(win_start)
            we = _hm_to_min(win_end)
            if we - ws < duration_min:
                continue
            slot_start = ws
            while slot_start + duration_min <= we:
                slot_end = slot_start + duration_min
                # Verifica conflito com appointments existentes
                conflict = False
                for a in day_apts:
                    a_s = _hm_to_min((a.get("time_start") or "00:00")[:5])
                    a_e = _hm_to_min((a.get("time_end") or "00:00")[:5])
                    if not (slot_end <= a_s or slot_start >= a_e):
                        conflict = True
                        break
                if not conflict:
                    found = (d, slot_start, slot_end)
                    break
                slot_start += 30
            if found:
                break
        d += _td(days=1)

    if not found:
        raise HTTPException(status_code=422, detail="Sem slots livres nos próximos 60 dias úteis em horário comercial")

    day, s_min, e_min = found
    title = f"Obra — {proposal.get('title') or proposal.get('label') or 'Proposta'}"
    apt_doc = {
        "id": str(uuid.uuid4()),
        "title": title,
        "client_name": proposal.get("client_name", ""),
        "client_phone": proposal.get("client_phone", ""),
        "date": day.isoformat(),
        "time_start": _min_to_hm(s_min),
        "time_end": _min_to_hm(e_min),
        "notes": f"Agendado a partir da proposta {proposal.get('label') or ''} · Valor {proposal.get('final_value', 0)} EUR",
        "proposal_id": proposal_id,
        "budget_id": proposal.get("budget_id", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user.get("id", ""),
    }
    await db.appointments.insert_one(apt_doc)
    apt_doc.pop("_id", None)

    # Construir URL do widget externo com query params para pré-preencher
    widget_base = "https://tech-app-obelisco.emergent.host/widget"
    from urllib.parse import urlencode
    proposal_label = proposal.get("label", "") or ""
    proposal_title = proposal.get("title", "") or ""
    proposal_value = proposal.get("final_value", 0) or 0
    auto_description = (
        f"Visita técnica relacionada com a proposta {proposal_label} — {proposal_title}.\n"
        f"Valor da proposta: {proposal_value:.2f} EUR."
    )
    qs = urlencode({
        "client": proposal.get("client_name", "") or "",
        "client_name": proposal.get("client_name", "") or "",
        "phone": proposal.get("client_phone", "") or "",
        "email": proposal.get("client_email", "") or "",
        "address": proposal.get("client_address", "") or "",
        "title": proposal_title,
        "proposal_id": proposal_id,
        "proposal_label": proposal_label,
        "value": proposal_value,
        "service_type": "visita_tecnica",
        "description": auto_description,
        "date": apt_doc["date"],
        "preferred_date": apt_doc["date"],
        "time_start": apt_doc["time_start"],
        "time_end": apt_doc["time_end"],
        "preferred_time": apt_doc["time_start"],
    })
    widget_url = f"{widget_base}?{qs}"

    return {"appointment": apt_doc, "widget_url": widget_url}


# --- Dashboard ---

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user=Depends(get_current_user)):
    total_obras = await db.works.count_documents({})
    obras_em_andamento = await db.works.count_documents({"status": "em_execucao"})
    obras_finalizadas = await db.works.count_documents({"status": "finalizado"})

    works_data = await db.works.find({}, {"_id": 0, "predicted_cost": 1, "real_cost": 1}).to_list(1000)
    total_predicted = sum(w.get("predicted_cost", 0) for w in works_data)
    total_real = sum(w.get("real_cost", 0) for w in works_data)

    total_orcamentos = await db.budgets.count_documents({})
    total_propostas = await db.proposals.count_documents({})

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    appointments_today = await db.appointments.count_documents({"date": today})

    recent_works = await db.works.find({}, {"_id": 0}).sort("created_at", -1).to_list(5)
    recent_budgets = await db.budgets.find({}, {"_id": 0}).sort("created_at", -1).to_list(5)

    return {
        "total_obras": total_obras,
        "obras_em_andamento": obras_em_andamento,
        "obras_finalizadas": obras_finalizadas,
        "lucro_estimado": round(total_predicted - total_real, 2),
        "total_orcamentos": total_orcamentos,
        "total_propostas": total_propostas,
        "appointments_today": appointments_today,
        "total_predicted": round(total_predicted, 2),
        "total_real": round(total_real, 2),
        "recent_works": recent_works,
        "recent_budgets": recent_budgets
    }


from emergentintegrations.llm.chat import LlmChat, UserMessage
import json as json_module


# --- Catalogo de Categorias e Itens para Eletricistas/Telecom ---

CATEGORIES_CATALOG = [
    {
        "id": "cabos_fios",
        "name": "Cabos e Fios",
        "items": [
            {"name": "Cabo H05VV-F 3G2,5mm", "unit": "metro"},
            {"name": "Cabo H05VV-F 3G1,5mm", "unit": "metro"},
            {"name": "Cabo H05VV-F 5G2,5mm", "unit": "metro"},
            {"name": "Cabo VV 3x2,5mm", "unit": "metro"},
            {"name": "Cabo VV 3x4mm", "unit": "metro"},
            {"name": "Fio H07V-U 2,5mm (azul/castanho/verde-amarelo)", "unit": "metro"},
            {"name": "Fio H07V-U 1,5mm", "unit": "metro"},
            {"name": "Fio H07V-U 4mm", "unit": "metro"},
            {"name": "Fio H07V-U 6mm", "unit": "metro"},
            {"name": "Cabo coaxial RG6", "unit": "metro"},
            {"name": "Cabo UTP Cat5e", "unit": "metro"},
            {"name": "Cabo UTP Cat6", "unit": "metro"},
            {"name": "Cabo UTP Cat6 (bobine 305m)", "unit": "unidade"},
            {"name": "Cabo fibra optica monomodo", "unit": "metro"},
            {"name": "Cabo de alarme 4 condutores", "unit": "metro"},
            {"name": "Cabo LSZH 3G2,5mm", "unit": "metro"},
        ]
    },
    {
        "id": "quadros_protecao",
        "name": "Quadros e Protecao",
        "items": [
            {"name": "Quadro eletrico 12 modulos", "unit": "unidade"},
            {"name": "Quadro eletrico 24 modulos", "unit": "unidade"},
            {"name": "Quadro eletrico 36 modulos", "unit": "unidade"},
            {"name": "Disjuntor monofasico 10A", "unit": "unidade"},
            {"name": "Disjuntor monofasico 16A", "unit": "unidade"},
            {"name": "Disjuntor monofasico 20A", "unit": "unidade"},
            {"name": "Disjuntor monofasico 32A", "unit": "unidade"},
            {"name": "Disjuntor trifasico 32A", "unit": "unidade"},
            {"name": "Diferencial 40A 30mA bipolar", "unit": "unidade"},
            {"name": "Diferencial 63A 30mA tetrapolar", "unit": "unidade"},
            {"name": "Descarregador de sobretensao Tipo 2", "unit": "unidade"},
            {"name": "Contactor 25A", "unit": "unidade"},
            {"name": "Rele horario digital", "unit": "unidade"},
            {"name": "Barramento de ligacao", "unit": "unidade"},
        ]
    },
    {
        "id": "tomadas_interruptores",
        "name": "Tomadas e Interruptores",
        "items": [
            {"name": "Tomada Schuko encastrar (branca)", "unit": "unidade"},
            {"name": "Tomada Schuko dupla encastrar", "unit": "unidade"},
            {"name": "Interruptor simples encastrar", "unit": "unidade"},
            {"name": "Interruptor duplo encastrar", "unit": "unidade"},
            {"name": "Comutador de escada", "unit": "unidade"},
            {"name": "Comutador de lustre", "unit": "unidade"},
            {"name": "Inversor de grupo", "unit": "unidade"},
            {"name": "Tomada RJ45 Cat6 encastrar", "unit": "unidade"},
            {"name": "Tomada TV/SAT encastrar", "unit": "unidade"},
            {"name": "Espelho/Placa 1 posto", "unit": "unidade"},
            {"name": "Espelho/Placa 2 postos", "unit": "unidade"},
            {"name": "Tomada industrial CEE 16A", "unit": "unidade"},
            {"name": "Tomada industrial CEE 32A", "unit": "unidade"},
            {"name": "Regulador de intensidade (dimmer)", "unit": "unidade"},
        ]
    },
    {
        "id": "iluminacao",
        "name": "Iluminacao",
        "items": [
            {"name": "Downlight LED encastrar 12W", "unit": "unidade"},
            {"name": "Downlight LED encastrar 18W", "unit": "unidade"},
            {"name": "Projetor LED exterior 30W", "unit": "unidade"},
            {"name": "Projetor LED exterior 50W", "unit": "unidade"},
            {"name": "Projetor LED exterior 100W", "unit": "unidade"},
            {"name": "Fita LED 5m branco quente", "unit": "unidade"},
            {"name": "Fita LED 5m RGB", "unit": "unidade"},
            {"name": "Lampada LED E27 10W", "unit": "unidade"},
            {"name": "Lampada LED E14 6W", "unit": "unidade"},
            {"name": "Lampada LED GU10 7W", "unit": "unidade"},
            {"name": "Armadura estanque LED 36W 120cm", "unit": "unidade"},
            {"name": "Painel LED 60x60 40W", "unit": "unidade"},
            {"name": "Aplique LED exterior", "unit": "unidade"},
            {"name": "Sensor de movimento PIR", "unit": "unidade"},
            {"name": "Sensor crepuscular", "unit": "unidade"},
        ]
    },
    {
        "id": "calhas_tubos",
        "name": "Calhas e Tubos",
        "items": [
            {"name": "Tubo VD 20mm (vara 3m)", "unit": "unidade"},
            {"name": "Tubo VD 25mm (vara 3m)", "unit": "unidade"},
            {"name": "Tubo VD 32mm (vara 3m)", "unit": "unidade"},
            {"name": "Tubo corrugado 20mm (rolo 50m)", "unit": "unidade"},
            {"name": "Tubo corrugado 25mm (rolo 50m)", "unit": "unidade"},
            {"name": "Tubo corrugado 32mm (rolo 25m)", "unit": "unidade"},
            {"name": "Calha DLP 40x25mm (2m)", "unit": "unidade"},
            {"name": "Calha DLP 60x40mm (2m)", "unit": "unidade"},
            {"name": "Calha de chao 50x12mm", "unit": "metro"},
            {"name": "Caixa de derivacao 100x100", "unit": "unidade"},
            {"name": "Caixa de derivacao 150x150", "unit": "unidade"},
            {"name": "Caixa de aparelhagem (fundo)", "unit": "unidade"},
            {"name": "Braçadeira clip 20mm", "unit": "unidade"},
            {"name": "Braçadeira clip 25mm", "unit": "unidade"},
            {"name": "Abraçadeira nylon 200mm (saco 100)", "unit": "unidade"},
        ]
    },
    {
        "id": "telecomunicacoes",
        "name": "Telecomunicacoes",
        "items": [
            {"name": "Router WiFi 6 dual band", "unit": "unidade"},
            {"name": "Access Point WiFi 6", "unit": "unidade"},
            {"name": "Switch Gigabit 8 portas", "unit": "unidade"},
            {"name": "Switch Gigabit 16 portas", "unit": "unidade"},
            {"name": "Switch PoE 8 portas", "unit": "unidade"},
            {"name": "Patch panel 24 portas Cat6", "unit": "unidade"},
            {"name": "Conector RJ45 Cat6 (saco 100)", "unit": "unidade"},
            {"name": "Patch cord Cat6 1m", "unit": "unidade"},
            {"name": "Patch cord Cat6 3m", "unit": "unidade"},
            {"name": "Bastidor rack 6U parede", "unit": "unidade"},
            {"name": "Bastidor rack 12U parede", "unit": "unidade"},
            {"name": "Bastidor rack 42U chao", "unit": "unidade"},
            {"name": "Organizador de cabos 1U", "unit": "unidade"},
            {"name": "Tomada fibra optica SC/APC", "unit": "unidade"},
            {"name": "Camera IP PoE 4MP", "unit": "unidade"},
            {"name": "NVR 8 canais PoE", "unit": "unidade"},
        ]
    },
    {
        "id": "ferramentas_acessorios",
        "name": "Ferramentas e Acessorios",
        "items": [
            {"name": "Multimetro digital profissional", "unit": "unidade"},
            {"name": "Alicate de corte diagonal", "unit": "unidade"},
            {"name": "Alicate universal isolado 1000V", "unit": "unidade"},
            {"name": "Alicate de cravar RJ45", "unit": "unidade"},
            {"name": "Alicate descarnar cabos", "unit": "unidade"},
            {"name": "Detetor de tensao sem contacto", "unit": "unidade"},
            {"name": "Chave de fendas isolada 1000V (jogo)", "unit": "unidade"},
            {"name": "Fita isoladora preta 20m", "unit": "unidade"},
            {"name": "Fita isoladora (pack 10 cores)", "unit": "unidade"},
            {"name": "Terminais de cravar (caixa sortida)", "unit": "unidade"},
            {"name": "Passa fios aço 20m", "unit": "unidade"},
            {"name": "Passa fios nylon 15m", "unit": "unidade"},
            {"name": "Testador de cabos RJ45/RJ11", "unit": "unidade"},
        ]
    },
    {
        "id": "carregamento_ev",
        "name": "Carregamento Veiculo Eletrico",
        "items": [
            {"name": "Wallbox monofasico 7.4kW", "unit": "unidade"},
            {"name": "Wallbox trifasico 11kW", "unit": "unidade"},
            {"name": "Wallbox trifasico 22kW", "unit": "unidade"},
            {"name": "Cabo de carregamento Tipo 2 (5m)", "unit": "unidade"},
            {"name": "Protecao diferencial Tipo B 40A", "unit": "unidade"},
        ]
    },
    {
        "id": "energia_solar",
        "name": "Energia Solar",
        "items": [
            {"name": "Painel solar fotovoltaico 400W", "unit": "unidade"},
            {"name": "Painel solar fotovoltaico 550W", "unit": "unidade"},
            {"name": "Inversor hibrido monofasico 5kW", "unit": "unidade"},
            {"name": "Inversor string trifasico 10kW", "unit": "unidade"},
            {"name": "Bateria de litio 5kWh", "unit": "unidade"},
            {"name": "Bateria de litio 10kWh", "unit": "unidade"},
            {"name": "Estrutura montagem telhado (kit 4 paineis)", "unit": "unidade"},
            {"name": "Cabo solar 6mm (vermelho)", "unit": "metro"},
            {"name": "Cabo solar 6mm (preto)", "unit": "metro"},
            {"name": "Conector MC4 (par)", "unit": "unidade"},
        ]
    },
    {
        "id": "domotica",
        "name": "Domotica e Automacao",
        "items": [
            {"name": "Interruptor inteligente WiFi", "unit": "unidade"},
            {"name": "Tomada inteligente WiFi", "unit": "unidade"},
            {"name": "Lampada inteligente E27 WiFi RGB", "unit": "unidade"},
            {"name": "Controlador de estores WiFi", "unit": "unidade"},
            {"name": "Hub Zigbee/Z-Wave", "unit": "unidade"},
            {"name": "Sensor de porta/janela", "unit": "unidade"},
            {"name": "Sensor de temperatura/humidade", "unit": "unidade"},
            {"name": "Campainha video WiFi (video doorbell)", "unit": "unidade"},
            {"name": "Fechadura inteligente", "unit": "unidade"},
        ]
    },
    {
        "id": "mao_obra_eletricista",
        "name": "Mao de Obra - Eletricista",
        "items": [
            {"name": "Instalacao de ponto de luz (completo)", "unit": "unidade"},
            {"name": "Instalacao de tomada (completo)", "unit": "unidade"},
            {"name": "Instalacao de interruptor/comutador", "unit": "unidade"},
            {"name": "Montagem de quadro eletrico residencial", "unit": "unidade"},
            {"name": "Montagem de quadro eletrico industrial", "unit": "unidade"},
            {"name": "Substituicao de quadro eletrico", "unit": "unidade"},
            {"name": "Passagem de cabos (por metro linear)", "unit": "metro"},
            {"name": "Abertura de roco em parede (por metro)", "unit": "metro"},
            {"name": "Instalacao de disjuntor/diferencial", "unit": "unidade"},
            {"name": "Instalacao de descarregador de sobretensao", "unit": "unidade"},
            {"name": "Verificacao e teste de instalacao eletrica", "unit": "unidade"},
            {"name": "Diagnostico de avaria eletrica", "unit": "unidade"},
            {"name": "Reparacao de curto-circuito", "unit": "unidade"},
            {"name": "Instalacao de projetor LED exterior", "unit": "unidade"},
            {"name": "Instalacao de downlight encastrado", "unit": "unidade"},
            {"name": "Instalacao de fita LED (por metro)", "unit": "metro"},
            {"name": "Instalacao de sensor de movimento", "unit": "unidade"},
            {"name": "Instalacao de wallbox carregamento EV", "unit": "unidade"},
            {"name": "Instalacao de painel solar (por painel)", "unit": "unidade"},
            {"name": "Instalacao de inversor solar", "unit": "unidade"},
            {"name": "Ligacao a rede eletrica (RESP)", "unit": "unidade"},
            {"name": "Certificacao de instalacao eletrica", "unit": "unidade"},
            {"name": "Manutencao preventiva de quadro eletrico", "unit": "unidade"},
            {"name": "Medicao de terra e resistencia de isolamento", "unit": "unidade"},
            {"name": "Instalacao de sistema de terra", "unit": "unidade"},
            {"name": "Hora de trabalho eletricista (normal)", "unit": "hora"},
            {"name": "Hora de trabalho eletricista (urgente/noturno)", "unit": "hora"},
            {"name": "Deslocacao tecnica (Grande Lisboa)", "unit": "unidade"},
            {"name": "Deslocacao tecnica (fora de Lisboa)", "unit": "unidade"},
        ]
    },
    {
        "id": "mao_obra_telecom",
        "name": "Mao de Obra - Telecomunicacoes",
        "items": [
            {"name": "Instalacao de ponto de rede Cat6 (completo)", "unit": "unidade"},
            {"name": "Instalacao de ponto de rede Cat5e (completo)", "unit": "unidade"},
            {"name": "Cablagem estruturada residencial (ate 8 pontos)", "unit": "unidade"},
            {"name": "Cablagem estruturada escritorio (ate 24 pontos)", "unit": "unidade"},
            {"name": "Montagem de bastidor/rack com patch panel", "unit": "unidade"},
            {"name": "Organizacao de cablagem em bastidor", "unit": "unidade"},
            {"name": "Certificacao de ponto de rede Cat6", "unit": "unidade"},
            {"name": "Instalacao de router/access point WiFi", "unit": "unidade"},
            {"name": "Configuracao de rede WiFi (SSID, seguranca, mesh)", "unit": "unidade"},
            {"name": "Instalacao de switch gerido", "unit": "unidade"},
            {"name": "Instalacao de switch PoE", "unit": "unidade"},
            {"name": "Passagem de cabo UTP (por metro)", "unit": "metro"},
            {"name": "Passagem de fibra optica (por metro)", "unit": "metro"},
            {"name": "Fusao de fibra optica (por fusao)", "unit": "unidade"},
            {"name": "Instalacao de tomada fibra optica", "unit": "unidade"},
            {"name": "Instalacao de camera IP/CCTV", "unit": "unidade"},
            {"name": "Configuracao de NVR/DVR", "unit": "unidade"},
            {"name": "Instalacao de sistema CCTV completo (ate 4 cameras)", "unit": "unidade"},
            {"name": "Instalacao de sistema CCTV completo (ate 8 cameras)", "unit": "unidade"},
            {"name": "Instalacao de videoporteiro/video doorbell", "unit": "unidade"},
            {"name": "Instalacao de sistema de alarme", "unit": "unidade"},
            {"name": "Configuracao de acesso remoto (VPN/DDNS)", "unit": "unidade"},
            {"name": "Diagnostico de rede/conectividade", "unit": "unidade"},
            {"name": "Instalacao de central telefonica IP", "unit": "unidade"},
            {"name": "Instalacao de sistema IPTV/distribuicao TV", "unit": "unidade"},
            {"name": "Hora de trabalho tecnico telecom (normal)", "unit": "hora"},
            {"name": "Hora de trabalho tecnico telecom (urgente/noturno)", "unit": "hora"},
        ]
    },
    {
        "id": "trabalhos_burocraticos",
        "name": "Trabalhos Burocraticos",
        "items": [
            # Pedidos junto de distribuidoras
            {"name": "Pedido de ramal novo (E-REDES)", "unit": "unidade"},
            {"name": "Pedido de aumento de potencia contratada (E-REDES)", "unit": "unidade"},
            {"name": "Pedido de reducao de potencia contratada (E-REDES)", "unit": "unidade"},
            {"name": "Pedido de alteracao de tarifario (E-REDES)", "unit": "unidade"},
            {"name": "Pedido de mudanca de contador para exterior (E-REDES)", "unit": "unidade"},
            {"name": "Pedido de ligacao provisoria de obra (E-REDES)", "unit": "unidade"},
            {"name": "Pedido de contagem trifasica (E-REDES)", "unit": "unidade"},
            {"name": "Pedido de instalacao de posto de transformacao (PT)", "unit": "unidade"},
            # ITED / ITUR / Certificacoes
            {"name": "Emissao de Ficha Tecnica de Instalacao (FTI/ITED2)", "unit": "unidade"},
            {"name": "Emissao de Termo de Responsabilidade ITED/ITUR", "unit": "unidade"},
            {"name": "Certificacao de instalacao eletrica (CERTIEL)", "unit": "unidade"},
            {"name": "Inspecao periodica CERTIEL", "unit": "unidade"},
            {"name": "Renovacao de certificado energetico (ADENE)", "unit": "unidade"},
            {"name": "Emissao de certificado energetico novo", "unit": "unidade"},
            # Projetos e memorias
            {"name": "Elaboracao de projeto eletrico (RITE)", "unit": "unidade"},
            {"name": "Memoria descritiva e justificativa", "unit": "unidade"},
            {"name": "Peca escrita e desenhada para camara municipal", "unit": "unidade"},
            {"name": "Deposito de projeto na entidade competente", "unit": "unidade"},
            # Camaras / Licenciamentos
            {"name": "Pedido de licenca de obra na Camara Municipal", "unit": "unidade"},
            {"name": "Pedido de ocupacao de via publica", "unit": "unidade"},
            {"name": "Comunicacao previa de obras", "unit": "unidade"},
            # Alteracoes de dados de cliente
            {"name": "Alteracao de titularidade do contrato de fornecimento", "unit": "unidade"},
            {"name": "Atualizacao de dados fiscais/morada junto do comercializador", "unit": "unidade"},
            {"name": "Denuncia de contrato/mudanca de comercializador", "unit": "unidade"},
            # Outros trabalhos administrativos
            {"name": "Elaboracao de orcamento e apresentacao ao cliente", "unit": "unidade"},
            {"name": "Reuniao tecnica presencial no local (deslocacao)", "unit": "unidade"},
            {"name": "Reuniao tecnica online (Teams/Zoom)", "unit": "hora"},
            {"name": "Acompanhamento de vistoria/fiscalizacao", "unit": "unidade"},
            {"name": "Preparacao e submissao de documentacao a seguradora", "unit": "unidade"},
            {"name": "Elaboracao de peritagem/relatorio tecnico", "unit": "unidade"},
            {"name": "Traducao de documentacao tecnica", "unit": "unidade"},
            {"name": "Hora de servico administrativo (backoffice)", "unit": "hora"},
        ]
    },
]


# --- Categories & Price Lookup Endpoints ---

class PriceLookupRequest(BaseModel):
    item_name: str

class ProposalSettingsUpdate(BaseModel):
    payment_methods: Optional[List[str]] = None
    payment_split: Optional[str] = None
    validity_days: Optional[int] = None
    warranty_text: Optional[str] = None
    conditions: Optional[List[str]] = None
    notes: Optional[str] = None


# --- Logo endpoint (serves logo as base64 to avoid CORS) ---

@api_router.get("/logo")
async def get_logo():
    logo_path = Path(__file__).parent / "logo.png"
    if not logo_path.exists():
        raise HTTPException(status_code=404, detail="Logo não encontrado")
    import base64
    with open(logo_path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode()
    return {"logo": f"data:image/png;base64,{b64}"}


# --- Proposal Settings ---

DEFAULT_PROPOSAL_SETTINGS = {
    "payment_methods": ["Transferencia Bancaria", "MB Way", "Multibanco", "Cartao de Credito/Debito"],
    "payment_split": "50% no inicio dos trabalhos, 50% na conclusao",
    "validity_days": 30,
    "warranty_text": "Garantia de 2 anos sobre mao de obra e materiais fornecidos",
    "conditions": [
        "Valores em EUR, IVA NAO incluido (a acrescer a taxa legal em vigor)",
        "Deslocacao incluida na zona da Grande Lisboa",
        "Material e mao de obra incluidos",
        "Alteracoes ao orcamento podem afetar o valor final",
    ],
    "notes": "",
}

@api_router.get("/proposal-settings")
async def get_proposal_settings(user=Depends(get_current_user)):
    settings = await db.proposal_settings.find_one({}, {"_id": 0})
    if not settings:
        return DEFAULT_PROPOSAL_SETTINGS
    return settings

@api_router.put("/proposal-settings")
async def update_proposal_settings(input: ProposalSettingsUpdate, user=Depends(get_current_user)):
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nada para atualizar")
    existing = await db.proposal_settings.find_one({})
    if existing:
        await db.proposal_settings.update_one({"_id": existing["_id"]}, {"$set": update_data})
    else:
        doc = {**DEFAULT_PROPOSAL_SETTINGS, **update_data}
        await db.proposal_settings.insert_one(doc)
    settings = await db.proposal_settings.find_one({}, {"_id": 0})
    return settings


@api_router.get("/categories")
async def get_categories(user=Depends(get_current_user)):
    # Merge static catalog with custom items from materials_db
    import copy
    result = copy.deepcopy(CATEGORIES_CATALOG)
    cat_map = {c["name"]: c for c in result}

    # Get custom items from materials_db that are not in the static catalog
    custom_materials = await db.materials_db.find({"active": True, "custom": True}, {"_id": 0}).to_list(2000)
    for mat in custom_materials:
        cat_name = mat.get("category", "")
        if not cat_name:
            continue
        if cat_name in cat_map:
            # Add to existing category if not already there
            existing_names = {i["name"] for i in cat_map[cat_name]["items"]}
            if mat["description"] not in existing_names:
                cat_map[cat_name]["items"].append({"name": mat["description"], "unit": mat.get("unit", "unidade")})
        else:
            # Create new category
            new_cat = {"id": cat_name.lower().replace(" ", "_"), "name": cat_name, "items": [{"name": mat["description"], "unit": mat.get("unit", "unidade")}]}
            result.append(new_cat)
            cat_map[cat_name] = new_cat

    return result


class SaveCustomItemInput(BaseModel):
    category: str
    name: str
    unit_cost: float = 0
    unit: str = "unidade"

@api_router.post("/categories/save-item")
async def save_custom_item(input: SaveCustomItemInput, user=Depends(get_current_user)):
    """Save a custom item to the materials database so it appears in future category dropdowns"""
    if not input.name or not input.category:
        raise HTTPException(status_code=400, detail="Nome e categoria obrigatorios")

    # Check if already exists
    existing = await db.materials_db.find_one({"description": input.name, "category": input.category})
    if existing:
        # Update price if changed
        if input.unit_cost > 0 and input.unit_cost != existing.get("purchase_price", 0):
            history = existing.get("price_history", [])
            history.append({"price": input.unit_cost, "date": datetime.now(timezone.utc).isoformat()})
            await db.materials_db.update_one(
                {"_id": existing["_id"]},
                {"$set": {"purchase_price": input.unit_cost, "market_price": input.unit_cost, "price_history": history, "price_updated_at": datetime.now(timezone.utc).isoformat()}}
            )
        return {"message": "Item atualizado", "new": False}

    # Create new material
    doc = {
        "id": str(uuid.uuid4()), "code": "", "description": input.name,
        "category": input.category, "subcategory": "", "brand": "", "supplier": "",
        "unit": input.unit, "purchase_price": input.unit_cost, "market_price": input.unit_cost,
        "waste_pct": 5, "notes": "Adicionado automaticamente via orcamento", "active": True,
        "custom": True,
        "price_history": [{"price": input.unit_cost, "date": datetime.now(timezone.utc).isoformat()}] if input.unit_cost > 0 else [],
        "price_updated_at": datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.materials_db.insert_one(doc)
    doc.pop("_id", None)
    return {"message": "Item guardado na categoria", "new": True, "item": doc}


@api_router.post("/price-lookup")
async def price_lookup(input: PriceLookupRequest, user=Depends(get_current_user)):
    llm_key = os.environ.get("EMERGENT_LLM_KEY")
    if not llm_key:
        raise HTTPException(status_code=500, detail="Chave LLM nao configurada")

    try:
        chat = LlmChat(
            api_key=llm_key,
            session_id=f"price-{uuid.uuid4()}",
            system_message=(
                "Es um especialista em precos de material eletrico e telecomunicacoes em Portugal, "
                "com experiencia em orcamentacao para empresas de eletricistas na zona de Lisboa. "
                "O utilizador vai dar-te o nome de um item/material. "
                "Deves responder APENAS com um objeto JSON valido com estes campos: "
                '{"price": <preco medio de compra em EUR>, '
                '"price_min": <preco minimo>, '
                '"price_max": <preco maximo>, '
                '"margin": <margem recomendada como decimal, ex: 0.65 para 65%>, '
                '"install_cost": <custo estimado de instalacao/mao de obra por unidade em EUR>, '
                '"unit": "metro/unidade/pack", '
                '"source": "<breve descrição>"} '
                "REGRAS PARA A MARGEM: "
                "- A margem deve cobrir: mao de obra do eletricista (media 15-25 EUR/hora em Lisboa), "
                "deslocacao (taxa media 35 EUR na Grande Lisboa), "
                "desgaste de ferramentas, seguro, impostos da empresa, e lucro. "
                "- Para itens simples (cabos, tomadas, interruptores): margem entre 0.80 e 1.20 (80-120%). "
                "- Para itens complexos (quadros, wallbox, paineis solares): margem entre 0.40 e 0.70 (40-70%). "
                "- Para ferramentas (sem instalacao): margem entre 0.15 e 0.30 (15-30%). "
                "- install_cost e o valor aproximado da mao de obra para instalar 1 unidade desse item. "
                "Baseia-te em precos de retalho em Portugal (Leroy Merlin, Voltimum, Material Eletrico Online, Jolar, etc). "
                "Se nao conseguires estimar, devolve price: 0. "
                "NAO incluas texto adicional, apenas o JSON."
            )
        ).with_model("openai", "gpt-5.2")

        user_msg = UserMessage(text=f"Preco medio de compra e margem de instalacao em Lisboa para: {input.item_name}")
        response = await chat.send_message(user_msg)

        # Parse the JSON response
        response_text = response.strip()
        # Handle markdown code blocks
        if response_text.startswith("```"):
            response_text = response_text.split("\n", 1)[1] if "\n" in response_text else response_text[3:]
            if response_text.endswith("```"):
                response_text = response_text[:-3]
            response_text = response_text.strip()

        price_data = json_module.loads(response_text)
        return {
            "item_name": input.item_name,
            "price": price_data.get("price", 0),
            "price_min": price_data.get("price_min", 0),
            "price_max": price_data.get("price_max", 0),
            "margin": price_data.get("margin", 0.6),
            "install_cost": price_data.get("install_cost", 0),
            "unit": price_data.get("unit", "unidade"),
            "source": price_data.get("source", "Estimativa IA"),
        }
    except json_module.JSONDecodeError:
        logger.error(f"Price lookup JSON parse error for: {input.item_name}, response: {response_text}")
        return {
            "item_name": input.item_name,
            "price": 0,
            "price_min": 0,
            "price_max": 0,
            "margin": 0.6,
            "install_cost": 0,
            "unit": "unidade",
            "source": "Erro ao interpretar resposta",
        }
    except Exception as e:
        logger.error(f"Price lookup error for: {input.item_name}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Erro na pesquisa de preco: {str(e)}")


# ============================================================
# PROFESSIONAL BUDGETING ENGINE
# ============================================================

SPECIALTIES = [
    {"id": "instalacoes_eletricas", "name": "Instalacoes Eletricas"},
    {"id": "ited", "name": "ITED / Telecomunicacoes"},
    {"id": "cctv", "name": "CCTV / Videovigilancia"},
    {"id": "intrusao", "name": "Sistemas de Intrusao"},
    {"id": "bastidores", "name": "Bastidores / Sala Tecnica"},
    {"id": "engenharia", "name": "Engenharia / Certificacao"},
    {"id": "trabalhos_prep", "name": "Trabalhos Preparatorios"},
]

DEFAULT_LABOR = [
    {"id": "eletricista", "type": "eletricista", "description": "Eletricista certificado", "cost_hour": 15, "sell_hour": 35, "charges": "SS+seguro", "notes": ""},
    {"id": "ajudante", "type": "ajudante", "description": "Ajudante de eletricista", "cost_hour": 10, "sell_hour": 22, "charges": "SS+seguro", "notes": ""},
    {"id": "tecnico_ited", "type": "tecnico_ited", "description": "Tecnico ITED certificado", "cost_hour": 18, "sell_hour": 40, "charges": "SS+seguro+cert", "notes": ""},
    {"id": "tecnico_cctv", "type": "tecnico_cctv", "description": "Tecnico CCTV", "cost_hour": 16, "sell_hour": 38, "charges": "SS+seguro", "notes": ""},
    {"id": "tecnico_intrusao", "type": "tecnico_intrusao", "description": "Tecnico sistemas intrusao", "cost_hour": 16, "sell_hour": 38, "charges": "SS+seguro", "notes": ""},
    {"id": "encarregado", "type": "encarregado", "description": "Encarregado de obra", "cost_hour": 20, "sell_hour": 45, "charges": "SS+seguro", "notes": ""},
    {"id": "engenheiro", "type": "engenheiro", "description": "Engenheiro / Direcao tecnica", "cost_hour": 30, "sell_hour": 65, "charges": "SS+seguro", "notes": ""},
]

DEFAULT_PRODUCTIVITIES = [
    {"item": "Tomada simples encastrar", "unit": "un", "time_min": 20, "difficulty": "baixa", "technician": "eletricista", "loss_pct": 5},
    {"item": "Tomada dupla encastrar", "unit": "un", "time_min": 25, "difficulty": "baixa", "technician": "eletricista", "loss_pct": 5},
    {"item": "Interruptor simples", "unit": "un", "time_min": 15, "difficulty": "baixa", "technician": "eletricista", "loss_pct": 5},
    {"item": "Comutador de escada", "unit": "un", "time_min": 20, "difficulty": "baixa", "technician": "eletricista", "loss_pct": 5},
    {"item": "Ponto de luz completo", "unit": "un", "time_min": 30, "difficulty": "baixa", "technician": "eletricista", "loss_pct": 10},
    {"item": "Downlight LED encastrar", "unit": "un", "time_min": 25, "difficulty": "baixa", "technician": "eletricista", "loss_pct": 5},
    {"item": "Projetor LED exterior", "unit": "un", "time_min": 35, "difficulty": "media", "technician": "eletricista", "loss_pct": 5},
    {"item": "Passagem de cabo (metro)", "unit": "m", "time_min": 3, "difficulty": "baixa", "technician": "eletricista", "loss_pct": 10},
    {"item": "Passagem de tubo embebido", "unit": "m", "time_min": 8, "difficulty": "media", "technician": "eletricista", "loss_pct": 10},
    {"item": "Abertura de roco", "unit": "m", "time_min": 15, "difficulty": "alta", "technician": "eletricista", "loss_pct": 15},
    {"item": "Quadro eletrico residencial", "unit": "un", "time_min": 240, "difficulty": "alta", "technician": "eletricista", "loss_pct": 10},
    {"item": "Quadro eletrico industrial", "unit": "un", "time_min": 480, "difficulty": "muito_alta", "technician": "eletricista", "loss_pct": 15},
    {"item": "Disjuntor/Diferencial", "unit": "un", "time_min": 10, "difficulty": "baixa", "technician": "eletricista", "loss_pct": 0},
    {"item": "Wallbox carregamento EV", "unit": "un", "time_min": 180, "difficulty": "alta", "technician": "eletricista", "loss_pct": 10},
    {"item": "Painel solar (por painel)", "unit": "un", "time_min": 60, "difficulty": "alta", "technician": "eletricista", "loss_pct": 5},
    {"item": "Inversor solar", "unit": "un", "time_min": 120, "difficulty": "alta", "technician": "eletricista", "loss_pct": 5},
    {"item": "Ponto rede Cat6", "unit": "un", "time_min": 30, "difficulty": "media", "technician": "tecnico_ited", "loss_pct": 10},
    {"item": "Ponto rede Cat5e", "unit": "un", "time_min": 25, "difficulty": "media", "technician": "tecnico_ited", "loss_pct": 10},
    {"item": "Passagem cabo UTP", "unit": "m", "time_min": 3, "difficulty": "baixa", "technician": "tecnico_ited", "loss_pct": 10},
    {"item": "Passagem fibra optica", "unit": "m", "time_min": 5, "difficulty": "media", "technician": "tecnico_ited", "loss_pct": 5},
    {"item": "Fusao fibra optica", "unit": "un", "time_min": 15, "difficulty": "alta", "technician": "tecnico_ited", "loss_pct": 5},
    {"item": "Montagem bastidor/rack", "unit": "un", "time_min": 240, "difficulty": "alta", "technician": "tecnico_ited", "loss_pct": 5},
    {"item": "Camera IP CCTV", "unit": "un", "time_min": 90, "difficulty": "media", "technician": "tecnico_cctv", "loss_pct": 5},
    {"item": "NVR/DVR configuracao", "unit": "un", "time_min": 120, "difficulty": "media", "technician": "tecnico_cctv", "loss_pct": 5},
    {"item": "Detetor de intrusao", "unit": "un", "time_min": 30, "difficulty": "media", "technician": "tecnico_intrusao", "loss_pct": 5},
    {"item": "Central de alarme", "unit": "un", "time_min": 180, "difficulty": "alta", "technician": "tecnico_intrusao", "loss_pct": 5},
    {"item": "Videoporteiro", "unit": "un", "time_min": 120, "difficulty": "media", "technician": "tecnico_ited", "loss_pct": 5},
    {"item": "Certificacao instalacao", "unit": "un", "time_min": 120, "difficulty": "media", "technician": "engenheiro", "loss_pct": 0},
    {"item": "Sensor movimento PIR", "unit": "un", "time_min": 20, "difficulty": "baixa", "technician": "eletricista", "loss_pct": 5},
    {"item": "Fita LED por metro", "unit": "m", "time_min": 10, "difficulty": "baixa", "technician": "eletricista", "loss_pct": 10},
]

DEFAULT_SYSTEM_SETTINGS = {
    "iva_rate": 23,
    "min_margin": 15,
    "target_margin": 30,
    "treasury_settings": {
        "anomaly_threshold_pct": 18,
    },
    "indirect_costs": {
        "deslocacao": 3, "ferramentas": 2, "consumiveis": 1.5,
        "gestao_obra": 4, "supervisao": 3, "logistica": 2,
        "seguros": 1.5, "administrativos": 3,
    },
    "risk_levels": {"baixo": 3, "medio": 5, "alto": 8, "muito_alto": 12},
    "proposal_modes": {
        "basico": {"margin_factor": 0.85, "risk": "baixo", "label": "Basico - Margem reduzida para fechar obra"},
        "profissional": {"margin_factor": 1.0, "risk": "medio", "label": "Profissional - Margem equilibrada"},
        "premium": {"margin_factor": 1.20, "risk": "alto", "label": "Premium - Margem alta com protecao"},
    },
    "company_info": {
        "name": "Obelisco Radical", "subtitle": "Eletricidade & Telecomunicacoes",
        "phone": "+351 911 132 401", "email": "obeliscoradical@gmail.com",
        "website": "www.obeliscoradical.pt", "address": "Grande Lisboa", "nif": "",
    },
}


# --- Pro Models ---

class LaborInput(BaseModel):
    type: str
    description: str = ""
    cost_hour: float = 0
    sell_hour: float = 0
    charges: str = ""
    notes: str = ""

class ProductivityInput(BaseModel):
    item: str
    unit: str = "un"
    time_min: float = 0
    difficulty: str = "media"
    technician: str = "eletricista"
    loss_pct: float = 5
    notes: str = ""

class MaterialInput(BaseModel):
    code: str = ""
    description: str
    category: str = ""
    subcategory: str = ""
    brand: str = ""
    supplier: str = ""
    supplier_nif: str = ""
    unit: str = "un"
    purchase_price: float = 0
    market_price: float = 0
    waste_pct: float = 5
    stock_current: float = 0
    stock_min: float = 0
    vat_rate: float = 23
    notes: str = ""
    active: bool = True

class StockMovementInput(BaseModel):
    material_id: str
    movement_type: str    # "entrada" or "saida"
    quantity: float
    reason: str = ""         # compra, consumo obra, ajuste, devolucao
    obra_id: Optional[str] = None
    notes: str = ""

class SystemSettingsInput(BaseModel):
    iva_rate: Optional[float] = None
    min_margin: Optional[float] = None
    target_margin: Optional[float] = None
    treasury_settings: Optional[dict] = None
    indirect_costs: Optional[dict] = None
    risk_levels: Optional[dict] = None
    proposal_modes: Optional[dict] = None
    company_info: Optional[dict] = None

class ProBudgetItem(BaseModel):
    category: str = ""
    name: str = ""
    quantity: float = 1
    unit_cost: float = 0
    margin: float = 0
    specialty: str = "instalacoes_eletricas"
    labor_type: str = "eletricista"
    labor_cost_hour: float = 0
    productivity_min: float = 0
    waste_pct: float = 5
    supply_type: str = "included"

class ProBudgetCreate(BaseModel):
    title: str
    client_name: str
    client_phone: str = ""
    items: List[ProBudgetItem] = []
    risk_level: str = "medio"
    global_margin: float = 0
    notes: str = ""


# --- System Settings Endpoints ---

def _merge_system_settings(doc: Optional[dict]) -> dict:
    base = {**DEFAULT_SYSTEM_SETTINGS}
    if not doc:
        return base
    merged = {**base, **doc}
    merged["treasury_settings"] = {**base.get("treasury_settings", {}), **(doc.get("treasury_settings") or {})}
    merged["indirect_costs"] = {**base.get("indirect_costs", {}), **(doc.get("indirect_costs") or {})}
    merged["risk_levels"] = {**base.get("risk_levels", {}), **(doc.get("risk_levels") or {})}
    merged["proposal_modes"] = {**base.get("proposal_modes", {}), **(doc.get("proposal_modes") or {})}
    merged["company_info"] = {**base.get("company_info", {}), **(doc.get("company_info") or {})}
    return merged

@api_router.get("/system-settings")
async def get_system_settings(user=Depends(get_current_user)):
    s = await db.system_settings.find_one({}, {"_id": 0})
    return _merge_system_settings(s)

@api_router.put("/system-settings")
async def update_system_settings(input: SystemSettingsInput, user=Depends(get_current_user)):
    data = {k: v for k, v in input.model_dump().items() if v is not None}
    existing = await db.system_settings.find_one({})
    if existing:
        await db.system_settings.update_one({"_id": existing["_id"]}, {"$set": data})
    else:
        await db.system_settings.insert_one({**DEFAULT_SYSTEM_SETTINGS, **data})
    updated = await db.system_settings.find_one({}, {"_id": 0})
    return _merge_system_settings(updated)

@api_router.get("/specialties")
async def get_specialties(user=Depends(get_current_user)):
    return SPECIALTIES


# --- Labor Endpoints ---

@api_router.get("/labor")
async def get_labor(user=Depends(get_current_user)):
    items = await db.labor_db.find({}, {"_id": 0}).to_list(100)
    return items

@api_router.post("/labor")
async def create_labor(input: LaborInput, user=Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), **input.model_dump(), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.labor_db.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/labor/{labor_id}")
async def update_labor(labor_id: str, input: LaborInput, user=Depends(get_current_user)):
    data = input.model_dump()
    result = await db.labor_db.update_one({"id": labor_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tipo de mao de obra não encontrado")
    return await db.labor_db.find_one({"id": labor_id}, {"_id": 0})

@api_router.delete("/labor/{labor_id}")
async def delete_labor(labor_id: str, user=Depends(get_current_user)):
    r = await db.labor_db.delete_one({"id": labor_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    return {"message": "Eliminado"}


# --- Productivity Endpoints ---

@api_router.get("/productivity")
async def get_productivity(user=Depends(get_current_user)):
    items = await db.productivity_db.find({}, {"_id": 0}).to_list(500)
    return items

@api_router.post("/productivity")
async def create_productivity(input: ProductivityInput, user=Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), **input.model_dump(), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.productivity_db.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/productivity/{prod_id}")
async def update_productivity(prod_id: str, input: ProductivityInput, user=Depends(get_current_user)):
    result = await db.productivity_db.update_one({"id": prod_id}, {"$set": input.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    return await db.productivity_db.find_one({"id": prod_id}, {"_id": 0})

@api_router.delete("/productivity/{prod_id}")
async def delete_productivity(prod_id: str, user=Depends(get_current_user)):
    r = await db.productivity_db.delete_one({"id": prod_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    return {"message": "Eliminado"}


# --- Materials DB Endpoints ---

@api_router.get("/materials")
async def get_materials(user=Depends(get_current_user)):
    items = await db.materials_db.find({}, {"_id": 0}).sort("category", 1).to_list(2000)
    return items

@api_router.post("/materials")
async def create_material(input: MaterialInput, user=Depends(get_current_user)):
    doc = {
        "id": str(uuid.uuid4()), **input.model_dump(),
        "price_history": [{"price": input.purchase_price, "date": datetime.now(timezone.utc).isoformat()}],
        "price_updated_at": datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.materials_db.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/materials/{mat_id}")
async def update_material(mat_id: str, input: MaterialInput, user=Depends(get_current_user)):
    old = await db.materials_db.find_one({"id": mat_id}, {"_id": 0})
    if not old:
        raise HTTPException(status_code=404, detail="Material não encontrado")
    data = input.model_dump()
    # Keep price history
    history = old.get("price_history", [])
    if data["purchase_price"] != old.get("purchase_price", 0):
        history.append({"price": data["purchase_price"], "date": datetime.now(timezone.utc).isoformat()})
        data["price_updated_at"] = datetime.now(timezone.utc).isoformat()
    data["price_history"] = history
    await db.materials_db.update_one({"id": mat_id}, {"$set": data})
    return await db.materials_db.find_one({"id": mat_id}, {"_id": 0})

@api_router.delete("/materials/{mat_id}")
async def delete_material(mat_id: str, user=Depends(get_current_user)):
    r = await db.materials_db.delete_one({"id": mat_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    return {"message": "Eliminado"}


# --- Stock Movements ---

@api_router.post("/stock/movement")
async def create_stock_movement(input: StockMovementInput, user=Depends(get_current_user)):
    material = await db.materials_db.find_one({"id": input.material_id}, {"_id": 0})
    if not material:
        raise HTTPException(status_code=404, detail="Material não encontrado")
    if input.quantity <= 0:
        raise HTTPException(status_code=400, detail="Quantidade deve ser maior que zero")
    if input.movement_type not in ("entrada", "saida"):
        raise HTTPException(status_code=400, detail="Tipo de movimento inválido (entrada/saida)")

    current = material.get("stock_current", 0) or 0
    delta = input.quantity if input.movement_type == "entrada" else -input.quantity
    new_stock = current + delta
    if new_stock < 0:
        raise HTTPException(status_code=400, detail=f"Stock insuficiente. Disponível: {current} {material.get('unit','un')}")

    mov = {
        "id": str(uuid.uuid4()),
        "material_id": input.material_id,
        "material_name": material.get("description", ""),
        "movement_type": input.movement_type,
        "quantity": input.quantity,
        "stock_before": current,
        "stock_after": new_stock,
        "reason": input.reason,
        "obra_id": input.obra_id,
        "notes": input.notes,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
        "created_by_name": user.get("name", ""),
    }
    await db.stock_movements.insert_one(mov)
    await db.materials_db.update_one({"id": input.material_id}, {"$set": {"stock_current": new_stock}})
    mov.pop("_id", None)
    return mov

@api_router.get("/stock/movements")
async def list_stock_movements(material_id: Optional[str] = None, user=Depends(get_current_user)):
    q = {"material_id": material_id} if material_id else {}
    movs = await db.stock_movements.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return movs

@api_router.get("/stock/low")
async def list_low_stock(user=Depends(get_current_user)):
    """Materials where current stock <= min stock (and min > 0)"""
    mats = await db.materials_db.find({"stock_min": {"$gt": 0}}, {"_id": 0}).to_list(2000)
    low = [m for m in mats if (m.get("stock_current", 0) or 0) <= (m.get("stock_min", 0) or 0)]
    return low


# --- Budget Calculation Engine ---

@api_router.post("/calculate-budget")
async def calculate_budget(input: ProBudgetCreate, user=Depends(get_current_user)):
    settings = await db.system_settings.find_one({}, {"_id": 0}) or DEFAULT_SYSTEM_SETTINGS
    labor_list = await db.labor_db.find({}, {"_id": 0}).to_list(100)
    labor_map = {lb["type"]: lb for lb in labor_list}

    iva_rate = settings.get("iva_rate", 23)
    indirect_pcts = settings.get("indirect_costs", {})
    total_indirect_pct = sum(indirect_pcts.values())
    risk_pct = settings.get("risk_levels", {}).get(input.risk_level, 5)
    global_margin = input.global_margin if input.global_margin > 0 else settings.get("target_margin", 30)

    items_calc = []
    specialty_totals = {}
    total_material = 0
    total_labor = 0
    total_hours = 0
    alerts = []

    for item in input.items:
        # Material cost
        mat_cost = item.unit_cost * item.quantity * (1 + item.waste_pct / 100)

        # Labor cost
        labor_info = labor_map.get(item.labor_type, {})
        cost_h = item.labor_cost_hour if item.labor_cost_hour > 0 else labor_info.get("cost_hour", 15)
        sell_h = labor_info.get("sell_hour", 35)
        prod_min = item.productivity_min if item.productivity_min > 0 else 20
        time_hours = (item.quantity * prod_min) / 60
        labor_cost = time_hours * cost_h
        labor_sell = time_hours * sell_h

        # Direct cost
        direct = mat_cost + labor_cost

        # Item margin (use item margin if set, otherwise global)
        item_margin = item.margin if item.margin > 0 else global_margin

        # Indirect costs for this item
        indirect_val = direct * (total_indirect_pct / 100)
        risk_val = (direct + indirect_val) * (risk_pct / 100)
        subtotal = direct + indirect_val + risk_val
        margin_val = subtotal * (item_margin / 100)
        sale_price = subtotal + margin_val

        total_material += mat_cost
        total_labor += labor_cost
        total_hours += time_hours

        # Specialty aggregation
        sp = item.specialty or "instalacoes_eletricas"
        if sp not in specialty_totals:
            specialty_totals[sp] = {"material": 0, "labor": 0, "direct": 0, "sale": 0, "hours": 0}
        specialty_totals[sp]["material"] += mat_cost
        specialty_totals[sp]["labor"] += labor_cost
        specialty_totals[sp]["direct"] += direct
        specialty_totals[sp]["sale"] += sale_price
        specialty_totals[sp]["hours"] += time_hours

        # Alerts per item
        if item.unit_cost == 0:
            alerts.append({"type": "warning", "msg": f"Item '{item.name}' sem preco de material"})
        if item.productivity_min == 0:
            alerts.append({"type": "info", "msg": f"Item '{item.name}' sem produtividade definida"})
        if sale_price < direct:
            alerts.append({"type": "danger", "msg": f"Item '{item.name}' com preco venda abaixo do custo!"})

        items_calc.append({
            "name": item.name, "category": item.category, "specialty": sp,
            "quantity": item.quantity, "unit_cost": item.unit_cost,
            "waste_pct": item.waste_pct, "material_cost": round(mat_cost, 2),
            "labor_type": item.labor_type, "time_hours": round(time_hours, 2),
            "labor_cost": round(labor_cost, 2), "labor_sell": round(labor_sell, 2),
            "direct_cost": round(direct, 2), "indirect_cost": round(indirect_val, 2),
            "risk_cost": round(risk_val, 2), "margin_pct": item_margin,
            "margin_value": round(margin_val, 2), "sale_price": round(sale_price, 2),
            "supply_type": item.supply_type,
        })

    total_direct = total_material + total_labor
    total_indirect_val = total_direct * (total_indirect_pct / 100)
    total_risk_val = (total_direct + total_indirect_val) * (risk_pct / 100)
    subtotal_before_margin = total_direct + total_indirect_val + total_risk_val
    total_margin_val = subtotal_before_margin * (global_margin / 100)
    total_sale = subtotal_before_margin + total_margin_val
    iva_val = total_sale * (iva_rate / 100)

    if global_margin < settings.get("min_margin", 15):
        alerts.append({"type": "danger", "msg": f"Margem global ({global_margin}%) abaixo do minimo ({settings.get('min_margin', 15)}%)"})
    if total_sale < total_direct:
        alerts.append({"type": "danger", "msg": "Preco de venda total abaixo do custo direto!"})
    if risk_pct >= 8 and total_risk_val < total_direct * 0.05:
        alerts.append({"type": "warning", "msg": "Risco alto mas provisao insuficiente"})

    # Specialty name mapping
    sp_names = {s["id"]: s["name"] for s in SPECIALTIES}
    specialty_summary = []
    for sp_id, totals in specialty_totals.items():
        specialty_summary.append({"id": sp_id, "name": sp_names.get(sp_id, sp_id), **{k: round(v, 2) for k, v in totals.items()}})

    return {
        "items": items_calc,
        "summary": {
            "total_material": round(total_material, 2),
            "total_labor": round(total_labor, 2),
            "total_direct": round(total_direct, 2),
            "total_indirect_pct": round(total_indirect_pct, 2),
            "total_indirect": round(total_indirect_val, 2),
            "indirect_breakdown": {k: round(total_direct * v / 100, 2) for k, v in indirect_pcts.items()},
            "risk_level": input.risk_level,
            "risk_pct": risk_pct,
            "total_risk": round(total_risk_val, 2),
            "subtotal_before_margin": round(subtotal_before_margin, 2),
            "margin_pct": global_margin,
            "total_margin": round(total_margin_val, 2),
            "total_sale": round(total_sale, 2),
            "iva_rate": iva_rate,
            "iva_value": round(iva_val, 2),
            "total_with_iva": round(total_sale + iva_val, 2),
            "total_hours": round(total_hours, 2),
            "by_specialty": specialty_summary,
        },
        "alerts": alerts,
    }


# --- Alerts Endpoint ---

@api_router.get("/alerts")
async def get_alerts(user=Depends(get_current_user)):
    alerts = []
    # Check budgets without proposals
    drafts = await db.budgets.count_documents({"status": "rascunho"})
    if drafts > 0:
        alerts.append({"type": "info", "msg": f"{drafts} orcamento(s) em rascunho sem proposta gerada"})
    # Check works over budget
    works = await db.works.find({}, {"_id": 0, "title": 1, "predicted_cost": 1, "real_cost": 1, "status": 1}).to_list(100)
    for w in works:
        if w.get("real_cost", 0) > w.get("predicted_cost", 0) and w.get("predicted_cost", 0) > 0:
            alerts.append({"type": "danger", "msg": f"Obra '{w.get('title', '')}' com custo real acima do previsto"})
    # Check materials without price
    no_price = await db.materials_db.count_documents({"purchase_price": 0, "active": True})
    if no_price > 0:
        alerts.append({"type": "warning", "msg": f"{no_price} material(is) sem preco definido"})
    # Check productivities
    no_prod = await db.productivity_db.count_documents({"time_min": 0})
    if no_prod > 0:
        alerts.append({"type": "warning", "msg": f"{no_prod} item(ns) sem produtividade definida"})
    return alerts


# --- Enhanced Dashboard ---

@api_router.get("/dashboard/overview")
async def get_dashboard_overview(user=Depends(get_current_user)):
    """Dashboard consolidado (financeiro + comercial + operacional + stock + alertas)."""
    from datetime import date as _date_cls, timedelta as _td, datetime as _dt

    now = _dt.now(timezone.utc)  # noqa: F841
    today = _date_cls.today()
    y = today.year
    m = today.month
    month_prefix = f"{y:04d}-{m:02d}"

    # ---- Data
    budgets = await db.budgets.find({}, {"_id": 0}).to_list(2000)
    proposals = await db.proposals.find({}, {"_id": 0}).to_list(2000)
    works = await db.works.find({}, {"_id": 0}).to_list(2000)
    invoices = await db.invoices.find({}, {"_id": 0}).to_list(5000)
    expenses = await db.expenses.find({}, {"_id": 0}).to_list(10000)
    guides = await db.transport_guides.find({}, {"_id": 0}).to_list(5000)
    materials = await db.materials_db.find({"active": True}, {"_id": 0}).to_list(5000)
    payroll_runs = await db.payroll_runs.find({"year": y}, {"_id": 0}).to_list(500)
    fixed_costs = await db.fixed_cost_instances.find({}, {"_id": 0}).to_list(2000) if "fixed_cost_instances" in await db.list_collection_names() else []
    employee_loans = await db.employee_loans.find({}, {"_id": 0}).to_list(1000) if "employee_loans" in await db.list_collection_names() else []

    # ---- CAIXA DO MÊS
    month_received = 0.0
    for inv in invoices:
        for p in (inv.get("payments") or []):
            d = (p.get("date") or "")[:7]
            if d == month_prefix:
                month_received += float(p.get("amount", 0) or 0)

    month_expenses = 0.0
    for e in expenses:
        d = (e.get("date") or "")[:7]
        if d == month_prefix:
            month_expenses += float(e.get("value_gross", 0) or 0)

    month_payroll = sum(float(r.get("total_custo_empresa", 0) or 0) for r in payroll_runs if r.get("month") == m)

    month_saldo = month_received - month_expenses - month_payroll

    # ---- A RECEBER
    pending_amount = 0.0
    pending_count = 0
    overdue_amount = 0.0
    overdue_count = 0
    today_iso = today.isoformat()
    for inv in invoices:
        total = float(inv.get("value_total", 0) or 0)
        paid = sum(float(p.get("amount", 0) or 0) for p in (inv.get("payments") or []))
        balance = total - paid
        if balance > 0.01:
            pending_amount += balance
            pending_count += 1
            due = (inv.get("due_date") or "")[:10]
            if due and due < today_iso:
                overdue_amount += balance
                overdue_count += 1

    # ---- A PAGAR (despesas por pagar + próximos salários mês+1)
    unpaid_expenses = 0.0
    unpaid_expenses_count = 0
    for e in expenses:
        # heurística: expense com "paid"=false OR sem payment_status  E value_gross > 0 nos últimos 90 dias
        status = (e.get("payment_status") or "").lower()
        if status in ("pago", "paid"):
            continue
        try:
            d = _date_cls.fromisoformat((e.get("date") or "")[:10])
        except Exception:
            continue
        if (today - d).days <= 90:
            unpaid_expenses += float(e.get("value_gross", 0) or 0)
            unpaid_expenses_count += 1

    # Custos fixos do mês corrente ainda por pagar
    fixed_costs_pending = 0.0
    for fc in fixed_costs:
        d = (fc.get("date") or "")[:7]
        if d == month_prefix and (fc.get("status") or "").lower() not in ("pago", "paid"):
            fixed_costs_pending += float(fc.get("amount", 0) or 0)

    to_pay = unpaid_expenses + fixed_costs_pending

    # ---- COMERCIAL: propostas pendentes (enviadas mas não aceites/assinadas)
    pending_proposals = [p for p in proposals if (p.get("status") or "").lower() in ("sent", "enviada", "draft", "rascunho") and not p.get("signed_at")]
    pending_proposals_value = sum(float(p.get("final_value", 0) or 0) for p in pending_proposals)
    # Últimas 5 propostas para mostrar
    recent_proposals = sorted(proposals, key=lambda p: p.get("created_at") or "", reverse=True)[:5]

    # ---- OPERACIONAL: obras em execução
    active_works = [w for w in works if (w.get("status") or "").lower() in ("em_execucao", "em_execução")]
    active_works_value = sum(float(w.get("predicted_cost", 0) or 0) for w in active_works)

    # Obras atrasadas: em execução há > 60 dias sem fatura
    invoices_by_client = {}
    for i in invoices:
        cn = (i.get("client_name") or "").strip().lower()
        if cn:
            invoices_by_client.setdefault(cn, []).append(i)
    late_works = []
    for w in active_works:
        cn = (w.get("client_name") or "").strip().lower()
        has_invoice = bool(invoices_by_client.get(cn))
        try:
            d = _date_cls.fromisoformat((w.get("start_date") or w.get("created_at") or "")[:10])
            days = (today - d).days
        except Exception:
            days = 0
        if not has_invoice and days > 60:
            late_works.append({"id": w.get("id"), "title": w.get("title"), "days": days, "client": w.get("client_name")})

    # Guias por receber (emitidas mas não recebidas)
    pending_guides = [g for g in guides if g.get("status") in ("emitida", "em_transito")]
    pending_guides_summary = [
        {"id": g.get("id"), "number": g.get("number"), "obra_name": g.get("obra_name"), "assigned_employee_name": g.get("assigned_employee_name")}
        for g in pending_guides[:5]
    ]

    # ---- STOCK BAIXO
    low_stock = []
    for m_ in materials:
        cur = float(m_.get("stock_current", 0) or 0)
        low = float(m_.get("stock_min", 0) or 0)
        if low > 0 and cur < low:
            low_stock.append({
                "id": m_.get("id"),
                "description": m_.get("description", ""),
                "stock_current": cur,
                "stock_min": low,
                "unit": m_.get("unit", "un"),
            })
    low_stock.sort(key=lambda x: (x["stock_current"] - x["stock_min"]))
    low_stock = low_stock[:10]

    # ---- ALERTAS AGREGADOS
    alerts = []
    if overdue_count > 0:
        alerts.append({"level": "danger", "icon": "invoice", "text": f"{overdue_count} fatura(s) vencida(s) — {overdue_amount:.2f} € por receber."})
    if len(late_works) > 0:
        alerts.append({"level": "warning", "icon": "work", "text": f"{len(late_works)} obra(s) em execução há mais de 60 dias sem fatura."})
    if len(low_stock) > 0:
        alerts.append({"level": "warning", "icon": "stock", "text": f"{len(low_stock)} material(is) em stock crítico."})
    if len(pending_guides) > 3:
        alerts.append({"level": "info", "icon": "truck", "text": f"{len(pending_guides)} guia(s) emitida(s) aguardam confirmação do técnico."})

    # ---- Empréstimos activos
    active_loans = [ln for ln in employee_loans if (ln.get("status") or "").lower() in ("active", "ativo", "aberto") and float(ln.get("outstanding", ln.get("original_amount", 0)) or 0) > 0]
    loans_outstanding = sum(float(ln.get("outstanding", ln.get("original_amount", 0)) or 0) for ln in active_loans)

    # ---- Recent activity (últimas 5 movimentações)
    recent_activity = []
    for inv in sorted(invoices, key=lambda x: x.get("issue_date") or "", reverse=True)[:3]:
        recent_activity.append({"type": "invoice", "when": inv.get("issue_date"), "title": f"Fatura {inv.get('number')} · {inv.get('client_name')}", "amount": float(inv.get("value_total", 0) or 0)})
    for e in sorted(expenses, key=lambda x: x.get("date") or "", reverse=True)[:3]:
        recent_activity.append({"type": "expense", "when": e.get("date"), "title": f"Despesa · {e.get('supplier')}", "amount": -float(e.get("value_gross", 0) or 0)})
    for g in sorted(guides, key=lambda x: x.get("created_at") or "", reverse=True)[:3]:
        recent_activity.append({"type": "guide", "when": (g.get("created_at") or "")[:10], "title": f"Guia {g.get('number')} · {g.get('obra_name') or g.get('destination')}", "amount": None})
    recent_activity.sort(key=lambda x: x.get("when") or "", reverse=True)
    recent_activity = recent_activity[:8]

    return {
        "period": {"year": y, "month": m, "month_label": month_prefix},
        "highlights": {
            "cash_month": {
                "amount": round(month_saldo, 2),
                "received": round(month_received, 2),
                "expenses": round(month_expenses, 2),
                "payroll": round(month_payroll, 2),
            },
            "to_receive": {"amount": round(pending_amount, 2), "count": pending_count, "overdue_amount": round(overdue_amount, 2), "overdue_count": overdue_count},
            "to_pay": {"amount": round(to_pay, 2), "expenses_count": unpaid_expenses_count, "fixed_costs": round(fixed_costs_pending, 2)},
            "alerts": alerts,
        },
        "commercial": {
            "budgets_count": len(budgets),
            "proposals_count": len(proposals),
            "pending_proposals_count": len(pending_proposals),
            "pending_proposals_value": round(pending_proposals_value, 2),
            "recent_proposals": [
                {"id": p.get("id"), "title": p.get("title"), "client_name": p.get("client_name"), "final_value": float(p.get("final_value", 0) or 0), "status": p.get("status"), "signed_at": p.get("signed_at")}
                for p in recent_proposals
            ],
        },
        "operational": {
            "works_count": len(works),
            "active_works_count": len(active_works),
            "active_works_value": round(active_works_value, 2),
            "late_works": late_works[:5],
            "pending_guides_count": len(pending_guides),
            "pending_guides": pending_guides_summary,
        },
        "stock": {
            "materials_count": len(materials),
            "low_stock": low_stock,
            "low_stock_count": len(low_stock),
        },
        "hr": {
            "active_loans_count": len(active_loans),
            "loans_outstanding": round(loans_outstanding, 2),
        },
        "recent_activity": recent_activity,
    }


@api_router.get("/dashboard/financial")
async def get_financial_dashboard(user=Depends(get_current_user)):
    sys_settings = await db.system_settings.find_one({}, {"_id": 0}) or DEFAULT_SYSTEM_SETTINGS
    works = await db.works.find({}, {"_id": 0}).to_list(500)
    budgets = await db.budgets.find({}, {"_id": 0}).to_list(500)
    proposals = await db.proposals.find({}, {"_id": 0}).to_list(500)

    total_predicted = sum(w.get("predicted_cost", 0) for w in works)
    total_real = sum(w.get("real_cost", 0) for w in works)
    total_budget_value = sum(b.get("total_price", 0) for b in budgets)
    total_proposals_value = sum(p.get("final_value", 0) for p in proposals)

    materials_count = await db.materials_db.count_documents({"active": True})
    labor_count = await db.labor_db.count_documents({})
    productivity_count = await db.productivity_db.count_documents({})

    return {
        "settings": sys_settings,
        "totals": {
            "obras": len(works),
            "orcamentos": len(budgets),
            "propostas": len(proposals),
            "predicted_revenue": round(total_predicted, 2),
            "real_cost": round(total_real, 2),
            "estimated_profit": round(total_predicted - total_real, 2),
            "margin_pct": round(((total_predicted - total_real) / total_predicted * 100) if total_predicted > 0 else 0, 1),
            "total_budget_value": round(total_budget_value, 2),
            "total_proposals_value": round(total_proposals_value, 2),
        },
        "database": {
            "materials": materials_count,
            "labor_types": labor_count,
            "productivities": productivity_count,
        },
    }


@api_router.get("/dashboard/cashflow")
async def get_cashflow_dashboard(
    year: Optional[int] = None,
    month: Optional[int] = None,
    client: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Unified financial dashboard: entries (invoice payments) vs exits (expenses + payroll),
    monthly breakdown, top expense categories, pending collections, 30-day forecast.
    Filters: year (default current), month (optional, narrows KPIs), client (optional, narrows invoice-side metrics)."""
    from datetime import date as _date_cls, timedelta as _timedelta
    import re as _re

    now = datetime.now(timezone.utc)
    y = year or now.year
    prefix = f"{y:04d}"
    today = _date_cls.today().isoformat()
    today_dt = _date_cls.fromisoformat(today)
    horizon = (today_dt + _timedelta(days=30)).isoformat()

    client_filter = (client or "").strip()
    client_active = bool(client_filter)

    # Load invoices (optionally filtered by client)
    inv_q = {}
    if client_active:
        inv_q["client_name"] = {"$regex": _re.escape(client_filter), "$options": "i"}
    invoices = await db.invoices.find(inv_q, {"_id": 0}).to_list(5000)

    # Expenses & payroll only make sense when NOT filtering by client (they are company-wide)
    if client_active:
        expenses = []
        payroll_runs = []
    else:
        expenses = await db.expenses.find({"date": {"$regex": f"^{prefix}"}}, {"_id": 0}).to_list(5000)
        payroll_runs = await db.payroll_runs.find({"year": y}, {"_id": 0}).to_list(500)

    by_month = {m: {"entries": 0.0, "expenses": 0.0, "payroll": 0.0} for m in range(1, 13)}
    by_category = {}

    pending_collection = 0.0
    overdue_collection = 0.0
    upcoming_due = []
    upcoming_due_total = 0.0
    total_emitted_year = 0.0

    for inv in invoices:
        issue = (inv.get("issue_date") or "")[:10]
        if issue.startswith(prefix):
            total_emitted_year += float(inv.get("value_total", 0) or 0)
        paid = sum(float(p.get("amount", 0) or 0) for p in inv.get("payments", []))
        total = float(inv.get("value_total", 0) or 0)
        balance = round(total - paid, 2)
        due = (inv.get("due_date") or "")[:10]
        if balance > 0.01:
            pending_collection += balance
            if due and due < today:
                overdue_collection += balance
            if due and today <= due <= horizon:
                upcoming_due.append({
                    "id": inv.get("id"),
                    "number": inv.get("number"),
                    "client_name": inv.get("client_name"),
                    "due_date": due,
                    "balance": balance,
                })
                upcoming_due_total += balance
        for p in inv.get("payments", []):
            pdate = (p.get("date") or "")[:10]
            if pdate.startswith(prefix):
                try:
                    m = int(pdate[5:7])
                    by_month[m]["entries"] = round(by_month[m]["entries"] + float(p.get("amount", 0) or 0), 2)
                except Exception:
                    pass

    for e in expenses:
        try:
            m = int((e.get("date") or "")[5:7])
        except Exception:
            m = 0
        gross = float(e.get("value_gross", 0) or 0)
        if m:
            by_month[m]["expenses"] = round(by_month[m]["expenses"] + gross, 2)
        cat = e.get("category") or "Outros"
        by_category[cat] = round(by_category.get(cat, 0) + gross, 2)

    for r in payroll_runs:
        m = int(r.get("month", 0) or 0)
        cost = float(r.get("total_custo_empresa", 0) or 0)
        if m:
            by_month[m]["payroll"] = round(by_month[m]["payroll"] + cost, 2)

    monthly = []
    total_entries_year = 0.0
    total_exits_expenses_year = 0.0
    total_exits_payroll_year = 0.0
    for m in range(1, 13):
        d = by_month[m]
        exits = round(d["expenses"] + d["payroll"], 2)
        net = round(d["entries"] - exits, 2)
        monthly.append({
            "month": m,
            "entries": round(d["entries"], 2),
            "expenses": round(d["expenses"], 2),
            "payroll": round(d["payroll"], 2),
            "exits": exits,
            "net": net,
        })
        total_entries_year += d["entries"]
        total_exits_expenses_year += d["expenses"]
        total_exits_payroll_year += d["payroll"]

    # KPI scope: if month filter set, scope totals to that month; else full year
    if month and 1 <= month <= 12:
        scoped = by_month[month]
        scope_entries = scoped["entries"]
        scope_exp = scoped["expenses"]
        scope_pay = scoped["payroll"]
        scope_label = f"{y:04d}-{month:02d}"
    else:
        scope_entries = total_entries_year
        scope_exp = total_exits_expenses_year
        scope_pay = total_exits_payroll_year
        scope_label = f"{y:04d}"

    scope_exits = scope_exp + scope_pay
    scope_net = scope_entries - scope_exits
    scope_margin_pct = round((scope_net / scope_entries * 100) if scope_entries > 0 else 0, 1)

    current_month_num = month if (month and 1 <= month <= 12) else (now.month if now.year == y else 12)
    cm = by_month.get(current_month_num, {"entries": 0, "expenses": 0, "payroll": 0})
    cm_exits = round(cm["expenses"] + cm["payroll"], 2)
    cm_net = round(cm["entries"] - cm_exits, 2)

    # Forecast: last up-to-3 months with activity before current_month_num
    def _last_n_avg(key, n=3):
        months_with_data = [by_month[mm][key] for mm in range(1, current_month_num)
                            if (by_month[mm]["entries"] + by_month[mm]["expenses"] + by_month[mm]["payroll"]) > 0]
        tail = months_with_data[-n:] if months_with_data else []
        return round(sum(tail) / len(tail), 2) if tail else 0.0

    avg_entries = _last_n_avg("entries")
    avg_expenses = _last_n_avg("expenses")
    avg_payroll = _last_n_avg("payroll")
    projected_entries = round(avg_entries, 2)
    projected_exits = round(avg_expenses + avg_payroll, 2)
    projected_net = round(projected_entries - projected_exits, 2)

    top_categories = sorted(
        [{"category": k, "value": v} for k, v in by_category.items()],
        key=lambda x: x["value"], reverse=True
    )[:6]

    upcoming_due.sort(key=lambda x: x["due_date"])

    return {
        "year": y,
        "month": month if (month and 1 <= month <= 12) else None,
        "client": client_filter or None,
        "client_filter_active": client_active,
        "scope_label": scope_label,
        "totals": {
            "entries": round(scope_entries, 2),
            "exits_expenses": round(scope_exp, 2),
            "exits_payroll": round(scope_pay, 2),
            "exits": round(scope_exits, 2),
            "net": round(scope_net, 2),
            "margin_pct": scope_margin_pct,
            "emitted_year": round(total_emitted_year, 2),
        },
        "current_month": {
            "month": current_month_num,
            "entries": round(cm["entries"], 2),
            "expenses": round(cm["expenses"], 2),
            "payroll": round(cm["payroll"], 2),
            "exits": cm_exits,
            "net": cm_net,
        },
        "monthly": monthly,
        "top_categories": top_categories,
        "collection": {
            "pending": round(pending_collection, 2),
            "overdue": round(overdue_collection, 2),
        },
        "forecast_30d": {
            "avg_entries": avg_entries,
            "avg_expenses": avg_expenses,
            "avg_payroll": avg_payroll,
            "projected_entries": projected_entries,
            "projected_exits": projected_exits,
            "projected_net": projected_net,
            "upcoming_due_total": round(upcoming_due_total, 2),
            "upcoming_due_count": len(upcoming_due),
            "upcoming_due": upcoming_due[:20],
        },
    }


# --- Annual Financial Report (Detailed PDF source) ---

@api_router.get("/reports/annual")
async def get_annual_report(
    year: Optional[int] = None,
    client: Optional[str] = None,
    category: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Aggregates EVERYTHING for an annual financial report (PDF source):
    invoices (linha-a-linha), expenses (linha-a-linha), payroll, works,
    monthly aggregates, cumulative cashflow, top categories of expense, top clients by revenue.

    Filters:
        - year (default current)
        - client: regex case-insensitive on invoices.client_name and expenses.obra_name
        - category: exact match on expenses.category
    """
    import re as _re
    now = datetime.now(timezone.utc)
    y = year or now.year
    prefix = f"{y:04d}"

    cli = (client or "").strip()
    cat = (category or "").strip()
    has_cli = bool(cli)
    has_cat = bool(cat)

    # ---------------- INVOICES ----------------
    inv_q = {}
    if has_cli:
        inv_q["client_name"] = {"$regex": _re.escape(cli), "$options": "i"}
    invoices_all = await db.invoices.find(inv_q, {"_id": 0}).to_list(20000)
    invoices_year = []
    for inv in invoices_all:
        issue = (inv.get("issue_date") or "")[:10]
        if issue.startswith(prefix):
            invoices_year.append(inv)

    # ---------------- EXPENSES ----------------
    exp_q = {"date": {"$regex": f"^{prefix}"}}
    if has_cat:
        exp_q["category"] = cat
    # Despesas não têm cliente — quando filtra cliente, despesas viram []
    if has_cli:
        expenses_year = []
    else:
        expenses_year = await db.expenses.find(exp_q, {"_id": 0}).sort("date", 1).to_list(20000)

    # ---------------- PAYROLL ----------------
    if has_cli:
        payroll_runs = []
    else:
        payroll_runs = await db.payroll_runs.find({"year": y}, {"_id": 0}).sort("month", 1).to_list(500)

    # ---------------- WORKS ----------------
    works_q = {}
    if has_cli:
        works_q["client_name"] = {"$regex": _re.escape(cli), "$options": "i"}
    works = await db.works.find(works_q, {"_id": 0}).to_list(2000)
    works_in_progress = [w for w in works if (w.get("status") or "").lower() in ("em_execucao", "em_execução", "em execucao", "em execução")]

    # ---------------- MONTHLY BREAKDOWN ----------------
    monthly = []
    by_cat_expense = {}
    by_client_revenue = {}
    total_in_year = 0.0
    total_out_var = 0.0
    total_out_fixed = 0.0
    total_out_obra = 0.0
    total_payroll_year = 0.0
    total_emitted_year = 0.0
    total_vat_paid = 0.0     # IVA suportado em despesas
    total_vat_charged = 0.0  # IVA liquidado em faturas

    # Per-invoice loop for revenue (use payments dates for entries; issue_date for emitted)
    for inv in invoices_year:
        emitted = float(inv.get("value_total", 0) or 0)
        total_emitted_year += emitted
        total_vat_charged += float(inv.get("vat_amount", 0) or 0)
        cn = inv.get("client_name") or "—"
        by_client_revenue[cn] = round(by_client_revenue.get(cn, 0.0) + emitted, 2)

    in_by_m = {m: 0.0 for m in range(1, 13)}
    for inv in invoices_all:
        for p in (inv.get("payments") or []):
            d = (p.get("date") or "")[:10]
            if d.startswith(prefix):
                try:
                    m = int(d[5:7])
                    amt = float(p.get("amount", 0) or 0)
                    in_by_m[m] = round(in_by_m[m] + amt, 2)
                    total_in_year += amt
                except Exception:
                    pass

    exp_by_m = {m: {"var": 0.0, "fix": 0.0, "obra": 0.0} for m in range(1, 13)}
    for e in expenses_year:
        try:
            m = int((e.get("date") or "")[5:7])
        except Exception:
            continue
        gross = float(e.get("value_gross", 0) or 0)
        vat = float(e.get("vat_amount", 0) or 0)
        total_vat_paid += vat
        t = (e.get("type") or "variavel").lower()
        if t == "fixo":
            exp_by_m[m]["fix"] = round(exp_by_m[m]["fix"] + gross, 2)
            total_out_fixed += gross
        elif t == "obra":
            exp_by_m[m]["obra"] = round(exp_by_m[m]["obra"] + gross, 2)
            total_out_obra += gross
        else:
            exp_by_m[m]["var"] = round(exp_by_m[m]["var"] + gross, 2)
            total_out_var += gross
        c = e.get("category") or "Outros"
        by_cat_expense[c] = round(by_cat_expense.get(c, 0.0) + gross, 2)

    pay_by_m = {m: 0.0 for m in range(1, 13)}
    for r in payroll_runs:
        m = int(r.get("month", 0) or 0)
        cost = float(r.get("total_custo_empresa", 0) or 0)
        if m:
            pay_by_m[m] = round(pay_by_m[m] + cost, 2)
            total_payroll_year += cost

    accumulated = 0.0
    months_pt = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    for m in range(1, 13):
        entries = round(in_by_m[m], 2)
        out_var = round(exp_by_m[m]["var"], 2)
        out_fix = round(exp_by_m[m]["fix"], 2)
        out_obra = round(exp_by_m[m]["obra"], 2)
        out_pay = round(pay_by_m[m], 2)
        total_out = round(out_var + out_fix + out_obra + out_pay, 2)
        net = round(entries - total_out, 2)
        accumulated = round(accumulated + net, 2)
        monthly.append({
            "month": m,
            "month_label": months_pt[m - 1],
            "entries": entries,
            "expenses_variable": out_var,
            "expenses_fixed": out_fix,
            "expenses_obra": out_obra,
            "payroll": out_pay,
            "total_out": total_out,
            "net": net,
            "accumulated": accumulated,
        })

    total_out_year = round(total_out_var + total_out_fixed + total_out_obra + total_payroll_year, 2)
    result_year = round(total_in_year - total_out_year, 2)
    margin_pct = round((result_year / total_in_year * 100) if total_in_year > 0 else 0, 1)

    cats_sorted = sorted(
        [{"category": k, "total": v, "pct": round(v / total_out_year * 100, 1) if total_out_year > 0 else 0}
         for k, v in by_cat_expense.items()],
        key=lambda x: x["total"], reverse=True,
    )
    clients_sorted = sorted(
        [{"client": k, "total": v, "pct": round(v / total_emitted_year * 100, 1) if total_emitted_year > 0 else 0}
         for k, v in by_client_revenue.items()],
        key=lambda x: x["total"], reverse=True,
    )

    # Pending collection (todos os anos, importante para fim de ano)
    pending = 0.0
    for inv in invoices_all:
        paid = sum(float(p.get("amount", 0) or 0) for p in (inv.get("payments") or []))
        total = float(inv.get("value_total", 0) or 0)
        bal = total - paid
        if bal > 0.01:
            pending += bal

    # Slimmed-down invoice and expense rows for the report (only fields we render)
    def _slim_inv(i):
        paid = sum(float(p.get("amount", 0) or 0) for p in (i.get("payments") or []))
        total = float(i.get("value_total", 0) or 0)
        return {
            "id": i.get("id"), "number": i.get("number"),
            "issue_date": (i.get("issue_date") or "")[:10],
            "due_date": (i.get("due_date") or "")[:10],
            "client_name": i.get("client_name") or "—",
            "client_nif": i.get("client_nif") or "",
            "value_net": float(i.get("value_net", 0) or 0),
            "vat_amount": float(i.get("vat_amount", 0) or 0),
            "value_total": total,
            "paid": round(paid, 2),
            "balance": round(total - paid, 2),
            "status": i.get("status") or "—",
        }
    def _slim_exp(e):
        return {
            "id": e.get("id"),
            "date": (e.get("date") or "")[:10],
            "supplier": e.get("supplier") or "—",
            "nif": e.get("nif") or "",
            "invoice_number": e.get("invoice_number") or "",
            "category": e.get("category") or "Outros",
            "type": e.get("type") or "variavel",
            "obra_name": e.get("obra_name") or "",
            "value_net": float(e.get("value_net", 0) or 0),
            "vat_amount": float(e.get("vat_amount", 0) or 0),
            "value_gross": float(e.get("value_gross", 0) or 0),
        }
    def _slim_work(w):
        return {
            "id": w.get("id"), "title": w.get("title") or "—",
            "client_name": w.get("client_name") or "—",
            "status": w.get("status") or "—",
            "predicted_cost": float(w.get("predicted_cost", 0) or 0),
            "real_cost": float(w.get("real_cost", 0) or 0),
            "start_date": w.get("start_date") or "",
            "end_date": w.get("end_date") or "",
        }
    def _slim_pay(r):
        return {
            "year": r.get("year"), "month": r.get("month"),
            "total_iliquido": float(r.get("total_iliquido", 0) or 0),
            "total_liquido": float(r.get("total_liquido", 0) or 0),
            "total_ss_empresa": float(r.get("total_ss_empresa", 0) or 0),
            "total_custo_empresa": float(r.get("total_custo_empresa", 0) or 0),
            "employees_count": len(r.get("employees", []) or []),
            "status": r.get("status") or "—",
        }

    invoices_year.sort(key=lambda i: (i.get("issue_date") or ""))
    return {
        "year": y,
        "filters": {
            "client": cli or None,
            "category": cat or None,
            "has_client_filter": has_cli,
            "has_category_filter": has_cat,
        },
        "scope_label": f"{y:04d}" + (f" · {cli}" if has_cli else "") + (f" · {cat}" if has_cat else ""),
        "kpis": {
            "total_in": round(total_in_year, 2),
            "total_emitted": round(total_emitted_year, 2),
            "total_out": total_out_year,
            "total_out_variable": round(total_out_var, 2),
            "total_out_fixed": round(total_out_fixed, 2),
            "total_out_obra": round(total_out_obra, 2),
            "total_payroll": round(total_payroll_year, 2),
            "result": result_year,
            "margin_pct": margin_pct,
            "vat_paid": round(total_vat_paid, 2),
            "vat_charged": round(total_vat_charged, 2),
            "vat_balance": round(total_vat_charged - total_vat_paid, 2),
            "pending_total": round(pending, 2),
            "invoices_count": len(invoices_year),
            "expenses_count": len(expenses_year),
            "works_count": len(works),
            "works_in_progress_count": len(works_in_progress),
        },
        "monthly": monthly,
        "categories_expense": cats_sorted,
        "clients_revenue": clients_sorted,
        "invoices": [_slim_inv(i) for i in invoices_year],
        "expenses": [_slim_exp(e) for e in expenses_year],
        "payroll_runs": [_slim_pay(r) for r in payroll_runs],
        "works": [_slim_work(w) for w in works],
        "works_in_progress": [_slim_work(w) for w in works_in_progress],
        "generated_at": now.isoformat(),
    }


# --- Seed Professional Data ---

async def seed_professional_data():
    # Seed labor
    if await db.labor_db.count_documents({}) == 0:
        for labor_item in DEFAULT_LABOR:
            labor_doc = {**labor_item, "created_at": datetime.now(timezone.utc).isoformat()}
            await db.labor_db.insert_one(labor_doc)
        logger.info(f"Labor DB seeded: {len(DEFAULT_LABOR)} types")

    # Seed productivity
    if await db.productivity_db.count_documents({}) == 0:
        for prod in DEFAULT_PRODUCTIVITIES:
            p_doc = {"id": str(uuid.uuid4()), **prod, "notes": "", "created_at": datetime.now(timezone.utc).isoformat()}
            await db.productivity_db.insert_one(p_doc)
        logger.info(f"Productivity DB seeded: {len(DEFAULT_PRODUCTIVITIES)} items")

    # Seed system settings
    if await db.system_settings.count_documents({}) == 0:
        await db.system_settings.insert_one(DEFAULT_SYSTEM_SETTINGS)
        logger.info("System settings seeded")

    # Seed materials from catalog
    if await db.materials_db.count_documents({}) == 0:
        count = 0
        for cat in CATEGORIES_CATALOG:
            for item in cat["items"]:
                doc = {
                    "id": str(uuid.uuid4()), "code": "", "description": item["name"],
                    "category": cat["name"], "subcategory": "", "brand": "", "supplier": "",
                    "unit": item.get("unit", "unidade"), "purchase_price": 0, "market_price": 0,
                    "waste_pct": 5, "notes": "", "active": True,
                    "price_history": [], "price_updated_at": "",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }
                await db.materials_db.insert_one(doc)
                count += 1
        logger.info(f"Materials DB seeded: {count} items from catalog")


# ============================================================
# FASE 2 - NEGOTIATION, HISTORY, USER MANAGEMENT
# ============================================================

# --- Negotiation / Discount Simulation ---

class NegotiationInput(BaseModel):
    budget_id: str = ""
    original_price: float = 0
    discount_type: str = "percentage"  # percentage or value
    discount_value: float = 0
    items: List[ProBudgetItem] = []
    risk_level: str = "medio"
    global_margin: float = 0

@api_router.post("/simulate-negotiation")
async def simulate_negotiation(input: NegotiationInput, user=Depends(get_current_user)):
    settings = await db.system_settings.find_one({}, {"_id": 0}) or DEFAULT_SYSTEM_SETTINGS
    min_margin = settings.get("min_margin", 15)

    # Calculate real cost from items
    labor_list = await db.labor_db.find({}, {"_id": 0}).to_list(100)
    labor_map = {lb["type"]: lb for lb in labor_list}

    indirect_pcts = settings.get("indirect_costs", {})
    total_indirect_pct = sum(indirect_pcts.values())
    risk_pct = settings.get("risk_levels", {}).get(input.risk_level, 5)

    total_cost = 0
    for item in input.items:
        mat_cost = item.unit_cost * item.quantity * (1 + item.waste_pct / 100)
        labor_info = labor_map.get(item.labor_type, {})
        cost_h = item.labor_cost_hour if item.labor_cost_hour > 0 else labor_info.get("cost_hour", 15)
        prod_min = item.productivity_min if item.productivity_min > 0 else 20
        time_hours = (item.quantity * prod_min) / 60
        labor_cost = time_hours * cost_h
        total_cost += mat_cost + labor_cost

    total_indirect = total_cost * (total_indirect_pct / 100)
    total_risk = (total_cost + total_indirect) * (risk_pct / 100)
    break_even = total_cost + total_indirect + total_risk
    min_price = break_even * (1 + min_margin / 100)

    original = input.original_price if input.original_price > 0 else break_even * 1.3

    if input.discount_type == "percentage":
        discount_amount = original * (input.discount_value / 100)
    else:
        discount_amount = input.discount_value

    final_price = original - discount_amount
    final_margin_pct = ((final_price - break_even) / break_even * 100) if break_even > 0 else 0
    profit = final_price - break_even
    max_discount = original - min_price
    max_discount_pct = (max_discount / original * 100) if original > 0 else 0

    alerts = []
    if final_price < break_even:
        alerts.append({"type": "danger", "msg": "PRECO ABAIXO DO CUSTO! A obra tera prejuizo."})
    elif final_margin_pct < min_margin:
        alerts.append({"type": "warning", "msg": f"Margem ({final_margin_pct:.1f}%) abaixo do minimo ({min_margin}%)"})
    if discount_amount > max_discount:
        alerts.append({"type": "danger", "msg": f"Desconto excede o maximo seguro de {formatEuro_py(max_discount)}"})

    return {
        "original_price": round(original, 2),
        "discount_amount": round(discount_amount, 2),
        "discount_pct": round(input.discount_value if input.discount_type == "percentage" else (discount_amount / original * 100) if original > 0 else 0, 2),
        "final_price": round(final_price, 2),
        "break_even": round(break_even, 2),
        "total_cost": round(total_cost, 2),
        "total_indirect": round(total_indirect, 2),
        "total_risk": round(total_risk, 2),
        "profit": round(profit, 2),
        "final_margin_pct": round(final_margin_pct, 1),
        "min_margin": min_margin,
        "min_price": round(min_price, 2),
        "max_discount": round(max_discount, 2),
        "max_discount_pct": round(max_discount_pct, 1),
        "alerts": alerts,
    }

def formatEuro_py(v):
    return f"{v:,.2f} EUR"


# --- Work History & Comparison ---

class WorkHistoryEntry(BaseModel):
    date: str
    description: str
    hours: float = 0
    cost: float = 0

@api_router.post("/works/{work_id}/history")
async def add_work_history(work_id: str, entry: WorkHistoryEntry, user=Depends(get_current_user)):
    work = await db.works.find_one({"id": work_id})
    if not work:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    history_entry = {"id": str(uuid.uuid4()), **entry.model_dump(), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.works.update_one({"id": work_id}, {"$push": {"history": history_entry}})
    updated = await db.works.find_one({"id": work_id}, {"_id": 0})
    return updated

@api_router.get("/works/{work_id}/comparison")
async def get_work_comparison(work_id: str, user=Depends(get_current_user)):
    work = await db.works.find_one({"id": work_id}, {"_id": 0})
    if not work:
        raise HTTPException(status_code=404, detail="Obra não encontrada")
    predicted = work.get("predicted_cost", 0)
    real = work.get("real_cost", 0)
    history = work.get("history", [])
    total_hours = sum(h.get("hours", 0) for h in history)
    total_history_cost = sum(h.get("cost", 0) for h in history)
    margin_predicted = predicted * 0.3 if predicted > 0 else 0
    margin_real = predicted - real if predicted > 0 else 0
    return {
        "work": work,
        "comparison": {
            "predicted_cost": round(predicted, 2),
            "real_cost": round(real, 2),
            "cost_difference": round(predicted - real, 2),
            "cost_deviation_pct": round(((real - predicted) / predicted * 100) if predicted > 0 else 0, 1),
            "margin_predicted": round(margin_predicted, 2),
            "margin_real": round(margin_real, 2),
            "total_hours_logged": round(total_hours, 1),
            "total_cost_logged": round(total_history_cost, 2),
            "history_entries": len(history),
        }
    }

@api_router.get("/works/comparison-all")
async def get_all_works_comparison(user=Depends(get_current_user)):
    works = await db.works.find({}, {"_id": 0}).to_list(500)
    results = []
    for w in works:
        predicted = w.get("predicted_cost", 0)
        real = w.get("real_cost", 0)
        results.append({
            "id": w["id"], "title": w.get("title", ""), "client_name": w.get("client_name", ""),
            "status": w.get("status", ""), "predicted_cost": predicted, "real_cost": real,
            "deviation": round(real - predicted, 2) if predicted > 0 else 0,
            "deviation_pct": round(((real - predicted) / predicted * 100) if predicted > 0 else 0, 1),
            "margin_real_pct": round(((predicted - real) / predicted * 100) if predicted > 0 else 0, 1),
        })
    return results


# --- User Management ---

# Lista completa de módulos disponíveis na app (usados para permissões granulares)
ALL_MODULES = [
    "dashboard",       # /
    "orcamentos",      # /orcamentos
    "propostas",       # /propostas
    "obras",           # /obras
    "pipeline",        # /pipeline
    "materiais",       # /materiais
    "transporte_guias",# /guias-transporte
    "faturas",         # /faturas
    "despesas",        # /despesas
    "custos_fixos",    # /custos-fixos
    "financeiro",      # /financeiro
    "ponto_equilibrio",# /ponto-equilibrio
    "contabilista",    # /contabilista
    "salarios",        # /salarios
    "funcionarios",    # /funcionarios
    "assiduidade",     # /assiduidade
    "agenda",          # /agenda
    "biblioteca",      # /biblioteca
    "relatorios",      # /relatorios
    "tech_portal",     # /tech (portal do técnico)
    "configuracoes",   # /definicoes, /config-*
    "utilizadores",    # /utilizadores (só admin real)
]


class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: str = "consulta"
    module_permissions: Optional[dict] = None   # {module_key: bool}

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    email: Optional[str] = None
    module_permissions: Optional[dict] = None
    password: Optional[str] = None   # opcional — só se admin quiser resetar

VALID_ROLES = ["admin", "orcamentista", "comercial", "tecnico", "consulta"]

# Defaults por role (usado se admin não especificar módulos)
def default_modules_for_role(role: str) -> dict:
    all_true = {m: True for m in ALL_MODULES}
    if role == "admin":
        return all_true
    if role == "orcamentista":
        return {**{m: False for m in ALL_MODULES},
                "dashboard": True, "orcamentos": True, "propostas": True, "obras": True,
                "materiais": True, "pipeline": True, "agenda": True, "biblioteca": True}
    if role == "comercial":
        return {**{m: False for m in ALL_MODULES},
                "dashboard": True, "propostas": True, "obras": True, "pipeline": True, "agenda": True}
    if role == "tecnico":
        return {**{m: False for m in ALL_MODULES}, "tech_portal": True}
    # consulta
    return {**{m: False for m in ALL_MODULES}, "dashboard": True, "propostas": True}

ROLE_PERMISSIONS = {
    "admin": {"view_costs": True, "view_margins": True, "view_prices": True, "edit_budgets": True, "edit_settings": True, "manage_users": True},
    "orcamentista": {"view_costs": True, "view_margins": True, "view_prices": True, "edit_budgets": True, "edit_settings": False, "manage_users": False},
    "comercial": {"view_costs": False, "view_margins": False, "view_prices": True, "edit_budgets": False, "edit_settings": False, "manage_users": False},
    "tecnico": {"view_costs": False, "view_margins": False, "view_prices": False, "edit_budgets": False, "edit_settings": False, "manage_users": False},
    "consulta": {"view_costs": False, "view_margins": False, "view_prices": True, "edit_budgets": False, "edit_settings": False, "manage_users": False},
}

@api_router.get("/users")
async def get_users(user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Sem permissao")
    users = await db.users.find({}, {"password_hash": 0}).to_list(100)
    return [{
        "id": str(u["_id"]),
        "email": u["email"],
        "name": u["name"],
        "role": u.get("role", "consulta"),
        "module_permissions": u.get("module_permissions") or default_modules_for_role(u.get("role", "consulta")),
        "created_at": u.get("created_at", "")
    } for u in users]

@api_router.post("/users")
async def create_user(input: UserCreate, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Sem permissao")
    if input.role not in VALID_ROLES:
        raise HTTPException(status_code=400, detail=f"Role invalido. Usar: {VALID_ROLES}")
    existing = await db.users.find_one({"email": input.email.lower().strip()})
    if existing:
        raise HTTPException(status_code=400, detail="Email já existe")
    # Se não vier módulos, usa default do role
    modules = input.module_permissions or default_modules_for_role(input.role)
    # Garante que só chaves válidas ficam
    modules = {k: bool(v) for k, v in modules.items() if k in ALL_MODULES}
    doc = {
        "email": input.email.lower().strip(),
        "password_hash": hash_password(input.password),
        "name": input.name,
        "role": input.role,
        "module_permissions": modules,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.users.insert_one(doc)
    return {"id": str(result.inserted_id), "email": doc["email"], "name": doc["name"], "role": doc["role"], "module_permissions": modules}

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, input: UserUpdate, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Sem permissao")
    data = {k: v for k, v in input.model_dump().items() if v is not None}
    if "role" in data and data["role"] not in VALID_ROLES:
        raise HTTPException(status_code=400, detail=f"Role invalido. Usar: {VALID_ROLES}")
    if "module_permissions" in data:
        data["module_permissions"] = {k: bool(v) for k, v in data["module_permissions"].items() if k in ALL_MODULES}
    if "password" in data and data["password"]:
        data["password_hash"] = hash_password(data.pop("password"))
    else:
        data.pop("password", None)
    result = await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Utilizador não encontrado")
    return {"message": "Atualizado"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, user=Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Sem permissao")
    if user_id == user.get("id"):
        raise HTTPException(status_code=400, detail="Não pode eliminar a si proprio")
    result = await db.users.delete_one({"_id": ObjectId(user_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    return {"message": "Eliminado"}

@api_router.get("/roles")
async def get_roles(user=Depends(get_current_user)):
    return {
        "roles": VALID_ROLES,
        "permissions": ROLE_PERMISSIONS,
        "all_modules": ALL_MODULES,
        "default_modules_per_role": {r: default_modules_for_role(r) for r in VALID_ROLES},
    }


# ============================================================
# FASE 3 - EXCEL, VERSIONING, TEMPLATES, FAVORITES
# ============================================================

# --- Excel Export ---

@api_router.get("/budgets/{budget_id}/export-excel")
async def export_budget_excel(budget_id: str, user=Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    budget = await db.budgets.find_one({"id": budget_id}, {"_id": 0})
    if not budget:
        raise HTTPException(status_code=404, detail="Orçamento não encontrado")

    wb = Workbook()
    ws = wb.active
    ws.title = "Orcamento"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="FACC15", end_color="FACC15", fill_type="solid")
    yellow_font = Font(bold=True, color="FACC15", size=12)

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Orcamento: {budget['title']}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Cliente: {budget['client_name']}"
    ws["A3"] = f"Telefone: {budget.get('client_phone', '')}"
    ws["A4"] = f"Data: {datetime.now(timezone.utc).strftime('%d/%m/%Y')}"

    # Headers
    headers = ["Categoria", "Item", "Quantidade", "Custo Unit. (EUR)", "Margem", "Preco Unit. (EUR)", "Total (EUR)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Items
    for i, item in enumerate(budget.get("items", []), 7):
        ws.cell(row=i, column=1, value=item.get("category", ""))
        ws.cell(row=i, column=2, value=item.get("name", ""))
        ws.cell(row=i, column=3, value=item.get("quantity", 0))
        ws.cell(row=i, column=4, value=item.get("unit_cost", 0))
        ws.cell(row=i, column=5, value=f"{item.get('margin', 0) * 100:.0f}%")
        pvp = item.get("unit_cost", 0) * (1 + item.get("margin", 0))
        ws.cell(row=i, column=6, value=round(pvp, 2))
        ws.cell(row=i, column=7, value=round(pvp * item.get("quantity", 0), 2))

    # Totals
    last_row = 7 + len(budget.get("items", []))
    ws.cell(row=last_row + 1, column=6, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=last_row + 1, column=7, value=budget.get("total_price", 0)).font = yellow_font

    # Column widths
    for col_widths in [(1, 20), (2, 40), (3, 12), (4, 15), (5, 10), (6, 15), (7, 15)]:
        ws.column_dimensions[chr(64 + col_widths[0])].width = col_widths[1]

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"orcamento_{budget['title'].replace(' ', '_')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           headers={"Content-Disposition": f"attachment; filename={filename}"})


# --- Excel Import ---

@api_router.post("/budgets/import-excel")
async def import_budget_excel(file: UploadFile = File(...), user=Depends(get_current_user)):
    from openpyxl import load_workbook
    import io
    import re

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Ficheiro deve ser .xlsx ou .xls")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    def try_float(val, default=0):
        if val is None:
            return default
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().replace(',', '.').replace('€', '').replace('EUR', '').replace('%', '').strip()
        s = re.sub(r'[^\d.]', '', s)
        try:
            return float(s) if s else default
        except (ValueError, TypeError):
            return default

    def is_numeric(val):
        if val is None:
            return False
        if isinstance(val, (int, float)):
            return True
        try:
            float(str(val).strip().replace(',', '.'))
            return True
        except (ValueError, TypeError):
            return False

    # Read all rows into memory for analysis
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    if not all_rows:
        raise HTTPException(status_code=400, detail="Ficheiro vazio")

    logger.info(f"Excel import: {len(all_rows)} rows, {len(all_rows[0])} cols in '{file.filename}'")
    # Log first 3 rows for debug
    for ri, row in enumerate(all_rows[:3]):
        logger.info(f"  Row {ri}: {row}")

    # STEP 1: Detect header row (must contain clear header keywords, not data)
    UNIT_TOKENS = {'un', 'un.', 'und', 'und.', 'unid', 'uni', 'pc', 'pcs', 'mt', 'm', 'ml',
                   'm2', 'm²', 'm3', 'm³', 'kg', 'g', 'h', 'hr', 'vg', 'gl', 'cj', 'kit',
                   'lt', 'l', 'cx', 'par', 'rolo', 'saco'}

    header_keywords = {
        'code': ['codigo', 'cod.', 'cod', 'ref.', 'referencia', 'artigo', 'code'],
        'category': ['categoria', 'grupo', 'especialidade'],
        'name': ['descrição', 'descrição', 'designacao', 'designação', 'descr', 'denominacao', 'descritivo'],
        'unit': ['unidade', 'un.', 'und.', 'unid.'],
        'quantity': ['quantidade', 'qtd', 'qtd.', 'qty', 'quant.', 'quant'],
        'cost': ['preco', 'preço', 'custo', 'valor unit', 'p.unit', 'p.u.', 'pu', 'unitario', 'unitário', 'unit price'],
        'total': ['total', 'subtotal', 'parcial', 'valor total'],
        'margin': ['margem', 'margin', 'markup'],
    }

    header_map = {}
    header_row_idx = -1

    def cell_str(v):
        return str(v or '').lower().strip()

    for ri, row in enumerate(all_rows[:10]):
        if not row or not any(row):
            continue
        row_str = [cell_str(c) for c in row]
        # Header row MUST have multiple short non-numeric textual cells
        text_cells = [c for c in row_str if c and not is_numeric(c) and len(c) < 40]
        if len(text_cells) < 2:
            continue

        temp_map = {}
        for col_idx, cell_val in enumerate(row_str):
            if not cell_val or len(cell_val) > 40 or is_numeric(cell_val):
                continue
            for field, keywords in header_keywords.items():
                if field in temp_map:
                    continue
                if any(cell_val == kw or cell_val.startswith(kw) for kw in keywords):
                    temp_map[field] = col_idx
                    break

        # Must explicitly find a "name/descrição" keyword and at least one other header
        if 'name' in temp_map and len(temp_map) >= 2:
            header_map = temp_map
            header_row_idx = ri
            break

    logger.info(f"Excel import: header at row {header_row_idx}, map: {header_map}")

    # STEP 2: If no header, auto-detect by analyzing ALL data rows (column statistics)
    data_start = header_row_idx + 1 if header_row_idx >= 0 else 0

    if not header_map:
        max_cols = max((len(r) for r in all_rows), default=0)
        col_stats = []  # per column: {numeric, text, unit_match, avg_len, total}
        for ci in range(max_cols):
            stats = {'numeric': 0, 'text': 0, 'unit': 0, 'avg_len': 0, 'total': 0, 'int_small': 0}
            lens = []
            for row in all_rows[data_start:]:
                if ci >= len(row):
                    continue
                v = row[ci]
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue
                stats['total'] += 1
                if isinstance(v, (int, float)):
                    stats['numeric'] += 1
                    if float(v) == int(v) and 0 < float(v) <= 9999:
                        stats['int_small'] += 1
                elif isinstance(v, str):
                    s = v.strip()
                    if is_numeric(s):
                        stats['numeric'] += 1
                    else:
                        stats['text'] += 1
                        lens.append(len(s))
                        if s.lower() in UNIT_TOKENS:
                            stats['unit'] += 1
            if lens:
                stats['avg_len'] = sum(lens) / len(lens)
            col_stats.append(stats)

        logger.info(f"Excel import: col_stats={col_stats}")

        # Unit column = column where most text values are unit tokens
        unit_col = None
        best_unit_ratio = 0
        for ci, s in enumerate(col_stats):
            if s['text'] > 0 and s['unit'] / s['text'] >= 0.5:
                ratio = s['unit'] / max(s['total'], 1)
                if ratio > best_unit_ratio:
                    best_unit_ratio = ratio
                    unit_col = ci

        # Name column = non-numeric column with largest avg text length (exclude unit col)
        name_col = None
        best_len = 0
        for ci, s in enumerate(col_stats):
            if ci == unit_col:
                continue
            if s['avg_len'] > best_len and s['text'] >= s['numeric'] and s['avg_len'] > 10:
                best_len = s['avg_len']
                name_col = ci

        # Quantity column = first numeric column after the name col (or after unit col)
        anchor = unit_col if unit_col is not None else name_col
        qty_col = None
        if anchor is not None:
            for ci in range(anchor + 1, len(col_stats)):
                s = col_stats[ci]
                if s['numeric'] > 0 and s['numeric'] >= s['text']:
                    qty_col = ci
                    break
        if qty_col is None:
            # Fallback: any numeric column
            for ci, s in enumerate(col_stats):
                if ci in (name_col,) or ci == unit_col:
                    continue
                if s['numeric'] > 0:
                    qty_col = ci
                    break

        # Cost column = next numeric column after quantity
        cost_col = None
        if qty_col is not None:
            for ci in range(qty_col + 1, len(col_stats)):
                s = col_stats[ci]
                if s['numeric'] > 0:
                    cost_col = ci
                    break

        if name_col is not None:
            header_map['name'] = name_col
        if unit_col is not None:
            header_map['unit'] = unit_col
        if qty_col is not None:
            header_map['quantity'] = qty_col
        if cost_col is not None:
            header_map['cost'] = cost_col

        logger.info(f"Excel import: auto-detected map: {header_map} (unit_col={unit_col}, name_col={name_col}, qty_col={qty_col}, cost_col={cost_col})")

    # STEP 3: Extract items
    items = []
    name_idx = header_map.get('name')
    code_idx = header_map.get('code')
    cat_idx = header_map.get('category')
    unit_idx = header_map.get('unit')
    qty_idx = header_map.get('quantity')
    cost_idx = header_map.get('cost')
    margin_idx = header_map.get('margin')

    for row in all_rows[data_start:]:
        if not row or not any(row):
            continue
        cells = list(row)

        # Get name
        name = ''
        if name_idx is not None and name_idx < len(cells) and cells[name_idx]:
            name = str(cells[name_idx]).strip()
        if not name and code_idx is not None and code_idx < len(cells) and cells[code_idx]:
            name = str(cells[code_idx]).strip()
        if not name:
            # Fallback: longest text cell in row, excluding unit tokens
            longest = ''
            for cv in cells:
                if cv and isinstance(cv, str):
                    s = cv.strip()
                    if len(s) > len(longest) and s.lower() not in UNIT_TOKENS and not is_numeric(s):
                        longest = s
            name = longest
        if not name or len(name) < 2:
            continue

        # Get unit
        unit = ''
        if unit_idx is not None and unit_idx < len(cells) and cells[unit_idx]:
            unit = str(cells[unit_idx]).strip()

        # Get quantity
        qty = 1
        if qty_idx is not None and qty_idx < len(cells):
            qty = try_float(cells[qty_idx], 1)
        if qty <= 0:
            qty = 1

        # Get cost
        cost = 0
        if cost_idx is not None and cost_idx < len(cells):
            cost = try_float(cells[cost_idx], 0)

        # Get category
        category = ''
        if cat_idx is not None and cat_idx < len(cells) and cells[cat_idx]:
            val = cells[cat_idx]
            if isinstance(val, str) and not is_numeric(val):
                category = val.strip()

        # Get margin
        margin = 0.6
        if margin_idx is not None and margin_idx < len(cells):
            m = try_float(cells[margin_idx], 0)
            margin = m / 100 if m > 1 else (m if m > 0 else 0.6)

        items.append({"category": category, "name": name, "unit": unit, "quantity": qty, "unit_cost": cost, "margin": margin, "discount_type": "percentage", "discount_value": 0})
        logger.info(f"  Item: '{name}' qty={qty} cost={cost}")

    if not items:
        raise HTTPException(status_code=400, detail="Nenhum item encontrado. Verifique que o Excel tem colunas com descrição e quantidade.")

    total_cost, total_price = calc_budget_totals(items, "percentage", 0)

    doc = {
        "id": str(uuid.uuid4()),
        "title": f"Importado: {file.filename}",
        "client_name": "A definir",
        "client_phone": "",
        "items": items,
        "discount_type": "percentage",
        "discount_value": 0,
        "payment_methods": [],
        "payment_split": "",
        "payment_notes": "",
        "total_cost": total_cost,
        "total_price": total_price,
        "status": "rascunho",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
    }
    await db.budgets.insert_one(doc)
    doc.pop("_id", None)
    logger.info(f"Excel imported: {len(items)} items, header_map={header_map}")
    return doc


# --- Budget Versioning ---

@api_router.post("/budgets/{budget_id}/save-version")
async def save_budget_version(budget_id: str, user=Depends(get_current_user)):
    budget = await db.budgets.find_one({"id": budget_id}, {"_id": 0})
    if not budget:
        raise HTTPException(status_code=404, detail="Orçamento não encontrado")

    versions = await db.budget_versions.find({"budget_id": budget_id}).to_list(100)
    version_num = len(versions) + 1

    version_doc = {
        "id": str(uuid.uuid4()),
        "budget_id": budget_id,
        "version": version_num,
        "snapshot": budget,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
    }
    await db.budget_versions.insert_one(version_doc)
    version_doc.pop("_id", None)
    return version_doc

@api_router.get("/budgets/{budget_id}/versions")
async def get_budget_versions(budget_id: str, user=Depends(get_current_user)):
    versions = await db.budget_versions.find({"budget_id": budget_id}, {"_id": 0}).sort("version", -1).to_list(100)
    return versions

@api_router.get("/budgets/{budget_id}/versions/{version_id}")
async def get_budget_version(budget_id: str, version_id: str, user=Depends(get_current_user)):
    version = await db.budget_versions.find_one({"id": version_id, "budget_id": budget_id}, {"_id": 0})
    if not version:
        raise HTTPException(status_code=404, detail="Versão não encontrada")
    return version

@api_router.post("/budgets/{budget_id}/duplicate")
async def duplicate_budget(budget_id: str, user=Depends(get_current_user)):
    budget = await db.budgets.find_one({"id": budget_id}, {"_id": 0})
    if not budget:
        raise HTTPException(status_code=404, detail="Orçamento não encontrado")

    new_doc = {**budget}
    new_doc["id"] = str(uuid.uuid4())
    new_doc["title"] = f"{budget['title']} (copia)"
    new_doc["status"] = "rascunho"
    new_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    new_doc["created_by"] = user["id"]
    await db.budgets.insert_one(new_doc)
    new_doc.pop("_id", None)
    return new_doc


# --- Text Templates ---

class TextTemplateInput(BaseModel):
    name: str
    category: str = "geral"
    content: str = ""

@api_router.get("/text-templates")
async def get_text_templates(user=Depends(get_current_user)):
    templates = await db.text_templates.find({}, {"_id": 0}).to_list(200)
    if not templates:
        defaults = [
            {"id": str(uuid.uuid4()), "name": "Objeto da Empreitada", "category": "proposta", "content": "Fornecimento, montagem, instalacao e ensaio de todas as instalacoes eletricas e de telecomunicacoes, conforme mapa de quantidades anexo."},
            {"id": str(uuid.uuid4()), "name": "Exclusao Fornecimento Dono Obra", "category": "exclusoes", "content": "Exclui-se do presente valor o fornecimento dos materiais/equipamentos a definir pelo Dono de Obra, mantendo-se incluidos, quando aplicavel, os respetivos trabalhos de instalacao, ligacao e ensaio."},
            {"id": str(uuid.uuid4()), "name": "Exclusao Geral", "category": "exclusoes", "content": "Excluem-se trabalhos de construcao civil, pintura, reposicao de pavimentos e quaisquer outros trabalhos nao mencionados no mapa de quantidades."},
            {"id": str(uuid.uuid4()), "name": "Prazo Standard", "category": "prazo", "content": "O prazo de execucao estimado e de [X] dias uteis, contados a partir da data de adjudicacao e disponibilizacao dos espacos para intervencao."},
            {"id": str(uuid.uuid4()), "name": "Garantia Standard", "category": "garantia", "content": "Garantia de [X] anos sobre a mao de obra e de acordo com a garantia do fabricante para os materiais fornecidos."},
            {"id": str(uuid.uuid4()), "name": "Observacoes Tecnicas", "category": "tecnico", "content": "Todos os trabalhos serao executados por tecnicos certificados, de acordo com as normas RTIEBT e regulamentos em vigor. Sera emitido certificado de conformidade no final da obra."},
            {"id": str(uuid.uuid4()), "name": "Condicoes ITED", "category": "tecnico", "content": "As instalacoes ITED serao executadas por instaladores certificados ANACOM, com emissao de ficha tecnica e certificacao obrigatoria."},
            {"id": str(uuid.uuid4()), "name": "Condicoes CCTV", "category": "tecnico", "content": "O sistema de videovigilancia sera instalado em conformidade com a legislacao vigente (Lei 34/2013). A configuracao e registo junto da CNPD e responsabilidade do cliente."},
        ]
        for t in defaults:
            t["created_at"] = datetime.now(timezone.utc).isoformat()
            await db.text_templates.insert_one(t)
        return [{k: v for k, v in t.items() if k != "_id"} for t in defaults]
    return templates

@api_router.post("/text-templates")
async def create_text_template(input: TextTemplateInput, user=Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), **input.model_dump(), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.text_templates.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/text-templates/{template_id}")
async def update_text_template(template_id: str, input: TextTemplateInput, user=Depends(get_current_user)):
    result = await db.text_templates.update_one({"id": template_id}, {"$set": input.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Template não encontrado")
    return await db.text_templates.find_one({"id": template_id}, {"_id": 0})

@api_router.delete("/text-templates/{template_id}")
async def delete_text_template(template_id: str, user=Depends(get_current_user)):
    r = await db.text_templates.delete_one({"id": template_id})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    return {"message": "Eliminado"}


# --- Favorite Items ---

@api_router.get("/favorites")
async def get_favorites(user=Depends(get_current_user)):
    favs = await db.favorites.find({"user_id": user["id"]}, {"_id": 0}).to_list(200)
    return favs

@api_router.post("/favorites")
async def add_favorite(item: BudgetItemModel, user=Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), "user_id": user["id"], **item.model_dump(), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.favorites.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/favorites/{fav_id}")
async def remove_favorite(fav_id: str, user=Depends(get_current_user)):
    r = await db.favorites.delete_one({"id": fav_id, "user_id": user["id"]})
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nao encontrado")
    return {"message": "Removido"}

async def seed_admin():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@obelisco.pt")
    admin_password = os.environ.get("ADMIN_PASSWORD", "obelisco2024")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "Admin",
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Admin criado: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}}
        )
        logger.info("Password do admin atualizada")
    await db.users.create_index("email", unique=True)

@app.on_event("startup")
async def startup():
    await seed_admin()
    await seed_professional_data()
    await db.expenses.create_index(
        "hard_dedupe_key",
        unique=True,
        partialFilterExpression={"hard_dedupe_key": {"$exists": True}, "dedupe_exempt": False},
        name="uniq_expenses_hard_dedupe_key",
    )
    # Start Telegram bot scheduler (commands + payment reminders)
    from telegram_scheduler import run_telegram_scheduler
    asyncio.create_task(run_telegram_scheduler(db))
    logger.info("Obelisco Manager API iniciada")

app.include_router(api_router)
app.include_router(public_router)

# Payroll module
from payroll import create_payroll_router
app.include_router(create_payroll_router(db, get_current_user))

# Expenses / Custos module
from expenses import create_expenses_router
app.include_router(create_expenses_router(db, get_current_user))

# Invoices / Faturacao module
from invoices import create_invoices_router
app.include_router(create_invoices_router(db, get_current_user))

from proposal_import import create_proposal_import_router
app.include_router(create_proposal_import_router(get_current_user))

# Notifications
from notifications import create_notifications_router, create_notification, notify_admins
from tech_extras import _get_tech_user_dep  # já usado internamente

async def _get_admin_users():
    admins = await db.users.find({"role": "admin"}, {"password_hash": 0}).to_list(100)
    return [{"id": str(a.get("_id") or a.get("id") or ""), "name": a.get("name")} for a in admins if (a.get("_id") or a.get("id"))]

_notif_router, _notif_tech_router = create_notifications_router(db, get_current_user, _get_tech_user_dep(db))
app.include_router(_notif_router)
app.include_router(_notif_tech_router)

from employee_loans import create_loans_router
app.include_router(create_loans_router(db, get_current_user))

from fixed_costs import create_fixed_costs_router
app.include_router(create_fixed_costs_router(db, get_current_user))

# Transport Guides module (Guias de Transporte)
from transport_guides import create_transport_guides_router
app.include_router(create_transport_guides_router(db, get_current_user))

# Stock invoice import module (OCR de fatura → atualizar stock)
from stock_invoice_import import create_stock_import_router
app.include_router(create_stock_import_router(db, get_current_user))

# Break-even / Faturamento Ideal
from breakeven import create_breakeven_router
app.include_router(create_breakeven_router(db, get_current_user))

# Contabilista IA
from contabilista import create_contabilista_router
app.include_router(create_contabilista_router(db, get_current_user))

# Portal Técnico — endpoints extra (ponto, chat, agenda, perfil, fotos)
from tech_extras import create_tech_extras_router
app.include_router(create_tech_extras_router(db, get_current_user))

# Pedidos de Serviço (migrado de Obelisco-Tecnicos-main)
from service_orders import create_service_orders_router
app.include_router(create_service_orders_router(db, get_current_user))

# Perfil 360° Cliente (mini-CRM)
from client_profile import create_client_profile_router
app.include_router(create_client_profile_router(db, get_current_user))

# Push Notifications (Web Push API — smartwatches, phones, desktop)
from push_notifications import create_push_router
app.include_router(create_push_router(db, get_current_user))

# Bank Statement Analysis (Análise Bancária)
from bank_analysis import create_bank_analysis_router
app.include_router(create_bank_analysis_router(db, get_current_user))

_default_origins = [
    os.environ.get("FRONTEND_URL", "http://localhost:3000"),
    "https://tech-app-obelisco.emergent.host",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_default_origins,
    allow_origin_regex=r"https://.*\.emergentagent\.com|https://.*\.emergent\.host",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown():
    client.close()
